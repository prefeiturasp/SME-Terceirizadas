from datetime import datetime

from openpyxl import Workbook, drawing
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def montar_cabecalho_padrao(row_idx, ws):
    thin = Side(border_style='thin', color='000000')
    cabecalho_label = ['DRE', 'Lote', 'Tipo de unid.', 'Unid. educacional', 'Tipo de Turma', 'Período', 'Matriculados']
    for col, label in enumerate(cabecalho_label):
        ws.row_dimensions[row_idx].height = 25
        celula = ws.cell(row=row_idx, column=(col + 1))
        celula.font = Font(size='11', bold=True)
        celula.alignment = Alignment(horizontal='left', vertical='center')
        celula.fill = PatternFill('solid', start_color='a9d18e')
        celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        celula.value = label


def montar_faixas_etarias(ws, row_calculada, dado, faixas_etarias):
    thin = Side(border_style='thin', color='000000')
    for row in range((row_calculada + 1), (row_calculada + 12)):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        if row == (row_calculada + 1):
            celula = ws.cell(row=row, column=1)
            celula.font = Font(size='11', bold=True)
            celula.alignment = Alignment(horizontal='center', vertical='center')
            celula.fill = PatternFill('solid', start_color='e2f0d9')
            celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            celula.value = 'Faixa Etária'

            celula = ws.cell(row=row, column=6)
            celula.font = Font(size='11', bold=True)
            celula.alignment = Alignment(horizontal='center', vertical='center')
            celula.fill = PatternFill('solid', start_color='e2f0d9')
            celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            celula.value = 'Alunos Matriculados'

        elif row == (row_calculada + 11):
            celula = ws.cell(row=row, column=1)
            celula.font = Font(size='11', bold=True)
            celula.alignment = Alignment(horizontal='left', vertical='center')
            celula.fill = PatternFill('solid', start_color='d9d9d9')
            celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            celula.value = 'Total'

            celula = ws.cell(row=row, column=6)
            celula.font = Font(size='11')
            celula.alignment = Alignment(horizontal='center', vertical='center')
            celula.fill = PatternFill('solid', start_color='d9d9d9')
            celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            celula.value = dado['matriculados']

        else:
            celula = ws.cell(row=row, column=1)
            celula.fill = PatternFill('solid', start_color='ffffff')
            celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            celula = ws.cell(row=row, column=6)
            celula.fill = PatternFill('solid', start_color='ffffff')
            celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            celula.alignment = Alignment(horizontal='center', vertical='center')

    for idx, faixa in enumerate(faixas_etarias):
        row = (row_calculada + 2 + idx)
        celula = ws.cell(row=row, column=1)
        celula.value = faixa['nome']
        celula = ws.cell(row=row, column=6)
        celula.value = dado['alunos_por_faixa_etaria'][dado['periodo_escolar']][faixa['uuid']]


def build_xlsx_alunos_matriculados(dados, nome_arquivo, output):
    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    thin = Side(border_style='thin', color='000000')

    # imagem
    img = drawing.image.Image('sme_terceirizadas/relatorios/static/images/logo-sigpae.png')
    img.anchor = 'A1'
    img.width = 100
    img.height = 50

    # titulo e logo
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=7)
    celula = ws.cell(row=1, column=1)
    celula.fill = PatternFill('solid', start_color='a9d18e')
    celula.font = Font(size='11', bold=True)
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    celula.value = 'Relatório SIGPAE - Alunos Matriculados'
    ws.add_image(img)

    # impresso por
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    ws.row_dimensions[4].height = 20
    celula = ws.cell(row=4, column=1)
    celula.font = Font(size='11', bold=True, italic=True)
    celula.alignment = Alignment(horizontal='left', vertical='center')
    celula.fill = PatternFill('solid', start_color='ffffff')
    celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    celula.value = f'Solicitado por {dados["usuario"]}, impresso em {datetime.today().strftime("%d/%m/%Y")}'

    montar_cabecalho_padrao(5, ws)

    cabecalho_count = 6
    keys = ['dre', 'lote', 'tipo_unidade', 'escola', 'tipo_turma', 'periodo_escolar', 'matriculados']
    for row, dado in enumerate(dados['queryset']):
        for col, key in enumerate(keys):
            row_calculada = (row + cabecalho_count)
            celula = ws.cell(row=row_calculada, column=(col + 1))
            celula.fill = PatternFill('solid', start_color='ffffff')
            celula.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            celula.value = dado[key]
        if dado['eh_cei'] or dado['eh_cemei']:
            cabecalho_count = cabecalho_count + 12
            montar_faixas_etarias(ws, row_calculada, dado, dados['faixas_etarias'])
            if not len(dados['queryset']) == (row + 1):
                montar_cabecalho_padrao((row_calculada + 12), ws)
    wb.save(output)
    output.seek(0)
