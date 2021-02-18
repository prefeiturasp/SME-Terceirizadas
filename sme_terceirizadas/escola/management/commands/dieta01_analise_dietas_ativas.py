import openpyxl
import timeit
from datetime import date
from time import sleep
from pathlib import Path
from django.core.management.base import BaseCommand
from utility.carga_dados.helper import excel_to_list_with_openpyxl, progressbar
from sme_terceirizadas.escola.models import Escola
from sme_terceirizadas.produto.models import ProtocoloDeDietaEspecial
from sme_terceirizadas.eol_servico.utils import EOLService

DATA = date.today().isoformat().replace('-', '_')
home = str(Path.home())
dict_codigos_escolas = {}
dict_codigo_aluno_por_codigo_escola = {}


def get_codigo_eol_escola(valor):
    return valor.strip().zfill(6)


def get_codigo_eol_aluno(valor):
    return str(valor).strip().zfill(7)


def gera_dict_codigos_escolas(items_codigos_escolas):
    for item in items_codigos_escolas:
        dict_codigos_escolas[str(item['CÓDIGO UNIDADE'])] = str(item['CODIGO EOL'])


def gera_dict_codigo_aluno_por_codigo_escola(items):
    for item in items:
        try:
            codigo_eol_escola = dict_codigos_escolas[item['CodEscola']]
        except Exception as e:
            # Grava os CodEscola não existentes em unidades_da_rede_28.01_.xlsx
            with open(f'{home}/codescola_nao_existentes.txt', 'a') as f:
                f.write(f"{item['CodEscola']}\n")

        cod_eol_aluno = get_codigo_eol_aluno(item['CodEOLAluno'])
        dict_codigo_aluno_por_codigo_escola[cod_eol_aluno] = get_codigo_eol_escola(codigo_eol_escola)


def get_escolas_unicas(items):
    # A partir da planilha, pegar todas as escolas únicas "escolas_da_planilha"
    escolas = []
    for item in items:
        escolas.append(item['CodEscola'])
    return set(escolas)


def escreve_xlsx(codigo_eol_escola_nao_existentes, arquivo_saida):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Código EOL das Escolas não identificadas no SIGPAE')
    ws['A1'] = 'codigo_eol_escola'
    for i, item in enumerate(progressbar(list(codigo_eol_escola_nao_existentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_primeira_aba(arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.worksheets[0]
    ws['A1'] = 'Este arquivo contém as planilhas:'
    ws['A2'] = 'Código EOL das Escolas não identificadas no SIGPAE'
    ws['A3'] = 'Código EOL dos Alunos não matriculados na escola'
    ws['A4'] = 'CodEscola não existentes em unidades_da_rede...'
    ws['A5'] = 'Dados do SIGPAE para as escolas da planilha'
    ws['A6'] = 'CodDiagnostico inexistentes'
    ws['A7'] = 'ProtocoloDieta inexistentes'
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_alunos_nao_matriculados_na_escola(alunos_nao_matriculados_na_escola_lista, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('Código EOL dos Alunos não matriculados na escola')
    ws['A1'] = 'codigo_eol_aluno'
    ws['B1'] = 'nome_aluno'
    ws['C1'] = 'codigo_eol_escola'
    for i, item in enumerate(progressbar(list(alunos_nao_matriculados_na_escola_lista), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item[0])
        ws[f'B{i+2}'] = str(item[1])
        ws[f'C{i+2}'] = str(item[2])
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_codescola_nao_existentes(codescola_nao_existentes, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('CodEscola não existentes em unidades_da_rede...')
    ws['A1'] = 'CodEscola'
    for i, item in enumerate(progressbar(list(codescola_nao_existentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def retorna_codigo_eol_escolas_nao_identificadas(items, arquivo_saida):
    aux = []
    codescola_nao_existentes = []
    for item in items:
        try:
            aux.append(get_codigo_eol_escola(dict_codigos_escolas[item['CodEscola']]))
        except Exception as e:
            # Grava os CodEscola não existentes em unidades_da_rede_28.01_.xlsx
            print('CodEscola não existentes', item['CodEscola'])
            codescola_nao_existentes.append(item['CodEscola'])

    cod_escola_unicos = set(aux)
    codigo_eol_sigpae_lista = Escola.objects.values_list('codigo_eol', flat=True)
    codigo_eol_escola_nao_existentes = cod_escola_unicos - set(codigo_eol_sigpae_lista)

    escreve_xlsx(codigo_eol_escola_nao_existentes, arquivo_saida)

    if set(codescola_nao_existentes):
        escreve_xlsx_codescola_nao_existentes(set(codescola_nao_existentes), arquivo_saida)


def retorna_alunos_nao_matriculados_na_escola(escolas_da_planilha, items, arquivo_saida):
    # Percorrendo "escolas_da_planilha", a partir da planilha pegar todos os alunos de cada escola do iterador.
    tic = timeit.default_timer()

    alunos_nao_matriculados_na_escola_lista = []
    for i, escola in enumerate(list(escolas_da_planilha)[:5]):
        print(i, 'Lendo API do EOL...')
        if i % 4 == 0:
            sleep(2)
        # Pegar os alunos de cada escola do iterador.
        # Retorna uma tupla com CodEOLAluno e NomeAluno
        alunos_da_planilha_por_escola_lista = [
            (get_codigo_eol_aluno(item['CodEOLAluno']), item['NomeAluno'])
            for item in items if item['CodEscola'] == escola
        ]
        print('Alunos', i, escola)
        print(i, 'alunos_da_planilha_por_escola_lista', alunos_da_planilha_por_escola_lista)
        # Pegar todos os alunos por escola na API usando a função get_informacoes_escola_turma_aluno
        codigo_eol_escola = get_codigo_eol_escola(dict_codigos_escolas[escola])
        print('codigo_eol_escola', codigo_eol_escola)
        alunos_da_api_por_escola_lista = EOLService.get_informacoes_escola_turma_aluno(codigo_eol_escola)
        for cod_aluno, nome_aluno in alunos_da_planilha_por_escola_lista:
            pertence = any(
                get_codigo_eol_aluno(aluno['cd_aluno']) == cod_aluno
                for aluno in alunos_da_api_por_escola_lista
            )
            if not pertence:
                print(f'>>> Aluno {cod_aluno} - {nome_aluno} não pertence a escola {codigo_eol_escola}')
                alunos_nao_matriculados_na_escola_lista.append((cod_aluno, nome_aluno, codigo_eol_escola))
                print()

    if alunos_nao_matriculados_na_escola_lista:
        escreve_xlsx_alunos_nao_matriculados_na_escola(alunos_nao_matriculados_na_escola_lista, arquivo_saida)

    toc = timeit.default_timer()
    print('Time', round(toc - tic, 2))


def escreve_xlsx_dados_sigpae(items, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('Dados do SIGPAE para as escolas da planilha')
    ws['A1'] = 'codigo_eol_escola'
    ws['B1'] = 'nome_da_escola'
    ws['C1'] = 'nome_dre'
    ws['D1'] = 'lote'
    ws['E1'] = 'tipo_gestao'
    ws['F1'] = 'contato_email'
    ws['G1'] = 'contato_telefone'
    ws['H1'] = 'contato_telefone2'
    ws['I1'] = 'contato_celular'
    i = 0  # indice criado manualmente pra não inserir linhas em branco na planilha.
    for item in progressbar(items, 'Escrevendo...'):
        escola = Escola.objects.filter(codigo_eol=item).first()
        if escola:
            ws[f'A{i+2}'] = escola.codigo_eol
            ws[f'B{i+2}'] = escola.nome
            if escola.diretoria_regional:
                ws[f'C{i+2}'] = escola.diretoria_regional.nome
            if escola.lote:
                ws[f'D{i+2}'] = escola.lote.nome
            if escola.tipo_gestao:
                ws[f'E{i+2}'] = escola.tipo_gestao.nome
            if escola.contato:
                ws[f'F{i+2}'] = escola.contato.email
                ws[f'G{i+2}'] = escola.contato.telefone
                ws[f'H{i+2}'] = escola.contato.telefone2
                ws[f'I{i+2}'] = escola.contato.celular
            i += 1
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def retorna_dados_sigpae(items, arquivo_saida):
    aux = []
    for item in items:
        try:
            aux.append(get_codigo_eol_escola(dict_codigos_escolas[item['CodEscola']]))
        except Exception as e:
            pass

    cod_escola_unicos = set(aux)
    escreve_xlsx_dados_sigpae(list(cod_escola_unicos), arquivo_saida)


def escreve_xlsx_cod_diagnostico_inexistentes(cod_diagnostico_inexistentes, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('CodDiagnostico inexistentes')
    ws['A1'] = 'cod_diagnostico'
    for i, item in enumerate(progressbar(list(cod_diagnostico_inexistentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_protocolo_dieta_inexistentes(protocolo_dieta_inexistentes, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('ProtocoloDieta inexistentes')
    ws['A1'] = 'protocolo_dieta'
    for i, item in enumerate(progressbar(list(protocolo_dieta_inexistentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def retorna_cod_diagnostico_inexistentes(items, arquivo_saida):
    cod_diagnostico_unicos = [item['CodDiagnostico'] for item in items]
    protocolo_de_dieta_especial = ProtocoloDeDietaEspecial.objects.values_list('nome', flat=True)
    cod_diagnostico_inexistentes = set(cod_diagnostico_unicos) - set(protocolo_de_dieta_especial)
    if cod_diagnostico_inexistentes:
        escreve_xlsx_cod_diagnostico_inexistentes(cod_diagnostico_inexistentes, arquivo_saida)


def retorna_protocolo_dieta_inexistentes(items, arquivo_saida):
    protocolo_dieta_unicos = [item['ProtocoloDieta'] for item in items]
    protocolo_de_dieta_especial = ProtocoloDeDietaEspecial.objects.values_list('nome', flat=True)
    protocolo_dieta_inexistentes = set(protocolo_dieta_unicos) - set(protocolo_de_dieta_especial)
    if protocolo_dieta_inexistentes:
        escreve_xlsx_protocolo_dieta_inexistentes(protocolo_dieta_inexistentes, arquivo_saida)


class Command(BaseCommand):
    help = """
    Lê uma planilha específica com Dietas Ativas a serem integradas no sistema.
    Valida se o código EOL da unidade educacional da planilha Excel existe na base do SIGPAE.
    """

    def add_arguments(self, parser):
        parser.add_argument(
            '--arquivo', '-a',
            dest='arquivo',
            help='Informar caminho absoluto do arquivo xlsx.'
        )

    def handle(self, *args, **options):
        arquivo = options['arquivo']
        arquivo_codigos_escolas = f'{home}/Downloads/unidades_da_rede_28.01_.xlsx'
        arquivo_saida = f'{home}/Downloads/resultado_analise_dietas_ativas.xlsx'
        items = excel_to_list_with_openpyxl(arquivo, in_memory=False)

        # 1
        items_codigos_escolas = excel_to_list_with_openpyxl(arquivo_codigos_escolas, in_memory=False)
        gera_dict_codigos_escolas(items_codigos_escolas)
        retorna_codigo_eol_escolas_nao_identificadas(items, arquivo_saida)

        # 2
        # items_codigos_escolas = excel_to_list_with_openpyxl(arquivo_codigos_escolas, in_memory=False)

        # gera_dict_codigos_escolas(items_codigos_escolas)
        # gera_dict_codigo_aluno_por_codigo_escola(items)

        # # A partir da planilha, pegar todas as escolas únicas "escolas_da_planilha"
        # escolas_da_planilha = get_escolas_unicas(items)

        # # Percorrendo "escolas_da_planilha", a partir da planilha pegar todos os alunos de cada escola do iterador.
        # retorna_alunos_nao_matriculados_na_escola(escolas_da_planilha, items, arquivo_saida)

        # 5
        # Usa items_codigos_escolas
        # Usa gera_dict_codigos_escolas
        retorna_dados_sigpae(items, arquivo_saida)

        # 6 Verificar os campos "CodDiagnostico" e "ProtocoloDieta"
        retorna_cod_diagnostico_inexistentes(items, arquivo_saida)
        retorna_protocolo_dieta_inexistentes(items, arquivo_saida)

        escreve_xlsx_primeira_aba(arquivo_saida)
