import json
import openpyxl
from datetime import date, datetime
from pathlib import Path
from django.core.management.base import BaseCommand
from sme_terceirizadas.escola.models import Escola
from sme_terceirizadas.produto.models import ProtocoloDeDietaEspecial
from utility.carga_dados.helper import excel_to_list_with_openpyxl, progressbar


DATA = date.today().isoformat().replace('-', '_')
home = str(Path.home())
dict_codigos_escolas = {}
dict_codigo_aluno_por_codigo_escola = {}


def get_codigo_eol_escola(valor):
    return valor.strip().zfill(6)


def get_codigo_eol_aluno(valor):
    return str(valor).strip().zfill(7)


def gera_dict_codigos_escolas(items_codigos_escolas):
    for item in items_codigos_escolas:
        dict_codigos_escolas[str(item['CÓDIGO UNIDADE'])] = str(item['CODIGO EOL'])


def gera_dict_codigo_aluno_por_codigo_escola(items):
    for item in items:
        try:
            codigo_eol_escola = dict_codigos_escolas[item['CodEscola']]
        except Exception as e:
            # Grava os CodEscola não existentes em unidades_da_rede_28.01_.xlsx
            with open(f'{home}/codescola_nao_existentes.txt', 'a') as f:
                f.write(f"{item['CodEscola']}\n")

        cod_eol_aluno = get_codigo_eol_aluno(item['CodEOLAluno'])
        # chave: cod_eol_aluno, valor: codigo_eol_escola
        dict_codigo_aluno_por_codigo_escola[cod_eol_aluno] = get_codigo_eol_escola(codigo_eol_escola)


def get_escolas_unicas(items):
    """
    A partir da planilha, pegar todas as escolas únicas "escolas_da_planilha"
    Retorna escolas únicas.
    """
    escolas = []
    for item in items:
        escolas.append(item['CodEscola'])
    return set(escolas)


def escreve_xlsx(codigo_eol_escola_nao_existentes, arquivo_saida):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Código EOL das Escolas não identificadas no SIGPAE')
    ws['A1'] = 'codigo_eol_escola'
    for i, item in enumerate(progressbar(list(codigo_eol_escola_nao_existentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_primeira_aba(arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.worksheets[0]
    ws['A1'] = 'Este arquivo contém as planilhas:'
    ws['A2'] = 'Código EOL das Escolas não identificadas no SIGPAE'
    ws['A3'] = 'Código EOL dos Alunos não matriculados na escola'
    ws['A4'] = 'CodEscola não existentes em unidades_da_rede...'
    ws['A5'] = 'Alunos com nome diferente do EOL'
    ws['A6'] = 'Alunos com data nascimento diferente do EOL'
    ws['A7'] = 'Dados do SIGPAE para as escolas da planilha'
    ws['A8'] = 'CodDiagnostico inexistentes'
    ws['A9'] = 'ProtocoloDieta inexistentes'
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_alunos_nao_matriculados_na_escola(alunos_nao_matriculados_na_escola_lista, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('Código EOL dos Alunos não matriculados na escola')
    ws['A1'] = 'codigo_eol_aluno'
    ws['B1'] = 'nome_aluno'
    ws['C1'] = 'codigo_eol_escola'
    for i, item in enumerate(progressbar(list(alunos_nao_matriculados_na_escola_lista), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item[0])
        ws[f'B{i+2}'] = str(item[1])
        ws[f'C{i+2}'] = str(item[2])
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_codescola_nao_existentes(codescola_nao_existentes, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('CodEscola não existentes em unidades_da_rede...')
    ws['A1'] = 'CodEscola'
    for i, item in enumerate(progressbar(list(codescola_nao_existentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def retorna_codigo_eol_escolas_nao_identificadas(items, arquivo_saida):
    aux = []
    codescola_nao_existentes = []
    for item in items:
        try:
            aux.append(get_codigo_eol_escola(dict_codigos_escolas[item['CodEscola']]))
        except Exception as e:
            # Grava os CodEscola não existentes em unidades_da_rede_28.01_.xlsx
            print('CodEscola não existentes', item['CodEscola'])
            codescola_nao_existentes.append(item['CodEscola'])

    cod_escola_unicos = set(aux)
    codigo_eol_sigpae_lista = Escola.objects.values_list('codigo_eol', flat=True)
    codigo_eol_escola_nao_existentes = cod_escola_unicos - set(codigo_eol_sigpae_lista)

    escreve_xlsx(codigo_eol_escola_nao_existentes, arquivo_saida)

    if set(codescola_nao_existentes):
        escreve_xlsx_codescola_nao_existentes(set(codescola_nao_existentes), arquivo_saida)


def get_escolas(arquivo):
    # Lê os dados de 'escolas.json' e retorna um json.
    with open(arquivo, 'r') as f:
        data = json.load(f)
    return data


def retorna_alunos_nao_matriculados_na_escola(items, escolas, arquivo_saida):
    data = escolas
    alunos_nao_matriculados_na_escola_lista = []

    for aluno in items:
        escola = dict_codigo_aluno_por_codigo_escola[str(aluno['CodEOLAluno'])]
        if data.get(escola):
            pertence = any(
                get_codigo_eol_aluno(_aluno['cd_aluno']) == str(aluno['CodEOLAluno'])
                for _aluno in data.get(escola)
            )
            print('aluno:', aluno['CodEOLAluno'], 'escola:', escola, pertence)
        else:
            pertence = False

        if not pertence:
            tupla = (
                str(aluno['CodEOLAluno']),
                aluno['NomeAluno'],
                escola
            )
            alunos_nao_matriculados_na_escola_lista.append(tupla)

    if alunos_nao_matriculados_na_escola_lista:
        # Criar lista com o nome e cod_eol_aluno não matriculado na escola informada.
        escreve_xlsx_alunos_nao_matriculados_na_escola(alunos_nao_matriculados_na_escola_lista, arquivo_saida)


def escreve_xlsx_dados_sigpae(items, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('Dados do SIGPAE para as escolas da planilha')
    ws['A1'] = 'codigo_eol_escola'
    ws['B1'] = 'nome_da_escola'
    ws['C1'] = 'nome_dre'
    ws['D1'] = 'lote'
    ws['E1'] = 'tipo_gestao'
    ws['F1'] = 'contato_email'
    ws['G1'] = 'contato_telefone'
    ws['H1'] = 'contato_telefone2'
    ws['I1'] = 'contato_celular'
    i = 0  # indice criado manualmente pra não inserir linhas em branco na planilha.
    for item in progressbar(items, 'Escrevendo...'):
        escola = Escola.objects.filter(codigo_eol=item).first()
        if escola:
            ws[f'A{i+2}'] = escola.codigo_eol
            ws[f'B{i+2}'] = escola.nome
            if escola.diretoria_regional:
                ws[f'C{i+2}'] = escola.diretoria_regional.nome
            if escola.lote:
                ws[f'D{i+2}'] = escola.lote.nome
            if escola.tipo_gestao:
                ws[f'E{i+2}'] = escola.tipo_gestao.nome
            if escola.contato:
                ws[f'F{i+2}'] = escola.contato.email
                ws[f'G{i+2}'] = escola.contato.telefone
                ws[f'H{i+2}'] = escola.contato.telefone2
                ws[f'I{i+2}'] = escola.contato.celular
            i += 1
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def retorna_dados_sigpae(items, arquivo_saida):
    aux = []
    for item in items:
        try:
            aux.append(get_codigo_eol_escola(dict_codigos_escolas[item['CodEscola']]))
        except Exception as e:
            pass

    cod_escola_unicos = set(aux)
    escreve_xlsx_dados_sigpae(list(cod_escola_unicos), arquivo_saida)


def escreve_xlsx_cod_diagnostico_inexistentes(cod_diagnostico_inexistentes, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('CodDiagnostico inexistentes')
    ws['A1'] = 'cod_diagnostico'
    for i, item in enumerate(progressbar(list(cod_diagnostico_inexistentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_protocolo_dieta_inexistentes(protocolo_dieta_inexistentes, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('ProtocoloDieta inexistentes')
    ws['A1'] = 'protocolo_dieta'
    for i, item in enumerate(progressbar(list(protocolo_dieta_inexistentes), 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item)
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def retorna_cod_diagnostico_inexistentes(items, arquivo_saida):
    cod_diagnostico_unicos = [item['CodDiagnostico'] for item in items]
    protocolo_de_dieta_especial = ProtocoloDeDietaEspecial.objects.values_list('nome', flat=True)
    cod_diagnostico_inexistentes = set(cod_diagnostico_unicos) - set(protocolo_de_dieta_especial)
    if cod_diagnostico_inexistentes:
        escreve_xlsx_cod_diagnostico_inexistentes(cod_diagnostico_inexistentes, arquivo_saida)


def retorna_protocolo_dieta_inexistentes(items, arquivo_saida):
    protocolo_dieta_unicos = [item['ProtocoloDieta'] for item in items]
    protocolo_de_dieta_especial = ProtocoloDeDietaEspecial.objects.values_list('nome', flat=True)
    protocolo_dieta_inexistentes = set(protocolo_dieta_unicos) - set(protocolo_de_dieta_especial)
    if protocolo_dieta_inexistentes:
        escreve_xlsx_protocolo_dieta_inexistentes(protocolo_dieta_inexistentes, arquivo_saida)


def string_to_date(texto, formato):
    # Transforma string em data.
    return datetime.strptime(texto, formato).date()


def escreve_xlsx_alunos_com_nome_diferente(lista, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('Alunos com nome diferente do EOL')
    ws['A1'] = 'nome_aluno_planilha'
    ws['B1'] = 'nome_aluno_eol'
    for i, item in enumerate(progressbar(lista, 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item[0])
        ws[f'B{i+2}'] = str(item[1])
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def escreve_xlsx_alunos_com_nascimento_diferente(lista, arquivo_saida):
    nome_ = arquivo_saida.split('.')
    nome = f'{nome_[0]}_{DATA}.{nome_[1]}'
    wb = openpyxl.load_workbook(nome)
    ws = wb.create_sheet('Alunos com nascimento diferente do EOL')
    ws['A1'] = 'nascimento_planilha'
    ws['B1'] = 'nascimento_eol'
    for i, item in enumerate(progressbar(lista, 'Escrevendo...')):
        ws[f'A{i+2}'] = str(item[0])
        ws[f'B{i+2}'] = str(item[1])
    nome = arquivo_saida.split('.')
    wb.save(f'{nome[0]}_{DATA}.{nome[1]}')


def retorna_alunos_com_nome_diferente(items, escolas, arquivo_saida):
    aux = []
    for aluno in items:
        escola = dict_codigo_aluno_por_codigo_escola[str(aluno['CodEOLAluno'])]
        if escolas.get(escola):
            # print(aluno['CodEOLAluno'], 'escola:', escola)
            # print(aluno['CodEOLAluno'], aluno['DataNascimento'], aluno['NomeAluno'])
            aluno_localizado = list(filter(lambda x: x['cd_aluno'] == aluno['CodEOLAluno'], escolas.get(escola)))
            if aluno_localizado:
                nome_aluno_planilha = aluno['NomeAluno']
                nome_aluno_eol = aluno_localizado[0]['nm_aluno']
                # print(aluno_localizado[0]['nm_aluno'])
                # print(nome_aluno_planilha == nome_aluno_eol)
                if nome_aluno_planilha != nome_aluno_eol:
                    tupla = (nome_aluno_planilha, nome_aluno_eol)
                    print(nome_aluno_planilha)
                    print(nome_aluno_eol)
                    aux.append(tupla)
    if aux:
        escreve_xlsx_alunos_com_nome_diferente(aux, arquivo_saida)
    else:
        print('Todos alunos estão com o nome OK.')


def retorna_alunos_com_nascimento_diferente(items, escolas, arquivo_saida):
    aux = []
    for aluno in items:
        escola = dict_codigo_aluno_por_codigo_escola[str(aluno['CodEOLAluno'])]
        if escolas.get(escola):
            aluno_localizado = list(filter(lambda x: x['cd_aluno'] == aluno['CodEOLAluno'], escolas.get(escola)))
            if aluno_localizado:
                nascimento_planilha = string_to_date(aluno['DataNascimento'], '%d/%m/%Y')
                # print(nascimento_planilha)
                nascimento_eol = string_to_date(aluno_localizado[0]['dt_nascimento_aluno'], '%Y-%m-%dT%H:%M:%S')
                # print(nascimento_eol)
                # print(nascimento_planilha == nascimento_eol)
                if nascimento_planilha != nascimento_eol:
                    tupla = (nascimento_planilha, nascimento_eol)
                    print(nascimento_planilha)
                    print(nascimento_eol)
                    aux.append(tupla)
    if aux:
        escreve_xlsx_alunos_com_nascimento_diferente(aux, arquivo_saida)
    else:
        print('Todos alunos estão com a data de nascimento OK.')


class Command(BaseCommand):
    help = """
    Lê uma planilha específica com Dietas Ativas a serem integradas no sistema.
    Valida se o código EOL da unidade educacional da planilha Excel existe na base do SIGPAE.
    """

    def add_arguments(self, parser):
        parser.add_argument(
            '--arquivo', '-a',
            dest='arquivo',
            help='Informar caminho absoluto do arquivo xlsx.'
        )

    def handle(self, *args, **options):
        arquivo = options['arquivo']
        arquivo_codigos_escolas = f'{home}/Downloads/unidades_da_rede_28.01_.xlsx'
        arquivo_saida = f'{home}/Downloads/resultado_analise_dietas_ativas.xlsx'
        items = excel_to_list_with_openpyxl(arquivo, in_memory=False)
        escolas = get_escolas(f'{home}/escolas.json')

        # 1
        items_codigos_escolas = excel_to_list_with_openpyxl(arquivo_codigos_escolas, in_memory=False)
        gera_dict_codigos_escolas(items_codigos_escolas)
        retorna_codigo_eol_escolas_nao_identificadas(items, arquivo_saida)

        # 2
        # Usa items_codigos_escolas
        # Usa gera_dict_codigos_escolas
        gera_dict_codigo_aluno_por_codigo_escola(items)
        retorna_alunos_nao_matriculados_na_escola(items, escolas, arquivo_saida)

        # 3 - Retorna nome e data de nascimento que forem diferentes entre a planilha e o EOL.
        retorna_alunos_com_nome_diferente(items, escolas, arquivo_saida)
        retorna_alunos_com_nascimento_diferente(items, escolas, arquivo_saida)

        # ---

        # 5
        # Usa items_codigos_escolas
        # Usa gera_dict_codigos_escolas
        retorna_dados_sigpae(items, arquivo_saida)

        # 6 Verificar os campos "CodDiagnostico" e "ProtocoloDieta"
        retorna_cod_diagnostico_inexistentes(items, arquivo_saida)
        retorna_protocolo_dieta_inexistentes(items, arquivo_saida)

        escreve_xlsx_primeira_aba(arquivo_saida)
