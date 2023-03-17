import datetime

from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django_filters import rest_framework as filters
from openpyxl import Workbook, styles
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import permissions, serializers, status
from rest_framework.decorators import action
from rest_framework.mixins import CreateModelMixin, ListModelMixin, RetrieveModelMixin, UpdateModelMixin
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet, ModelViewSet, ReadOnlyModelViewSet

from ...dados_comuns.constants import (
    ADMINISTRADOR_DIETA_ESPECIAL,
    ADMINISTRADOR_GESTAO_ALIMENTACAO_TERCEIRIZADA,
    ADMINISTRADOR_GESTAO_PRODUTO,
    ADMINISTRADOR_SUPERVISAO_NUTRICAO,
    ADMINISTRADOR_UE,
    COGESTOR_DRE
)
from ...dados_comuns.permissions import UsuarioDiretoriaRegional, UsuarioEscolaTercTotal
from ...dados_comuns.utils import get_ultimo_dia_mes
from ...eol_servico.utils import EOLException
from ...escola.api.permissions import (
    PodeCriarAdministradoresDaCODAEGestaoAlimentacaoTerceirizada,
    PodeCriarAdministradoresDaCODAEGestaoDietaEspecial,
    PodeCriarAdministradoresDaCODAEGestaoProdutos,
    PodeCriarAdministradoresDaCODAESupervisaoNutricao,
    PodeCriarAdministradoresDaDiretoriaRegional,
    PodeCriarAdministradoresDaEscola
)
from ...escola.api.serializers import (
    AlunoSerializer,
    AlunoSimplesSerializer,
    CODAESerializer,
    DiretoriaRegionalParaFiltroSerializer,
    EscolaParaFiltroSerializer,
    EscolaPeriodoEscolarSerializer,
    LoteNomeSerializer,
    LoteParaFiltroSerializer,
    LoteSerializer,
    LoteSimplesSerializer,
    TipoUnidadeParaFiltroSerializer,
    UsuarioDetalheSerializer
)
from ...escola.api.serializers_create import (
    EscolaPeriodoEscolarCreateSerializer,
    EscolaSimplesUpdateSerializer,
    FaixaEtariaSerializer,
    LoteCreateSerializer,
    MudancaFaixasEtariasCreateSerializer
)
from ...inclusao_alimentacao.models import InclusaoAlimentacaoContinua
from ...paineis_consolidados.api.constants import FILTRO_DRE_UUID
from ...perfil.api.serializers import UsuarioUpdateSerializer, VinculoSerializer
from ..forms import AlunosPorFaixaEtariaForm
from ..models import (
    Aluno,
    Codae,
    DiaCalendario,
    DiretoriaRegional,
    Escola,
    EscolaPeriodoEscolar,
    FaixaEtaria,
    LogAlteracaoQuantidadeAlunosPorEscolaEPeriodoEscolar,
    LogAlunosMatriculadosPeriodoEscola,
    Lote,
    PeriodoEscolar,
    Subprefeitura,
    TipoGestao,
    TipoTurma,
    TipoUnidadeEscolar
)
from ..services import NovoSGPServicoLogado, NovoSGPServicoLogadoException
from ..utils import EscolaSimplissimaPagination
from .filters import AlunoFilter, DiretoriaRegionalFilter
from .permissions import PodeVerEditarFotoAlunoNoSGP
from .serializers import (
    AlunosMatriculadosPeriodoEscolaSerializer,
    DiaCalendarioSerializer,
    DiretoriaRegionalCompletaSerializer,
    DiretoriaRegionalLookUpSerializer,
    DiretoriaRegionalSimplissimaSerializer,
    EscolaListagemSimplissimaComDRESelializer,
    EscolaSimplesSerializer,
    EscolaSimplissimaSerializer,
    LogAlunosMatriculadosPeriodoEscolaSerializer,
    PeriodoEFaixaEtariaCounterSerializer,
    PeriodoEscolarSerializer,
    SubprefeituraSerializer,
    SubprefeituraSerializerSimples,
    TipoGestaoSerializer,
    TipoUnidadeEscolarSerializer
)


# https://www.django-rest-framework.org/api-guide/permissions/#custom-permissions
class VinculoViewSet(ReadOnlyModelViewSet):
    lookup_field = 'uuid'
    serializer_class = VinculoSerializer

    @action(detail=True, methods=['post'])
    def criar_equipe_administradora(self, request, uuid=None):
        try:
            instituicao = self.get_object()
            data = request.data.copy()
            data['instituicao'] = instituicao.nome
            usuario = UsuarioUpdateSerializer(data).create(validated_data=data)
            usuario.criar_vinculo_administrador(instituicao, nome_perfil=self.nome_perfil)  # noqa
            return Response(UsuarioDetalheSerializer(usuario).data)
        except serializers.ValidationError as e:
            return Response(data=dict(detail=e.args[0]), status=e.status_code)

    @action(detail=True)
    def get_equipe_administradora(self, request, uuid=None):
        instituicao = self.get_object()
        vinculos = instituicao.vinculos_que_podem_ser_finalizados
        return Response(self.get_serializer(vinculos, many=True).data)

    @action(detail=True, methods=['patch'])
    def finalizar_vinculo(self, request, uuid=None):
        instituicao = self.get_object()
        vinculo_uuid = request.data.get('vinculo_uuid')
        vinculo = instituicao.vinculos.get(uuid=vinculo_uuid)
        vinculo.finalizar_vinculo()
        return Response(self.get_serializer(vinculo).data)


class VinculoEscolaViewSet(VinculoViewSet):
    queryset = Escola.objects.all()
    permission_classes = [PodeCriarAdministradoresDaEscola]
    nome_perfil = ADMINISTRADOR_UE


class VinculoDiretoriaRegionalViewSet(VinculoViewSet):
    queryset = DiretoriaRegional.objects.all()
    permission_classes = [PodeCriarAdministradoresDaDiretoriaRegional]
    nome_perfil = COGESTOR_DRE


class VinculoCODAEGestaoAlimentacaoTerceirizadaViewSet(VinculoViewSet):
    queryset = Codae.objects.all()
    permission_classes = [PodeCriarAdministradoresDaCODAEGestaoAlimentacaoTerceirizada]  # noqa
    nome_perfil = ADMINISTRADOR_GESTAO_ALIMENTACAO_TERCEIRIZADA


class VinculoCODAEGestaoDietaEspecialViewSet(VinculoViewSet):
    queryset = Codae.objects.all()
    permission_classes = [PodeCriarAdministradoresDaCODAEGestaoDietaEspecial]
    nome_perfil = ADMINISTRADOR_DIETA_ESPECIAL


class VinculoCODAEGestaoProdutosViewSet(VinculoViewSet):
    queryset = Codae.objects.all()
    permission_classes = [PodeCriarAdministradoresDaCODAEGestaoProdutos]
    nome_perfil = ADMINISTRADOR_GESTAO_PRODUTO


class VinculoCODAESupervisaoNutricaoViewSet(VinculoViewSet):
    queryset = Codae.objects.all()
    permission_classes = [PodeCriarAdministradoresDaCODAESupervisaoNutricao]
    nome_perfil = ADMINISTRADOR_SUPERVISAO_NUTRICAO


class EscolaSimplesViewSet(ListModelMixin, RetrieveModelMixin, UpdateModelMixin, GenericViewSet):
    lookup_field = 'uuid'
    queryset = Escola.objects.all()

    def get_serializer_class(self):
        if self.action in ['update', 'partial_update']:
            return EscolaSimplesUpdateSerializer
        return EscolaSimplesSerializer


class EscolaSimplissimaViewSet(ListModelMixin, RetrieveModelMixin, GenericViewSet):
    lookup_field = 'uuid'
    queryset = Escola.objects.all()
    serializer_class = EscolaSimplissimaSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    pagination_class = EscolaSimplissimaPagination
    filterset_fields = ['codigo_eol', 'nome', 'diretoria_regional__uuid']

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        if 'page' in request.query_params:
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response({'results': serializer.data})

    @action(detail=False, methods=['GET'], url_path=f'{FILTRO_DRE_UUID}')
    def filtro_por_diretoria_regional(self, request, dre_uuid=None):
        escolas = Escola.objects.filter(diretoria_regional__uuid=dre_uuid)
        return Response(self.get_serializer(escolas, many=True).data)


class EscolaSimplissimaComDREViewSet(ReadOnlyModelViewSet):
    lookup_field = 'uuid'
    queryset = Escola.objects.all().prefetch_related('diretoria_regional')
    serializer_class = EscolaListagemSimplissimaComDRESelializer


class EscolaSimplissimaComDREUnpaginatedViewSet(EscolaSimplissimaComDREViewSet):
    pagination_class = None
    filterset_class = DiretoriaRegionalFilter

    @action(detail=False, methods=['GET'], url_path='terc-total')
    def terc_total(self, request):
        escolas = self.get_queryset().filter(tipo_gestao__nome='TERC TOTAL')
        dre = request.query_params.get('dre', None)
        if dre:
            escolas = escolas.filter(diretoria_regional__uuid=dre)
        return Response(self.get_serializer(escolas, many=True).data)


class EscolaQuantidadeAlunosPorPeriodoEFaixaViewSet(GenericViewSet):
    lookup_field = 'uuid'
    queryset = Escola.objects.all()
    serializer_class = PeriodoEFaixaEtariaCounterSerializer

    @action(detail=True, url_path='(?P<data_referencia_str>[^/.]+)')  # noqa C901
    def alunos_por_faixa_etaria(self, request, uuid, data_referencia_str):
        form = AlunosPorFaixaEtariaForm({
            'data_referencia': data_referencia_str
        })

        if not form.is_valid():
            return Response(form.errors)

        escola = self.get_object()
        data_referencia = form.cleaned_data['data_referencia']

        counter_faixas_etarias = escola.alunos_por_periodo_e_faixa_etaria(data_referencia)  # noqa
        serializer = PeriodoEFaixaEtariaCounterSerializer(counter_faixas_etarias)  # noqa
        return Response(serializer.data)


class PeriodoEscolarViewSet(ReadOnlyModelViewSet):
    lookup_field = 'uuid'
    queryset = PeriodoEscolar.objects.all()
    serializer_class = PeriodoEscolarSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ('nome',)

    #  TODO: Quebrar esse método um pouco, está complexo e sem teste
    @action(detail=True, url_path='alunos-por-faixa-etaria/(?P<data_referencia_str>[^/.]+)')
    def alunos_por_faixa_etaria(self, request, uuid, data_referencia_str):  # noqa C901
        form = AlunosPorFaixaEtariaForm({
            'data_referencia': data_referencia_str
        })

        if not form.is_valid():
            return Response(form.errors)

        periodo_escolar = self.get_object()
        if periodo_escolar.nome == 'PARCIAL':
            periodo_escolar = PeriodoEscolar.objects.get(nome='INTEGRAL')
        escola = self.request.user.vinculos.get(ativo=True).instituicao
        escola_periodo, created = EscolaPeriodoEscolar.objects.get_or_create(
            escola=escola,
            periodo_escolar=periodo_escolar
        )
        data_referencia = form.cleaned_data['data_referencia']

        try:
            faixa_alunos = escola_periodo.alunos_por_faixa_etaria(data_referencia)  # noqa
        except ObjectDoesNotExist:
            return Response(
                {'detail': 'Não há faixas etárias cadastradas. Contate a coordenadoria CODAE.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except EOLException:
            return Response(
                {'detail': 'API EOL indisponível para carregar as faixas etárias. Tente novamente mais tarde'},
                status=status.HTTP_400_BAD_REQUEST
            )

        results = []
        for uuid_faixa_etaria in faixa_alunos:
            results.append({
                'faixa_etaria': FaixaEtariaSerializer(FaixaEtaria.objects.get(uuid=uuid_faixa_etaria)).data,
                'count': faixa_alunos[uuid_faixa_etaria]
            })
        results = sorted(results, key=lambda x: x['faixa_etaria']['inicio'])

        return Response({
            'count': len(results),
            'results': results
        })

    @action(detail=False, methods=['GET'], url_path='inclusao-continua-por-mes',
            permission_classes=[UsuarioEscolaTercTotal | UsuarioDiretoriaRegional])
    def inclusao_continua_por_mes(self, request):
        try:
            for param in ['mes', 'ano']:
                if param not in request.query_params:
                    raise ValidationError(f'{param} é obrigatório via query_params')
            mes = request.query_params.get('mes')
            ano = request.query_params.get('ano')
            escola = request.query_params.get('escola')
            primeiro_dia_mes = datetime.date(int(ano), int(mes), 1)
            ultimo_dia_mes = get_ultimo_dia_mes(primeiro_dia_mes)
            instituicao = request.user.vinculo_atual.instituicao
            if(isinstance(instituicao, DiretoriaRegional) and escola):
                instituicao = Escola.objects.get(uuid=escola)
            periodos = dict(InclusaoAlimentacaoContinua.objects.filter(
                status='CODAE_AUTORIZADO',
                rastro_escola=instituicao).filter(
                    data_inicial__lte=ultimo_dia_mes,
                    data_final__gte=primeiro_dia_mes,
            ).values_list(
                'quantidades_por_periodo__periodo_escolar__nome',
                'quantidades_por_periodo__periodo_escolar__uuid'))
            return Response({'periodos': periodos if len(periodos) else None})
        except ValidationError as e:
            return Response({'detail': e}, status=status.HTTP_400_BAD_REQUEST)


class PeriodosComMatriculadosPorUEViewSet(ReadOnlyModelViewSet):
    queryset = Escola.objects.all()

    def list(self, request, uuid=None):
        escola = self.get_object()
        periodos = escola.alunos_matriculados_por_periodo.filter(
            tipo_turma=TipoTurma.REGULAR.name, quantidade_alunos__gt=0
        ).values_list('periodo_escolar__nome', flat=True)
        return Response(periodos)

    def get_object(self):
        uuid = self.request.query_params.get('escola_uuid', None)
        return get_object_or_404(self.get_queryset(), uuid=uuid.rstrip('/'))

    def get_serializer_class(self):
        return AlunosMatriculadosPeriodoEscolaSerializer


class DiretoriaRegionalViewSet(ReadOnlyModelViewSet):
    lookup_field = 'uuid'
    queryset = DiretoriaRegional.objects.all()
    serializer_class = DiretoriaRegionalCompletaSerializer


class DiretoriaRegionalSimplissimaViewSet(ReadOnlyModelViewSet):
    permission_classes = [permissions.AllowAny]
    lookup_field = 'uuid'
    queryset = DiretoriaRegional.objects.all()
    serializer_class = DiretoriaRegionalSimplissimaSerializer

    @action(detail=False, methods=['GET'], url_path='lista-completa')
    def lista_completa(self, request):
        response = {'results': DiretoriaRegionalLookUpSerializer(self.get_queryset(), many=True).data}
        return Response(response)


class TipoGestaoViewSet(ReadOnlyModelViewSet):
    lookup_field = 'uuid'
    queryset = TipoGestao.objects.all()
    serializer_class = TipoGestaoSerializer


class SubprefeituraViewSet(ReadOnlyModelViewSet):
    permission_classes = [permissions.AllowAny]
    lookup_field = 'uuid'
    queryset = Subprefeitura.objects.all()
    serializer_class = SubprefeituraSerializer

    @action(detail=False, methods=['get'], url_path='lista-completa')
    def lista_completa(self, request):
        response = {'results': SubprefeituraSerializerSimples(self.get_queryset(), many=True).data}
        return Response(response)


class LoteViewSet(ModelViewSet):
    lookup_field = 'uuid'
    serializer_class = LoteSimplesSerializer
    queryset = Lote.objects.all()

    @action(detail=False, methods=['GET'], url_path='meus-lotes-vinculados')
    def meus_lotes_vinculados(self, request):
        if (request.user.tipo_usuario == 'diretoriaregional'):
            lotes = self.queryset.filter(diretoria_regional=request.user.vinculo_atual.instituicao)
        else:
            lotes = self.queryset.filter(terceirizada=request.user.vinculo_atual.instituicao)
        return Response({'results': LoteSerializer(lotes, many=True).data})

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return LoteCreateSerializer
        return LoteSimplesSerializer

    def destroy(self, request, *args, **kwargs):
        return Response({'detail': 'Não é permitido excluir um Lote com escolas associadas'},
                        status=status.HTTP_401_UNAUTHORIZED)


class LoteSimplesViewSet(ModelViewSet):
    lookup_field = 'uuid'
    serializer_class = LoteNomeSerializer
    queryset = Lote.objects.all()
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ('diretoria_regional__uuid',)


class CODAESimplesViewSet(ModelViewSet):
    lookup_field = 'uuid'
    queryset = Codae.objects.all()
    serializer_class = CODAESerializer


class TipoUnidadeEscolarViewSet(ReadOnlyModelViewSet):
    lookup_field = 'uuid'
    serializer_class = TipoUnidadeEscolarSerializer
    queryset = TipoUnidadeEscolar.objects.all()
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ('pertence_relatorio_solicitacoes_alimentacao',)


class LogAlunosMatriculadosPeriodoEscolaViewSet(ModelViewSet):
    serializer_class = LogAlunosMatriculadosPeriodoEscolaSerializer
    queryset = LogAlunosMatriculadosPeriodoEscola.objects.all()
    pagination_class = None

    def get_queryset(self):
        queryset = LogAlunosMatriculadosPeriodoEscola.objects.all()

        escola_uuid = self.request.query_params.get('escola_uuid', '')
        mes = self.request.query_params.get('mes', '')
        ano = self.request.query_params.get('ano', '')
        tipo_turma = self.request.query_params.get('tipo_turma', '')
        periodo_escolar = self.request.query_params.get('periodo_escolar', '')

        queryset = queryset.filter(escola__uuid=escola_uuid,
                                   criado_em__month=mes,
                                   criado_em__year=ano,
                                   tipo_turma=tipo_turma,
                                   periodo_escolar__uuid=periodo_escolar)

        return queryset


class EscolaPeriodoEscolarViewSet(ModelViewSet):
    lookup_field = 'uuid'

    def get_queryset(self):
        if self.request.user.tipo_usuario == 'escola':
            escola = self.request.user.vinculos.get(ativo=True).instituicao
            return EscolaPeriodoEscolar.objects.filter(escola=escola)
        return EscolaPeriodoEscolar.objects.all()

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return EscolaPeriodoEscolarCreateSerializer
        return EscolaPeriodoEscolarSerializer

    def destroy(self, request, *args, **kwargs):
        return Response({'detail': 'Não é permitido excluir um periodo já existente'},
                        status=status.HTTP_401_UNAUTHORIZED)

    @action(detail=False, url_path='escola/(?P<escola_uuid>[^/.]+)')
    def filtro_por_escola(self, request, escola_uuid=None):
        periodos = EscolaPeriodoEscolar.objects.filter(
            escola__uuid=escola_uuid
        )
        page = self.paginate_queryset(periodos)
        serializer = self.get_serializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    #  TODO: Quebrar esse método um pouco, está complexo e sem teste
    @action(detail=True, url_path='alunos-por-faixa-etaria/(?P<data_referencia_str>[^/.]+)')
    def alunos_por_faixa_etaria(self, request, uuid, data_referencia_str):  # noqa C901
        """
        EscolaCEI: Deve retornar a quantidade de alunos por faixa etária e período escolar através do uuid do
        PeriodoEscolar

        Outras escolas: Deve retornar a quantidade de alunos por faixa etária e período escolar através do uuid do
        EscolaPeriodoEscolar
        """

        form = AlunosPorFaixaEtariaForm({
            'data_referencia': data_referencia_str
        })

        if not form.is_valid():
            return Response(form.errors)

        if request.user.vinculo_atual:
            escola = request.user.vinculo_atual.instituicao
            if escola.eh_cei:
                escola_periodo = EscolaPeriodoEscolar.objects.get(periodo_escolar__uuid=uuid, escola__uuid=escola.uuid)
        else:
            escola_periodo = self.get_object()

        data_referencia = form.cleaned_data['data_referencia']

        try:
            faixa_alunos = escola_periodo.alunos_por_faixa_etaria(data_referencia)  # noqa
        except ObjectDoesNotExist:
            return Response(
                {'detail': 'Não há faixas etárias cadastradas. Contate a coordenadoria CODAE.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        results = []
        for uuid_faixa_etaria in faixa_alunos:
            results.append({
                'faixa_etaria': FaixaEtariaSerializer(FaixaEtaria.objects.get(uuid=uuid_faixa_etaria)).data,
                'count': faixa_alunos[uuid_faixa_etaria]
            })

        return Response({
            'count': len(results),
            'results': results
        })

#  TODO: Quebrar esse método um pouco, está complexo e sem teste
    @action(detail=True, url_path='matriculados-na-data/(?P<data_referencia_str>[^/.]+)')  # noqa C901
    def matriculados_na_data(self, request, uuid, data_referencia_str):
        form = AlunosPorFaixaEtariaForm({
            'data_referencia': data_referencia_str
        })

        if not form.is_valid():
            return Response(form.errors)

        escola_periodo = self.get_object()
        data_referencia = form.cleaned_data['data_referencia']

        log = LogAlteracaoQuantidadeAlunosPorEscolaEPeriodoEscolar.objects.filter(
            criado_em__gte=data_referencia,
            escola=escola_periodo.escola,
            periodo_escolar=escola_periodo.periodo_escolar
        ).order_by('criado_em').first()

        if log:
            quantidade_alunos = log.quantidade_alunos_de
        else:
            quantidade_alunos = escola_periodo.quantidade_alunos

        return Response({
            'quantidade_alunos': {
                'convencional': quantidade_alunos
            }
        })


class AlunoViewSet(RetrieveModelMixin, ListModelMixin, GenericViewSet):
    lookup_field = 'codigo_eol'
    queryset = Aluno.objects.all()
    serializer_class = AlunoSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = AlunoFilter

    def get_serializer_class(self):
        if self.action == 'list':
            return AlunoSimplesSerializer
        return self.serializer_class

    def get_queryset(self):
        if self.action == 'retrieve':
            return self.queryset.select_related('escola__diretoria_regional')
        return self.queryset

    @action(detail=True, methods=['GET'], url_path='aluno-pertence-a-escola/(?P<escola_codigo_eol>[^/.]+)')
    def aluno_pertence_a_escola(self, request, codigo_eol, escola_codigo_eol):
        escola = Escola.objects.filter(codigo_eol=escola_codigo_eol).first()
        aluno = Aluno.objects.filter(codigo_eol=codigo_eol).first()

        resposta = True if aluno and aluno.escola == escola else False
        return Response({'pertence_a_escola': resposta})

    @action(detail=True, methods=['GET'], url_path='ver-foto', permission_classes=(PodeVerEditarFotoAlunoNoSGP,))
    def ver_foto(self, request, codigo_eol):
        try:
            novosgpservicologado = NovoSGPServicoLogado()
            codigo_eol_ = self.get_object().codigo_eol
            response = novosgpservicologado.pegar_foto_aluno(codigo_eol_)
            if response.status_code == status.HTTP_200_OK:
                return Response({'data': response.json()}, status=response.status_code)
            return Response(status=response.status_code)
        except NovoSGPServicoLogadoException as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['POST'], url_path='atualizar-foto', permission_classes=(PodeVerEditarFotoAlunoNoSGP,))
    def atualizar_foto(self, request, codigo_eol):
        try:
            novosgpservicologado = NovoSGPServicoLogado()
            codigo_eol_ = self.get_object().codigo_eol
            response = novosgpservicologado.atualizar_foto_aluno(codigo_eol_, request.FILES['file'])
            if response.status_code == status.HTTP_200_OK:
                return Response({'data': response.json()}, status=response.status_code)
            return Response(status=response.status_code)
        except NovoSGPServicoLogadoException as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['DELETE'], url_path='deletar-foto', permission_classes=(PodeVerEditarFotoAlunoNoSGP,))
    def deletar_foto(self, request, codigo_eol):
        try:
            novosgpservicologado = NovoSGPServicoLogado()
            codigo_eol_ = self.get_object().codigo_eol
            response = novosgpservicologado.deletar_foto_aluno(codigo_eol_)
            if response.status_code == status.HTTP_200_OK:
                return Response({'data': response.json()}, status=response.status_code)
            return Response(status=response.status_code)
        except NovoSGPServicoLogadoException as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=('GET',), url_path='quantidade-cemei-por-cei-emei',  # noqa C901
            permission_classes=(IsAuthenticated,))
    def quantidade_cemei_por_cei_emei(self, request):
        try:
            codigo_eol_escola = request.query_params.get('codigo_eol_escola', None)
            manha_e_tarde_sempre = request.query_params.get('manha_e_tarde_sempre', False)
            if not codigo_eol_escola:
                raise ValidationError('`codigo_eol_escola` como query_param é obrigatório')
            escola = Escola.objects.get(codigo_eol=codigo_eol_escola)
            if not escola.eh_cemei:
                raise ValidationError('escola não é CEMEI')
            return Response(escola.quantidade_alunos_por_cei_emei(manha_e_tarde_sempre == 'true'),
                            status=status.HTTP_200_OK)
        except ValidationError as e:
            return Response(e, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=('GET',), url_path='quantidade-alunos-por-periodo-cei-emei',  # noqa C901
            permission_classes=(IsAuthenticated,))
    def quantidade_alunos_por_periodo_cei_emei(self, request):
        try:
            codigo_eol_escola = request.query_params.get('codigo_eol_escola', None)
            if not codigo_eol_escola:
                raise ValidationError('`codigo_eol_escola` como query_param é obrigatório')
            escola = Escola.objects.get(codigo_eol=codigo_eol_escola)
            if not escola.eh_cemei:
                raise ValidationError('escola não é CEMEI')
            return Response(escola.quantidade_alunos_por_periodo_cei_emei, status=status.HTTP_200_OK)
        except ValidationError as e:
            return Response(e, status=status.HTTP_400_BAD_REQUEST)


class FaixaEtariaViewSet(CreateModelMixin, ListModelMixin, GenericViewSet):
    queryset = FaixaEtaria.objects.filter(ativo=True)

    def get_serializer_class(self):
        if self.action == 'create':
            return MudancaFaixasEtariasCreateSerializer
        return FaixaEtariaSerializer


class DiaCalendarioViewSet(ModelViewSet):
    serializer_class = DiaCalendarioSerializer
    queryset = DiaCalendario.objects.all()
    pagination_class = None

    def get_queryset(self):
        queryset = DiaCalendario.objects.all()

        escola_uuid = self.request.query_params.get('escola_uuid', '')
        mes = self.request.query_params.get('mes', '')
        ano = self.request.query_params.get('ano', '')

        queryset = queryset.filter(escola__uuid=escola_uuid,
                                   data__month=mes,
                                   data__year=ano)

        return queryset


def exportar_planilha_importacao_tipo_gestao_escola(request, **kwargs):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=planilha_importacao_tipo_gestao_escolas.xlsx'
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = 'UNIDADES COM TIPO DE GESTÃO'
    headers = [
        'CÓDIGO EOL',
        'CÓDIGO CODAE',
        'NOME UNIDADE',
        'TIPO',
    ]
    _font = styles.Font(name='Calibri', sz=11)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill('solid', fgColor='ffff99')
        cabecalho.font = styles.Font(name='Calibri', size=11, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style='thin', color='000000'),
            right=styles.Side(border_style='thin', color='000000'),
            top=styles.Side(border_style='thin', color='000000'),
            bottom=styles.Side(border_style='thin', color='000000')
        )
    dv = DataValidation(
        type='list',
        formula1='"PARCEIRA, DIRETA, MISTA, TERCEIRIZADA TOTAL"',
        allow_blank=True
    )
    dv.error = 'Tipo Inválido'
    dv.errorTitle = 'Tipo não permitido'
    ws.add_data_validation(dv)
    dv.add('D2:H1048576')

    for colunas in ws.columns:
        unmerged_cells = list(
            filter(lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells, colunas))
        length = max(len(str(cell.value)) for cell in unmerged_cells)
        ws.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.3
    workbook.save(response)

    return response


class RelatorioAlunosMatriculadosViewSet(ModelViewSet):

    @action(detail=False, methods=['GET'], url_path='filtros')
    def filtros(self, request):
        terceirizada = request.user.vinculo_atual.instituicao
        lotes = terceirizada.lotes.all()
        diretorias_regionais_uuids = lotes.values_list('diretoria_regional__uuid', flat=True).distinct()
        diretorias_regionais = DiretoriaRegional.objects.filter(uuid__in=diretorias_regionais_uuids)
        escolas = Escola.objects.filter(diretoria_regional__uuid__in=diretorias_regionais_uuids)
        tipos_unidade_uuids = escolas.values_list('tipo_unidade__uuid', flat=True)
        tipos_unidade_escolar = TipoUnidadeEscolar.objects.filter(uuid__in=tipos_unidade_uuids)
        filtros = {
            'lotes': LoteParaFiltroSerializer(lotes, many=True).data,
            'diretorias_regionais': DiretoriaRegionalParaFiltroSerializer(diretorias_regionais, many=True).data,
            'tipos_unidade_escolar': TipoUnidadeParaFiltroSerializer(tipos_unidade_escolar, many=True).data,
            'escolas': EscolaParaFiltroSerializer(escolas, many=True).data,
        }
        return Response(filtros, status=status.HTTP_200_OK)
