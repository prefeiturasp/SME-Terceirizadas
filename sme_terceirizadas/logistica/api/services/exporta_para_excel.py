import os
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from ..helpers import (
    retorna_motivo_insucesso,
    retorna_ocorrencias_alimento,
    retorna_status_alimento,
    retorna_status_guia_remessa
)


class RequisicoesExcelService(object):
    DEFAULT_BORDER = Border(right=Side(border_style='thin', color='24292E'),
                            left=Side(border_style='thin', color='24292E'),
                            top=Side(border_style='thin', color='24292E'),
                            bottom=Side(border_style='thin', color='24292E'))

    @classmethod
    def aplicar_tamanho_calculado_nas_celulas(cls, ws):
        # Ajuste automatico do tamanho das colunas
        for colunas in ws.columns:
            unmerged_cells = list(
                filter(lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells, colunas))
            length = max(len(str(cell.value)) for cell in unmerged_cells)
            ws.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.2

    @classmethod
    def aplicar_estilo_padrao(cls, ws, count_data, count_fields):
        for linha in range(1, (count_data + 2)):
            for coluna in range(1, (count_fields + 1)):
                celula = ws.cell(row=linha, column=coluna)
                celula.border = cls.DEFAULT_BORDER
                if linha == 1:
                    celula.fill = PatternFill(fill_type='solid', fgColor='198459')

        cls.aplicar_tamanho_calculado_nas_celulas(ws)

    @classmethod
    def aplicar_estilo_visao_distribuidor(cls, ws, count_data, count_fields):
        for linha in range(2, (count_data + 3)):
            for coluna in range(1, (count_fields + 1)):
                celula = ws.cell(row=linha, column=coluna)
                celula.border = cls.DEFAULT_BORDER
                if linha == 2:
                    celula.fill = PatternFill(fill_type='solid', fgColor='198459')

        cls.aplicar_tamanho_calculado_nas_celulas(ws)

    @classmethod
    def gera_arquivo(cls, wb):
        with NamedTemporaryFile(delete=False) as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            arquivo = tmp.read()
            tmp.close()
            os.unlink(tmp.name)
            return arquivo

    @classmethod # noqa C901
    def exportar_visao_distribuidor(cls, requisicoes, is_async=False):

        cabecalho = ['Nº da Requisição',
                     'Status da Requisição',
                     'Data de Entrega',
                     'Alimento',
                     'Nome da UE',
                     'Código EOL da UE',
                     'Endereço da UE',
                     'Bairro da UE',
                     'Cep da UE',
                     'Telefone UE',
                     'Número da Guia',
                     'Quantidade',
                     'Capacidade (Embalagem Fechada)',
                     'Quantidade (Fracionada)',
                     'Capacidade (Embalagem Fracionada)',
                     'Código Supri',
                     'Agrupamento']

        count_fields = len(cabecalho)
        count_data = requisicoes.count()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Visão Analítica Abastecimento'

        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=2, column=ind)
            celula.value = title
            celula.font = Font(size='13', bold=True, color='00FFFFFF')

        for ind, requisicao in enumerate(requisicoes, 3):
            ws.cell(row=ind, column=1, value=int(requisicao['numero_solicitacao']))
            ws.cell(row=ind, column=2, value=requisicao['status_requisicao'])
            ws.cell(row=ind, column=3, value=requisicao['guias__data_entrega'].strftime('%d/%m/%Y'))
            ws.cell(row=ind, column=4, value=requisicao['guias__alimentos__nome_alimento'])
            ws.cell(row=ind, column=5, value=requisicao['guias__nome_unidade'])
            ws.cell(row=ind, column=6, value=requisicao['guias__escola__codigo_eol'])
            ws.cell(row=ind, column=7, value=f'{requisicao["guias__endereco_unidade"]} '
                                             f'{requisicao["guias__numero_unidade"]}')
            ws.cell(row=ind, column=8, value=requisicao['guias__bairro_unidade'])
            ws.cell(row=ind, column=9, value=requisicao['guias__cep_unidade'])
            ws.cell(row=ind, column=10, value=requisicao['guias__telefone_unidade'])
            ws.cell(row=ind, column=11, value=int(requisicao['guias__numero_guia']))
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=12,
                        value=f'{requisicao["guias__alimentos__embalagens__qtd_volume"]} '
                              f'{requisicao["guias__alimentos__embalagens__descricao_embalagem"]}')
                ws.cell(row=ind, column=13,
                        value=f'{requisicao["guias__alimentos__embalagens__capacidade_embalagem"]} '
                              f'{requisicao["guias__alimentos__embalagens__unidade_medida"]}')
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FRACIONADA':
                ws.cell(row=ind, column=14,
                        value=f'{requisicao["guias__alimentos__embalagens__qtd_volume"]} '
                              f'{requisicao["guias__alimentos__embalagens__descricao_embalagem"]}')
                ws.cell(row=ind, column=15,
                        value=f'{requisicao["guias__alimentos__embalagens__capacidade_embalagem"]} '
                              f'{requisicao["guias__alimentos__embalagens__unidade_medida"]}')
            ws.cell(row=ind, column=16, value=requisicao['guias__alimentos__codigo_suprimento'])
            ws.cell(row=ind, column=17, value=requisicao['guias__escola__subprefeitura__agrupamento'])

        cls.aplicar_estilo_visao_distribuidor(ws, count_data, count_fields)
        arquivo = cls.gera_arquivo(wb)
        filename = 'visao-consolidada.xlsx'

        return arquivo if is_async else {'arquivo': arquivo, 'filename': filename}

    @classmethod
    def exportar_visao_dilog(cls, requisicoes, is_async=False):

        cabecalho = ['Nome do Distribuidor', 'Número da Requisição', 'Status da Requisição',
                     'Quantidade Total de Guias', 'Número da Guia', 'Data de Entrega', 'Código EOL da UE',
                     'Código CODAE da UE', 'Nome UE', 'Endereço UE', 'Número UE', 'Bairro UE', 'CEP UE', 'Cidade UE',
                     'Estado UE', 'Contato de Entrega', 'Telefone UE', 'Nome do Alimento', 'Código SUPRI',
                     'Código PAPA', 'Descrição Embalagem Fechada', 'Capacidade da Embalagem Fechada',
                     'Unidade de Medida da Embalagem Fechada', 'Quantidade de Volumes da Embalagem Fechada',
                     'Descrição Embalagem Fracionada', 'Capacidade da Embalagem Fracionada',
                     'Unidade de Medida da Embalagem Fracionada', 'Quantidade de Volumes da Embalagem Fracionada',
                     'Status da Guia']

        count_fields = len(cabecalho)
        count_data = requisicoes.count()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Visão Analítica Abastecimento'

        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = title
            celula.font = Font(size='13', bold=True, color='00FFFFFF')

        for ind, requisicao in enumerate(requisicoes, 2):
            ws.cell(row=ind, column=1, value=requisicao['distribuidor__nome_fantasia'])
            ws.cell(row=ind, column=2, value=requisicao['numero_solicitacao'])
            ws.cell(row=ind, column=3, value=requisicao['status_requisicao'])
            ws.cell(row=ind, column=4, value=requisicao['quantidade_total_guias'])
            ws.cell(row=ind, column=5, value=requisicao['guias__numero_guia'])
            ws.cell(row=ind, column=6, value=requisicao['guias__data_entrega'].strftime('%d/%m/%Y'))
            ws.cell(row=ind, column=7, value=requisicao['codigo_eol_unidade'])
            ws.cell(row=ind, column=8, value=requisicao['guias__codigo_unidade'])
            ws.cell(row=ind, column=9, value=requisicao['guias__nome_unidade'])
            ws.cell(row=ind, column=10, value=requisicao['guias__endereco_unidade'])
            ws.cell(row=ind, column=11, value=(requisicao['guias__numero_unidade']))
            ws.cell(row=ind, column=12, value=requisicao['guias__bairro_unidade'])
            ws.cell(row=ind, column=13, value=requisicao['guias__cep_unidade'])
            ws.cell(row=ind, column=14, value=requisicao['guias__cidade_unidade'])
            ws.cell(row=ind, column=15, value=requisicao['guias__estado_unidade'])
            ws.cell(row=ind, column=16, value=requisicao['guias__contato_unidade'])
            ws.cell(row=ind, column=17, value=requisicao['guias__telefone_unidade'])
            ws.cell(row=ind, column=18, value=requisicao['guias__alimentos__nome_alimento'])
            ws.cell(row=ind, column=19, value=requisicao['guias__alimentos__codigo_suprimento'])
            ws.cell(row=ind, column=20, value=requisicao['guias__alimentos__codigo_papa'])
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=21, value=requisicao['guias__alimentos__embalagens__descricao_embalagem'])
                ws.cell(row=ind, column=22, value=requisicao['guias__alimentos__embalagens__capacidade_embalagem'])
                ws.cell(row=ind, column=23, value=requisicao['guias__alimentos__embalagens__unidade_medida'])
                ws.cell(row=ind, column=24, value=requisicao['guias__alimentos__embalagens__qtd_volume'])
            else:
                ws.cell(row=ind, column=25, value=requisicao['guias__alimentos__embalagens__descricao_embalagem'])
                ws.cell(row=ind, column=26, value=requisicao['guias__alimentos__embalagens__capacidade_embalagem'])
                ws.cell(row=ind, column=27, value=requisicao['guias__alimentos__embalagens__unidade_medida'])
                ws.cell(row=ind, column=28, value=requisicao['guias__alimentos__embalagens__qtd_volume'])
            ws.cell(row=ind, column=29, value=requisicao['guias__status'])

        cls.aplicar_estilo_padrao(ws, count_data, count_fields)
        arquivo = cls.gera_arquivo(wb)

        return arquivo if is_async else {'arquivo': arquivo}

    @classmethod
    def cria_aba_insucesso(cls, ws, requisicoes, perfil):  # noqa C901
        offset = 0
        cabecalho = ['Número da Requisição', 'Quantidade Total de Guias', 'Número da Guia', 'Data de Entrega',
                     'Nome do Motorista', 'Placa do Veículo',
                     'Data de Registro do Insucesso', 'Hora de Registro do Insucesso',
                     'Código EOL', 'Código CODAE', 'Nome UE', 'Endereço UE', 'Número UE', 'Bairro UE',
                     'CEP UE', 'Cidade UE', 'Estado UE', 'Contato de Entrega', 'Telefone UE', 'Nome do Alimento',
                     'Código SUPRI', 'Código PAPA',
                     'Descrição Embalagem Fechada', 'Capacidade da Embalagem Fechada',
                     'Unidade de Medida da Embalagem Fechada',
                     'Quantidade Prevista (Volumes da Embalagem Fechada)',
                     'Descrição Embalagem Fracionada', 'Capacidade da Embalagem Fracionada',
                     'Unidade de Medida da Embalagem Fracionada',
                     'Quantidade Prevista (Volumes da Embalagem Fracionada)',
                     'Hora da tentativa de entrega', 'Motivo',
                     'Justificativa',
                     'Status da Guia de Remessa',
                     'Nome Completo do Conferente',
                     'Documento do Conferente']

        if perfil == 'DILOG':
            cabecalho.insert(0, 'Nome do Distribuidor')
            offset = 1

        count_fields = len(cabecalho)
        count_data = requisicoes.count()

        ws.title = 'Relatório de Insucesso'
        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = title
            celula.font = Font(size='13', bold=True, color='00FFFFFF')

        for ind, requisicao in enumerate(requisicoes, 2):
            if perfil == 'DILOG':
                ws.cell(row=ind, column=offset, value=requisicao['distribuidor__nome_fantasia'])
            ws.cell(row=ind, column=offset + 1, value=requisicao['numero_solicitacao'])
            ws.cell(row=ind, column=offset + 2, value=requisicao['quantidade_total_guias'])
            ws.cell(row=ind, column=offset + 3, value=requisicao['guias__numero_guia'])
            ws.cell(row=ind, column=offset + 4, value=requisicao['guias__data_entrega'])
            if requisicao['guias__insucessos__criado_em'] is not None:
                ws.cell(row=ind, column=offset + 5, value=requisicao['guias__insucessos__nome_motorista'])
                ws.cell(row=ind, column=offset + 6, value=requisicao['guias__insucessos__placa_veiculo'])
                ws.cell(row=ind, column=offset + 7,
                        value=requisicao['guias__insucessos__criado_em'].strftime('%d/%m/%Y'))
                ws.cell(row=ind, column=offset + 8,
                        value=requisicao['guias__insucessos__criado_em'].strftime('%H:%M:%S'))
            ws.cell(row=ind, column=offset + 9, value=requisicao['guias__escola__codigo_eol'])
            ws.cell(row=ind, column=offset + 10, value=requisicao['guias__codigo_unidade'])
            ws.cell(row=ind, column=offset + 11, value=requisicao['guias__nome_unidade'])
            ws.cell(row=ind, column=offset + 12, value=requisicao['guias__endereco_unidade'])
            ws.cell(row=ind, column=offset + 13, value=requisicao['guias__numero_unidade'])
            ws.cell(row=ind, column=offset + 14, value=requisicao['guias__bairro_unidade'])
            ws.cell(row=ind, column=offset + 15, value=requisicao['guias__cep_unidade'])
            ws.cell(row=ind, column=offset + 16, value=requisicao['guias__cidade_unidade'])
            ws.cell(row=ind, column=offset + 17, value=requisicao['guias__estado_unidade'])
            ws.cell(row=ind, column=offset + 18, value=requisicao['guias__contato_unidade'])
            ws.cell(row=ind, column=offset + 19, value=requisicao['guias__telefone_unidade'])
            ws.cell(row=ind, column=offset + 20, value=requisicao['guias__alimentos__nome_alimento'])
            ws.cell(row=ind, column=offset + 21, value=requisicao['guias__alimentos__codigo_suprimento'])
            ws.cell(row=ind, column=offset + 22, value=requisicao['guias__alimentos__codigo_papa'])
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=offset + 23,
                        value=requisicao['guias__alimentos__embalagens__descricao_embalagem'])
                ws.cell(row=ind, column=offset + 24,
                        value=requisicao['guias__alimentos__embalagens__capacidade_embalagem'])
                ws.cell(row=ind, column=offset + 25, value=requisicao['guias__alimentos__embalagens__unidade_medida'])
                ws.cell(row=ind, column=offset + 26, value=requisicao['guias__alimentos__embalagens__qtd_volume'])
            else:
                ws.cell(row=ind, column=offset + 27,
                        value=requisicao['guias__alimentos__embalagens__descricao_embalagem'])
                ws.cell(row=ind, column=offset + 28,
                        value=requisicao['guias__alimentos__embalagens__capacidade_embalagem'])
                ws.cell(row=ind, column=offset + 29, value=requisicao['guias__alimentos__embalagens__unidade_medida'])
                ws.cell(row=ind, column=offset + 30, value=requisicao['guias__alimentos__embalagens__qtd_volume'])
            if requisicao['guias__insucessos__criado_em'] is not None:
                ws.cell(row=ind, column=offset + 31,
                        value=requisicao['guias__insucessos__hora_tentativa'].strftime('%H:%M:%S'))
                ws.cell(row=ind, column=offset + 32,
                        value=retorna_motivo_insucesso(requisicao['guias__insucessos__motivo']))
                ws.cell(row=ind, column=offset + 33, value=requisicao['guias__insucessos__justificativa'])
                ws.cell(row=ind, column=offset + 34, value=retorna_status_guia_remessa(requisicao['guias__status']))
                ws.cell(row=ind, column=offset + 35, value=requisicao['guias__insucessos__criado_por__nome'])
                ws.cell(row=ind, column=offset + 36, value=requisicao['guias__insucessos__criado_por__cpf'])

        cls.aplicar_estilo_padrao(ws, count_data, count_fields)

    @classmethod
    def cria_aba_conferencia_dilog(cls, ws, requisicoes):  # noqa C901
        cabecalho = ['Data de Entrega',
                     'Nome do Alimento',
                     'Nome do Distribuidor',
                     'Número da Guia de Remessa',
                     'Status da Guia de Remessa',
                     'Nome da UE',
                     'Telefone da UE',
                     'Endereço da UE ',
                     'Bairro da UE',
                     'Contato da Entrega',
                     'Código EOL',
                     'Código PAPA',
                     'Código CODAE',
                     'Data e Hora do Recebimento (1ª Conferência) ',
                     'Data e hora do Registro (1ª Conferência) ',
                     'Quantidade Prevista (Embalagem Fechada) ',
                     'Quantidade Prevista (Embalagem Fracionada) ',
                     'Quantidade Recebida (Embalagem Fechada)',
                     'Quantidade Recebida (Embalagem Fracionada)',
                     'Nome Completo do Conferente (1ª Conferência)',
                     'Nome do Motorista (1ª Conferência)',
                     'Placa do Veículo (1ª Conferência)',
                     'Status de Recebimento do Alimento (1ª Conferência)',
                     'Ocorrências (1ª Conferência)',
                     'Observações (1ª Conferência)',
                     'Quantidade a ​Receber (Embalagem Fechada)',
                     'Quantidade a ​Receber (Embalagem Fracionada)',
                     'Data e Hora do Recebimento (Reposição)',
                     'Data e Hora de Registro (Reposição)',
                     'Status de Recebimento do Alimento (Reposição)',
                     'Quantidade Reposta (Embalagem Fechada)',
                     'Quantidade Reposta (Embalagem Fracionada)',
                     'Nome Completo do Conferente (Reposição)',
                     'Nome do Motorista (Reposição)',
                     'Placa do Veículo (Reposição)',
                     'Ocorrências (Reposição)',
                     'Observações (Reposição)']

        count_fields = len(cabecalho)
        count_data = requisicoes.count()

        ws.title = 'Relatório de Conferência'
        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = title
            celula.font = Font(size='13', bold=True, color='00FFFFFF')

        for ind, requisicao in enumerate(requisicoes, 2):

            qtd_recebida = 0

            if 'conferencia_alimento' in requisicao and requisicao['conferencia_alimento'] is not None:
                qtd_recebida = requisicao['conferencia_alimento'].qtd_recebido
            elif 'primeira_conferencia' in requisicao:
                qtd_recebida = requisicao['guias__alimentos__embalagens__qtd_volume']
            else:
                qtd_recebida = ''

            ws.cell(row=ind, column=1, value=requisicao['guias__data_entrega'])
            ws.cell(row=ind, column=2, value=requisicao['guias__alimentos__nome_alimento'])
            ws.cell(row=ind, column=3, value=requisicao['distribuidor__nome_fantasia'])
            ws.cell(row=ind, column=4, value=requisicao['guias__numero_guia'])
            ws.cell(row=ind, column=5, value=retorna_status_guia_remessa(requisicao['guias__status']))
            ws.cell(row=ind, column=6, value=requisicao['guias__nome_unidade'])
            ws.cell(row=ind, column=7, value=requisicao['guias__telefone_unidade'])
            ws.cell(row=ind, column=8, value=f'{requisicao["guias__endereco_unidade"]} '
                                             f'{requisicao["guias__numero_unidade"]}')
            ws.cell(row=ind, column=9, value=requisicao['guias__bairro_unidade'])
            ws.cell(row=ind, column=10, value=requisicao['guias__contato_unidade'])
            ws.cell(row=ind, column=11, value=requisicao['guias__escola__codigo_eol'])
            ws.cell(row=ind, column=12, value=requisicao['guias__alimentos__codigo_papa'])
            ws.cell(row=ind, column=13, value=requisicao['guias__codigo_unidade'])
            if 'primeira_conferencia' in requisicao:
                ws.cell(row=ind, column=14, value=f'{requisicao["primeira_conferencia"].data_recebimento} '
                                                  f'{requisicao["primeira_conferencia"].hora_recebimento}')
                ws.cell(row=ind, column=15,
                        value=requisicao['primeira_conferencia'].criado_em.strftime('%d/%m/%Y %H:%M:%S'))
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=16,
                        value=f'{requisicao["guias__alimentos__embalagens__qtd_volume"]} '
                              f'{requisicao["guias__alimentos__embalagens__descricao_embalagem"]}')
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FRACIONADA':
                ws.cell(row=ind, column=17,
                        value=f'{requisicao["guias__alimentos__embalagens__qtd_volume"]} '
                              f'{requisicao["guias__alimentos__embalagens__descricao_embalagem"]}')
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=18, value=qtd_recebida)
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FRACIONADA':
                ws.cell(row=ind, column=19, value=qtd_recebida)
            if 'primeira_conferencia' in requisicao:
                ws.cell(row=ind, column=20, value=requisicao['primeira_conferencia'].criado_por.nome)
                ws.cell(row=ind, column=21, value=requisicao['primeira_conferencia'].nome_motorista)
                ws.cell(row=ind, column=22, value=requisicao['primeira_conferencia'].placa_veiculo)
                if 'conferencia_alimento' in requisicao:
                    ws.cell(row=ind, column=23, value=retorna_status_alimento(
                        requisicao['conferencia_alimento'].status_alimento
                    ))
                    ws.cell(row=ind, column=24, value=retorna_ocorrencias_alimento(
                        requisicao['conferencia_alimento'].ocorrencia
                    ))
                    ws.cell(row=ind, column=25, value=requisicao['conferencia_alimento'].observacao)
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=26, value=requisicao['guias__alimentos__embalagens__qtd_a_receber'])
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FRACIONADA':
                ws.cell(row=ind, column=27, value=requisicao['guias__alimentos__embalagens__qtd_a_receber'])
            if 'primeira_reposicao' in requisicao:
                ws.cell(row=ind, column=28, value=f'{requisicao["primeira_reposicao"].data_recebimento} '
                                                  f'{requisicao["primeira_reposicao"].hora_recebimento}')
                ws.cell(row=ind, column=29,
                        value=requisicao['primeira_reposicao'].criado_em.strftime('%d/%m/%Y %H:%M:%S'))
            if 'primeira_reposicao' in requisicao:
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=30, value=retorna_status_alimento(
                        requisicao['reposicao_alimento'].status_alimento
                    ))
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=31, value=requisicao['reposicao_alimento'].qtd_recebido)
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FRACIONADA':
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=32, value=requisicao['reposicao_alimento'].qtd_recebido)
            if 'primeira_reposicao' in requisicao:
                ws.cell(row=ind, column=33, value=requisicao['primeira_reposicao'].criado_por.nome)
                ws.cell(row=ind, column=34, value=requisicao['primeira_reposicao'].nome_motorista)
                ws.cell(row=ind, column=35, value=requisicao['primeira_reposicao'].placa_veiculo)
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=36, value=retorna_ocorrencias_alimento(
                        requisicao['reposicao_alimento'].ocorrencia
                    ))
                    ws.cell(row=ind, column=37, value=requisicao['reposicao_alimento'].observacao)

        cls.aplicar_estilo_padrao(ws, count_data, count_fields)

    @classmethod
    def cria_aba_conferencia_dre(cls, ws, requisicoes):  # noqa C901
        offset = 0
        cabecalho = ['Código EOL',
                     'Código CODAE',
                     'Nome da UE',
                     'Telefone UE',
                     'Data de Entrega',
                     'Número da Guia',
                     'Status da Guia de Remessa',
                     'Data e Hora do Recebimento (1ª Conferência)',
                     'Data e Hora do Registro (1ª Conferência)',
                     'Data e Hora do Recebimento (Reposição)',
                     'Data e Hora do Registro (Reposição)',
                     'Nome do Alimento',
                     'Quantidade Prevista (Embalagem Fechada)',
                     'Quantidade Prevista (Embalagem Fracionada)',
                     'Quantidade a Repor (Embalagem Fechada)',
                     'Quantidade a Repor (Embalagem Fracionada)',
                     'Status de Recebimento do alimento (1ª Conferência)',
                     'Ocorrências (1ª Conferência)',
                     'Observações (1ª Conferência)',
                     'Nome Completo do Conferente (1ª Conferência)',
                     'Quantidade Reposta (Embalagem Fechada)',
                     'Quantidade Reposta (Embalagem Fracionada)',
                     'Status de Recebimento do Alimento (Reposição)',
                     'Ocorrências (Reposição)',
                     'Observações (Reposição)',
                     'Nome Completo do Conferente (Reposição)'
                     ]

        offset = 1

        count_fields = len(cabecalho)
        count_data = requisicoes.count()

        ws.title = 'Relatório de Conferência'
        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = title
            celula.font = Font(size='13', bold=True, color='00FFFFFF')

        for ind, requisicao in enumerate(requisicoes, 2):

            qtd_recebida = 0

            if 'conferencia_alimento' in requisicao and requisicao['conferencia_alimento'] is not None:
                qtd_recebida = requisicao['conferencia_alimento'].qtd_recebido
            elif 'primeira_conferencia' in requisicao:
                qtd_recebida = requisicao['guias__alimentos__embalagens__qtd_volume']
            else:
                qtd_recebida = ''

            ws.cell(row=ind, column=offset, value=requisicao['guias__escola__codigo_eol'])
            ws.cell(row=ind, column=offset + 1, value=requisicao['guias__codigo_unidade'])
            ws.cell(row=ind, column=offset + 2, value=requisicao['guias__nome_unidade'])
            ws.cell(row=ind, column=offset + 3, value=requisicao['guias__telefone_unidade'])
            ws.cell(row=ind, column=offset + 4, value=requisicao['guias__data_entrega'])
            ws.cell(row=ind, column=offset + 5, value=requisicao['guias__numero_guia'])
            ws.cell(row=ind, column=offset + 6, value=retorna_status_guia_remessa(requisicao['guias__status']))
            if 'primeira_conferencia' in requisicao:
                ws.cell(row=ind, column=offset + 7, value=f'{requisicao["primeira_conferencia"].data_recebimento} '
                                                          f'{requisicao["primeira_conferencia"].hora_recebimento}')
                ws.cell(row=ind, column=offset + 8,
                        value=f'{requisicao["primeira_conferencia"].criado_em.strftime("%d/%m/%Y")} '
                              f'{requisicao["primeira_conferencia"].criado_em.strftime("%H:%M:%S")}')
            if 'primeira_reposicao' in requisicao:
                ws.cell(row=ind, column=offset + 9, value=f'{requisicao["primeira_reposicao"].data_recebimento} '
                                                          f'{requisicao["primeira_reposicao"].hora_recebimento}')
                ws.cell(row=ind, column=offset + 10,
                        value=f'{requisicao["primeira_reposicao"].criado_em.strftime("%d/%m/%Y")} '
                              f'{requisicao["primeira_reposicao"].criado_em.strftime("%H:%M:%S")}')
            ws.cell(row=ind, column=offset + 11, value=requisicao['guias__alimentos__nome_alimento'])
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=offset + 12,
                        value=f'{requisicao["guias__alimentos__embalagens__qtd_volume"]} '
                              f'{requisicao["guias__alimentos__embalagens__descricao_embalagem"]}')
            else:
                ws.cell(row=ind, column=offset + 13,
                        value=f'{requisicao["guias__alimentos__embalagens__qtd_volume"]} '
                              f'{requisicao["guias__alimentos__embalagens__descricao_embalagem"]}')

            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=offset + 14, value=requisicao['guias__alimentos__embalagens__qtd_a_receber'])
            else:
                ws.cell(row=ind, column=offset + 15, value=requisicao['guias__alimentos__embalagens__qtd_a_receber'])
            if 'primeira_conferencia' in requisicao:
                if 'conferencia_alimento' in requisicao:
                    ws.cell(row=ind, column=offset + 16, value=retorna_status_alimento(
                        requisicao['conferencia_alimento'].status_alimento
                    ))
                    ws.cell(row=ind, column=offset + 17, value=retorna_ocorrencias_alimento(
                        requisicao['conferencia_alimento'].ocorrencia
                    ))
                    ws.cell(row=ind, column=offset + 18, value=requisicao['conferencia_alimento'].observacao)
                    ws.cell(row=ind, column=offset + 19, value=requisicao['primeira_conferencia'].criado_por.nome)
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=offset + 20, value=qtd_recebida)
            else:
                ws.cell(row=ind, column=offset + 21, value=qtd_recebida)
            if 'primeira_reposicao' in requisicao:
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=offset + 22, value=retorna_status_alimento(
                        requisicao['reposicao_alimento'].status_alimento
                    ))
                    ws.cell(row=ind, column=offset + 23, value=retorna_ocorrencias_alimento(
                        requisicao['reposicao_alimento'].ocorrencia
                    ))
                    ws.cell(row=ind, column=offset + 24, value=requisicao['reposicao_alimento'].observacao)
                    ws.cell(row=ind, column=offset + 25, value=requisicao['primeira_reposicao'].criado_por.nome)
        cls.aplicar_estilo_padrao(ws, count_data, count_fields)

    @classmethod
    def cria_aba_conferencia_distribuidor(cls, ws, requisicoes):  # noqa C901
        offset = 0
        cabecalho = ['Número da Requisição', 'Data de Entrega', 'Nome do Alimento', 'Código SUPRI',
                     'Número da Guia de Remessa', 'Status da Guia de Remessa', 'Nome da UE', 'Código CODAE', 'Agrup',
                     'Telefone da UE', 'Endereço da UE', 'Bairro da UE', 'CEP da UE', 'Contato da Entrega',
                     'Quantidade prevista (Embalagem Fechada)', 'Capacidade (Embalagem Fechada)',
                     'Quantidade prevista (Embalagem Fracionada)', 'Capacidade (Embalagem Fracionada)',
                     'Data e Hora do recebimento (1ª Conferência)', 'Nome completo do conferente (1ª Conferência)',
                     'Quantidade recebida (Embalagem Fechada)',
                     'Quantidade a ​repor, referente a quantidade a receber(Embalagem Fechada)',
                     'Quantidade recebida (Embalagem Fracionada)',
                     'Quantidade a ​repor, referente a quantidade a receber (Embalagem Fracionada)',
                     'Nome do motorista (1ª Conferência)', 'Placa do veículo (1ª Conferência)',
                     'Ocorrências (1ª Conferência)', 'Observação (1ª Conferência)',
                     'Reposição (Data e Hora do recebimento)', 'Data de registro da reposição (com hora)',
                     'Nome completo do conferente (Reposição)', 'Quantidade reposta (Embalagem Fechada)',
                     'Quantidade reposta (Embalagem Fracionada)', 'Nome do motorista (Reposição)',
                     'Placa do veículo (Reposição)', 'Status de recebimento do alimento (Reposição)',
                     'Ocorrências (Reposição)', 'Observações (Reposição)']

        count_fields = len(cabecalho)
        count_data = requisicoes.count()

        ws.title = 'Relatório de Conferência'
        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = title
            celula.font = Font(size='13', bold=True, color='00FFFFFF')

        for ind, requisicao in enumerate(requisicoes, 2):

            if 'conferencia_alimento' in requisicao and requisicao['conferencia_alimento'] is not None:
                qtd_recebida = requisicao['conferencia_alimento'].qtd_recebido
            elif 'primeira_conferencia' in requisicao:
                qtd_recebida = requisicao['guias__alimentos__embalagens__qtd_volume']
            else:
                qtd_recebida = ''

            ws.cell(row=ind, column=offset + 1, value=requisicao['numero_solicitacao'])
            ws.cell(row=ind, column=offset + 2, value=requisicao['guias__data_entrega'])
            ws.cell(row=ind, column=offset + 3, value=requisicao['guias__alimentos__nome_alimento'])
            ws.cell(row=ind, column=offset + 4, value=requisicao['guias__alimentos__codigo_suprimento'])
            ws.cell(row=ind, column=offset + 5, value=requisicao['guias__numero_guia'])
            ws.cell(row=ind, column=offset + 6, value=retorna_status_guia_remessa(requisicao['guias__status']))
            ws.cell(row=ind, column=offset + 7, value=requisicao['guias__nome_unidade'])
            ws.cell(row=ind, column=offset + 8, value=requisicao['guias__codigo_unidade'])
            ws.cell(row=ind, column=offset + 9, value=requisicao['guias__escola__subprefeitura__agrupamento'])
            ws.cell(row=ind, column=offset + 10, value=requisicao['guias__telefone_unidade'])
            ws.cell(row=ind, column=offset + 11, value=f'{requisicao["guias__endereco_unidade"]} '
                                                       f'{requisicao["guias__numero_unidade"]}')
            ws.cell(row=ind, column=offset + 12, value=requisicao['guias__bairro_unidade'])
            ws.cell(row=ind, column=offset + 13, value=requisicao['guias__cep_unidade'])
            ws.cell(row=ind, column=offset + 14, value=requisicao['guias__contato_unidade'])
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=offset + 15,
                        value=f'{requisicao["guias__alimentos__embalagens__qtd_volume"]} '
                              f'{requisicao["guias__alimentos__embalagens__descricao_embalagem"]}')
                ws.cell(row=ind, column=offset + 16,
                        value=f'{requisicao["guias__alimentos__embalagens__capacidade_embalagem"]} '
                              f'{requisicao["guias__alimentos__embalagens__unidade_medida"]}')
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FRACIONADA':
                ws.cell(row=ind, column=offset + 17,
                        value=f'{requisicao["guias__alimentos__embalagens__qtd_volume"]} '
                              f'{requisicao["guias__alimentos__embalagens__descricao_embalagem"]}')
                ws.cell(row=ind, column=offset + 18,
                        value=f'{requisicao["guias__alimentos__embalagens__capacidade_embalagem"]} '
                              f'{requisicao["guias__alimentos__embalagens__unidade_medida"]}')
            if 'primeira_conferencia' in requisicao:
                ws.cell(row=ind, column=offset + 19, value=f'{requisicao["primeira_conferencia"].data_recebimento} '
                                                           f'{requisicao["primeira_conferencia"].hora_recebimento}')
                ws.cell(row=ind, column=offset + 20, value=requisicao['primeira_conferencia'].criado_por.nome)
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                ws.cell(row=ind, column=offset + 21, value=qtd_recebida)
                ws.cell(row=ind, column=offset + 22, value=requisicao['guias__alimentos__embalagens__qtd_a_receber'])
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FRACIONADA':
                ws.cell(row=ind, column=offset + 23, value=qtd_recebida)
                ws.cell(row=ind, column=offset + 24, value=requisicao['guias__alimentos__embalagens__qtd_a_receber'])
            if 'primeira_conferencia' in requisicao:
                ws.cell(row=ind, column=offset + 25, value=requisicao['primeira_conferencia'].nome_motorista)
                ws.cell(row=ind, column=offset + 26, value=requisicao['primeira_conferencia'].placa_veiculo)
                if 'conferencia_alimento' in requisicao:
                    ws.cell(row=ind, column=offset + 27, value=retorna_ocorrencias_alimento(
                        requisicao['conferencia_alimento'].ocorrencia
                    ))
                    ws.cell(row=ind, column=offset + 28, value=requisicao['conferencia_alimento'].observacao)
            if 'primeira_reposicao' in requisicao:
                ws.cell(row=ind, column=offset + 29, value=f'{requisicao["primeira_reposicao"].data_recebimento} '
                                                           f'{requisicao["primeira_reposicao"].hora_recebimento}')
                ws.cell(row=ind, column=offset + 30,
                        value=f'{requisicao["primeira_reposicao"].criado_em.strftime("%d/%m/%Y")} '
                              f'{requisicao["primeira_reposicao"].criado_em.strftime("%H:%M:%S")}')
            if 'primeira_reposicao' in requisicao:
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=offset + 31, value=requisicao['primeira_reposicao'].criado_por.nome)
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FECHADA':
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=offset + 32, value=requisicao['reposicao_alimento'].qtd_recebido)
            if requisicao['guias__alimentos__embalagens__tipo_embalagem'] == 'FRACIONADA':
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=offset + 33, value=requisicao['reposicao_alimento'].qtd_recebido)
            if 'primeira_reposicao' in requisicao:
                ws.cell(row=ind, column=offset + 34, value=requisicao['primeira_reposicao'].nome_motorista)
                ws.cell(row=ind, column=offset + 35, value=requisicao['primeira_reposicao'].placa_veiculo)
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=offset + 36, value=retorna_status_alimento(
                        requisicao['reposicao_alimento'].status_alimento))
                if 'reposicao_alimento' in requisicao:
                    ws.cell(row=ind, column=offset + 37, value=retorna_ocorrencias_alimento(
                        requisicao['reposicao_alimento'].ocorrencia
                    ))
                    ws.cell(row=ind, column=offset + 38, value=requisicao['reposicao_alimento'].observacao)

        cls.aplicar_estilo_padrao(ws, count_data, count_fields)

    @classmethod  # noqa C901
    def exportar_entregas(cls, requisicoes, requisicoes_insucesso, perfil, tem_conferencia,
                          tem_insucesso, is_async=False):
        wb = Workbook()
        ws_conferencia = wb.active
        if perfil == 'DISTRIBUIDOR':
            conferencia = cls.cria_aba_conferencia_distribuidor(ws_conferencia, requisicoes)
        elif perfil == 'DRE':
            conferencia = cls.cria_aba_conferencia_dre(ws_conferencia, requisicoes)
        else:
            conferencia = cls.cria_aba_conferencia_dilog(ws_conferencia, requisicoes)

        if tem_conferencia and not tem_insucesso:
            conferencia
        elif tem_insucesso and not tem_conferencia:
            ws_insucesso = wb.active
            cls.cria_aba_insucesso(ws_insucesso, requisicoes_insucesso, perfil)
        else:
            conferencia

            ws_insucesso = wb.create_sheet('Relatório de Insucesso')
            cls.cria_aba_insucesso(ws_insucesso, requisicoes_insucesso, perfil)

        arquivo = cls.gera_arquivo(wb)

        return arquivo if is_async else {'arquivo': arquivo}
