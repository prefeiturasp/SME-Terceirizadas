import logging
from datetime import date
from tempfile import NamedTemporaryFile

from django.core.files import File
from django.db import transaction
from django.db.models import Q
from openpyxl import Workbook, load_workbook, styles

from sme_terceirizadas.dados_comuns.constants import DJANGO_ADMIN_TREINAMENTO_PASSWORD
from sme_terceirizadas.dieta_especial.models import ArquivoCargaUsuariosEscola
from sme_terceirizadas.escola.models import Escola
from sme_terceirizadas.perfil.models.perfil import Perfil, Vinculo
from sme_terceirizadas.perfil.models.usuario import Usuario

from .schemas import ArquivoCargaUsuariosDiretorSchema

logger = logging.getLogger('sigpae.importa_usuarios_escola_diretor')


class ProcessadorPlanilha:
    def __init__(self, usuario: Usuario, arquivo: ArquivoCargaUsuariosEscola) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.erros = []
        self.worksheet = self.abre_worksheet()

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def processamento(self):  # noqa C901
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        logger.info(f'Quantidade de linhas: {len(linhas)} - Quantidade colunas {len(linhas[0])}')

        for ind, linha in enumerate(linhas[1:], 2):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)
                usuario_schema = ArquivoCargaUsuariosDiretorSchema(**dicionario_dados)

                logger.info(f'USUARIO: {usuario_schema}\n')
                self.cria_usuario(ind, usuario_schema)
            except Exception as exc:
                self.erros.append(f'Linha {ind} - {exc}')

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def validacao_inicial(self) -> bool:
        return (self.existe_conteudo() and self.extensao_do_arquivo_esta_correta())

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = 'Não foi feito o upload da planilha'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def extensao_do_arquivo_esta_correta(self) -> bool:
        if not self.path.endswith('.xlsx'):
            self.arquivo.log = 'Planilha precisa ter extensão .xlsx'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {key: linha[index].value
                for index, key in enumerate(ArquivoCargaUsuariosDiretorSchema.schema()['properties'].keys())}

    def cria_usuario(self, ind, usuario_schema: ArquivoCargaUsuariosDiretorSchema):  # noqa C901
        try:
            self.__criar_usuario_diretor(usuario_schema)
        except Exception as exd:
            self.erros.append(f'Linha {ind} - {exd}')

        try:
            self.__criar_usuario_assitente(usuario_schema)
        except Exception as exa:
            self.erros.append(f'Linha {ind} - {exa}')

    @transaction.atomic
    def __criar_usuario_diretor(self, usuario_schema: ArquivoCargaUsuariosDiretorSchema):
        escola = self.consulta_escola(usuario_schema.codigo_eol_escola)
        self.checa_usuario(usuario_schema.rf_diretor, usuario_schema.email_diretor, usuario_schema.nome_diretor)
        self.__criar_usuario(escola, usuario_schema.nome_diretor, usuario_schema.email_diretor, 'DIRETOR')

    @transaction.atomic
    def __criar_usuario_assitente(self, usuario_schema: ArquivoCargaUsuariosDiretorSchema):
        escola = self.consulta_escola(usuario_schema.codigo_eol_escola)
        self.checa_usuario(usuario_schema.rf_assistente, usuario_schema.email_assistente,
                           usuario_schema.nome_assistente)
        self.__criar_usuario(escola, usuario_schema.nome_assistente, usuario_schema.email_assistente, 'DIRETOR')

    def consulta_escola(self, codigo_eol_escola):
        escola = Escola.objects.filter(codigo_eol=codigo_eol_escola).first()
        if not escola:
            raise Exception(f'Escola não encontrada para o código: {codigo_eol_escola}.')
        return escola

    def checa_usuario(self, registro_funcional, email, nome_planilha):
        usuario = Usuario.objects.filter(Q(registro_funcional=registro_funcional) | Q(email=email)).first()
        if usuario:
            raise Exception(f'Usuário pra esse email {email} ou com esse rf '
                            + f'{registro_funcional} já criado: Nome na planilha {nome_planilha}.')

    def __criar_usuario(self, escola, nome_usuario, email, cargo):
        perfil = Perfil.objects.get(nome=cargo)
        data_atual = date.today()
        usuario = Usuario.objects.create_superuser(
            email=email.strip(),
            password=DJANGO_ADMIN_TREINAMENTO_PASSWORD,
            nome=nome_usuario.strip(),
            cargo=cargo,
        )
        Vinculo.objects.create(
            instituicao=escola,
            perfil=perfil,
            usuario=usuario,
            data_inicial=data_atual,
            ativo=True,
        )

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = '\n'.join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
        else:
            self.arquivo.log = 'Planilha processada com sucesso.'
            self.arquivo.processamento_com_sucesso()

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = 'Erros'
        cabecalho = ws.cell(row=1, column=1, value='Erros encontrados no processamento da planilha')
        cabecalho.fill = styles.PatternFill('solid', fgColor='808080')
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f'arquivo_resultado_{self.arquivo.pk}.xlsx'
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))


def importa_usuarios_escola(usuario: Usuario, arquivo: ArquivoCargaUsuariosEscola) -> None:
    logger.debug(f'Iniciando o processamento do arquivo: {arquivo.uuid}')

    try:
        processador = ProcessadorPlanilha(usuario, arquivo)
        processador.processamento()
        processador.finaliza_processamento()
    except Exception as exc:
        logger.error(f'Erro genérico: {exc}')
