import logging
from datetime import date
from tempfile import NamedTemporaryFile
from typing import List

from django.core.files import File
from django.db import transaction
from django.db.models import Q
from openpyxl import Workbook, load_workbook, styles

from sme_terceirizadas.dados_comuns.constants import DJANGO_ADMIN_PASSWORD
from sme_terceirizadas.dieta_especial.models import (
    AlergiaIntolerancia,
    ArquivoCargaDietaEspecial,
    ClassificacaoDieta,
    SolicitacaoDietaEspecial
)
from sme_terceirizadas.escola.models import Aluno, Escola
from sme_terceirizadas.perfil.models.perfil import Perfil, Vinculo
from sme_terceirizadas.perfil.models.usuario import Usuario

from .schemas import ArquivoCargaDietaEspecialSchema

logger = logging.getLogger('sigpae.importa_dietas_especiais')


class ProcessadorPlanilha:
    def __init__(self, usuario: Usuario, arquivo: ArquivoCargaDietaEspecial) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.erros = []
        if self.arquivo.conteudo:
            self.worksheet = self.abre_worksheet()

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def processamento(self):  # noqa C901
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        logger.info(f'Quantidade de linhas: {len(linhas)} - Quantidade colunas {len(linhas[0])}')
        if self.worksheet.max_column != 12:
            self.erros.append('Erro: O número de colunas diferente das estrutura definida.')
            return

        for ind, linha in enumerate(linhas[1:], 2):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)

                # Faz validações dos campos usando o schema para a dieta
                solicitacao_dieta_schema = ArquivoCargaDietaEspecialSchema(**dicionario_dados)

                aluno = self.consulta_aluno(solicitacao_dieta_schema)

                escola = self.consulta_escola(solicitacao_dieta_schema)

                classificacao_dieta = self.consulta_classificacao(solicitacao_dieta_schema)

                diagnosticos = self.monta_diagnosticos(solicitacao_dieta_schema.codigo_diagnostico)

                self.checa_existencia_solicitacao(solicitacao_dieta_schema, aluno)
                self.cria_solicitacao(solicitacao_dieta_schema, aluno, classificacao_dieta, diagnosticos, escola)
            except Exception as exc:
                self.erros.append(f'Linha {ind} - {exc}')

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def validacao_inicial(self) -> bool:
        return (self.existe_conteudo() and self.extensao_do_arquivo_esta_correta())

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = 'Não foi feito o upload da planilha'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def extensao_do_arquivo_esta_correta(self) -> bool:
        if not self.path.endswith('.xlsx'):
            self.arquivo.log = 'Planilha precisa ter extensão .xlsx'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {key: linha[index].value
                for index, key in enumerate(ArquivoCargaDietaEspecialSchema.schema()['properties'].keys())}

    def consulta_aluno(self, solicitacao_dieta_schema) -> Aluno:
        aluno = Aluno.objects.filter(codigo_eol=solicitacao_dieta_schema.codigo_eol_aluno).first()
        if not aluno:
            raise Exception(f'Erro: Aluno com código eol {solicitacao_dieta_schema.codigo_eol_aluno} não encontrado.')

        if solicitacao_dieta_schema.nome_aluno.upper() != aluno.nome:
            raise Exception(f'Erro: Nome divergente para o código eol {aluno.codigo_eol}: '
                            + 'Nome aluno planilha: '
                            + f'{solicitacao_dieta_schema.nome_aluno.upper()} != Nome aluno sistema: {aluno.nome}')
        return aluno

    def consulta_escola(self, solicitacao_dieta_schema) -> Escola:
        escola = Escola.objects.filter(codigo_codae=solicitacao_dieta_schema.codigo_escola).first()
        if not escola:
            raise Exception(f'Erro: escola com código codae {solicitacao_dieta_schema.codigo_escola} não encontrada.')
        return escola

    def consulta_classificacao(self, dieta_schema) -> ClassificacaoDieta:
        classificacao_dieta = ClassificacaoDieta.objects.filter(
            nome=f'Tipo {dieta_schema.codigo_categoria_dieta}').first()
        if not classificacao_dieta:
            raise Exception(f'Erro: A categoria da dieta {dieta_schema.codigo_categoria_dieta} não encontrado.')
        return classificacao_dieta

    def monta_diagnosticos(self, codigo_diagnostico: str) -> List[AlergiaIntolerancia]:
        lista_nomes_diagnosticos = codigo_diagnostico.split(';')
        lista_diagnosticos = []
        for nome_diagnostico in lista_nomes_diagnosticos:
            # Checando se existe o diagnóstico tanto pelo nome que veio na planilha quanto pelo nome
            # em uppercase
            diag = AlergiaIntolerancia.objects.filter(
                Q(descricao=nome_diagnostico) | Q(descricao=nome_diagnostico.upper())).first()
            if not diag:
                diag = AlergiaIntolerancia.objects.create(descricao=nome_diagnostico.upper())
            else:
                diag.descricao = diag.descricao.upper()
                diag.save()
            lista_diagnosticos.append(diag)

        return lista_diagnosticos

    def consulta_relacao_lote_terceirizada(self, solicitacao) -> None:
        if solicitacao.rastro_terceirizada is None:
            raise Exception(f'Lote "{solicitacao.rastro_lote.nome}" relacionado à escola '
                            + f'"{solicitacao.rastro_escola.nome}" não está vinculado à uma terceirizada.')

    def eh_exatamente_mesma_solicitacao(self, solicitacao, solicitacao_dieta_schema):
        if (solicitacao.escola_destino.codigo_codae == solicitacao_dieta_schema.codigo_escola and
            solicitacao.aluno.codigo_eol == solicitacao_dieta_schema.codigo_eol_aluno and
            solicitacao.nome_protocolo.upper() == solicitacao_dieta_schema.protocolo_dieta.upper() and
            solicitacao_dieta_schema.codigo_categoria_dieta in solicitacao.classificacao.nome and
                [x.upper() for x in solicitacao_dieta_schema.codigo_diagnostico.split(';')] == [
                x.descricao.upper() for x in solicitacao.alergias_intolerancias.all()]):
            raise Exception('Erro: Já existe uma solicitação ativa que foi importada para o aluno com código eol: '
                            + f'{solicitacao_dieta_schema.codigo_eol_aluno} exatamente igual a esta')

    def checa_existencia_solicitacao(self, solicitacao_dieta_schema, aluno) -> None:
        if SolicitacaoDietaEspecial.objects.filter(aluno=aluno, ativo=True, eh_importado=True).exists():
            solicitacao = SolicitacaoDietaEspecial.objects.get(aluno=aluno, ativo=True, eh_importado=True)
            self.eh_exatamente_mesma_solicitacao(solicitacao, solicitacao_dieta_schema)
        if SolicitacaoDietaEspecial.objects.filter(aluno=aluno, ativo=True, eh_importado=False).exists():
            SolicitacaoDietaEspecial.objects.filter(aluno=aluno, ativo=True, eh_importado=False).update(ativo=False)

    @transaction.atomic
    def cria_solicitacao(self, solicitacao_dieta_schema, aluno, classificacao_dieta, diagnosticos, escola):  # noqa C901
        observacoes = """Essa Dieta Especial foi autorizada anteriormente a implantação do SIGPAE.
        Para ter acesso ao Protocolo da Dieta Especial,
        entre em contato com o Núcleo de Dieta Especial através do e-mail:
        smecodaedietaespecial@sme.prefeitura.sp.gov.br."""

        email_fake = f'fake{escola.id}@admin.com'
        usuario_escola = Usuario.objects.filter(email=email_fake).first()
        if not usuario_escola:
            perfil = Perfil.objects.get(nome='DIRETOR_UE')
            data_atual = date.today()
            # Esse Usuário não consegue acessar o sistema
            # Troquei a senha pra reforçar que essa não funciona
            # já que o usuário não tem acesso ao sistema
            usuario_escola = Usuario.objects.create(
                email=email_fake,
                password=DJANGO_ADMIN_PASSWORD,
                nome=escola.nome,
                cargo='DIRETOR',
                is_active=False
            )
            Vinculo.objects.create(
                instituicao=escola,
                perfil=perfil,
                usuario=usuario_escola,
                data_inicial=data_atual,
                ativo=True
            )

        solicitacao: SolicitacaoDietaEspecial = SolicitacaoDietaEspecial.objects.create(
            criado_por=usuario_escola,
            aluno=aluno,
            escola_destino=escola,
            ativo=True,
            nome_protocolo=solicitacao_dieta_schema.protocolo_dieta.upper(),
            classificacao=classificacao_dieta,
            observacoes=observacoes,
            conferido=True,
            eh_importado=True
        )
        solicitacao.inicia_fluxo(user=usuario_escola)
        solicitacao.alergias_intolerancias.add(*diagnosticos)
        self.consulta_relacao_lote_terceirizada(solicitacao)
        solicitacao.save()
        solicitacao.codae_autoriza(user=self.usuario, eh_importacao=True)

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = '\n'.join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
        else:
            self.arquivo.log = 'Planilha processada com sucesso.'
            self.arquivo.processamento_com_sucesso()

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = 'Erros'
        cabecalho = ws.cell(row=1, column=1, value='Erros encontrados no processamento da planilha')
        cabecalho.fill = styles.PatternFill('solid', fgColor='808080')
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f'arquivo_resultado_{self.arquivo.pk}.xlsx'
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))


def importa_dietas_especiais(usuario: Usuario, arquivo: ArquivoCargaDietaEspecial) -> None:
    logger.debug(f'Iniciando o processamento do arquivo: {arquivo.uuid}')

    try:
        processador = ProcessadorPlanilha(usuario, arquivo)
        processador.processamento()
        processador.finaliza_processamento()
    except Exception as exc:
        logger.error(f'Erro genérico: {exc}')
