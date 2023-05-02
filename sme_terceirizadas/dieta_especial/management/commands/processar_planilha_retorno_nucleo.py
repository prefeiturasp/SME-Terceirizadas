from django.core.management.base import BaseCommand
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from sme_terceirizadas.dieta_especial.models import ProtocoloPadraoDietaEspecial, SolicitacaoDietaEspecial
from sme_terceirizadas.terceirizada.models import Edital


class Command(BaseCommand):
    help = """
    Processar planilhas de devolutiva das POs para Dietas Especiais com protocolos válidos
    """

    def extrair_dados_planilha(self):
        active_sheet = load_workbook('retorno-nucleo-dieta.xlsx').active
        linhas = list(active_sheet.rows)
        lista_objetos = []
        for linha in enumerate(linhas[1:]):
            objeto = {}
            for c, coluna in enumerate(linha[1]):
                objeto[linhas[0][c].value] = coluna.value
            lista_objetos.append(objeto)
        return lista_objetos

    def formatar_tamanho_celulas(self, ws):
        for column in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[column].width = 50

    def exportar_planilha(self, lista):
        wb = Workbook()
        ws = wb.active
        self.formatar_tamanho_celulas(ws)
        ws.title = 'Dietas não relacionadas'
        cabecalho = ['uuid', 'escola', 'aluno', 'lote', 'edital', 'nome_protocolo']
        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = str(title)
            celula.font = Font(size='13', bold=True)

        for ind, dict_solicitacao in enumerate(lista, 2):
            ws.cell(row=ind, column=1, value=str(dict_solicitacao['uuid']))
            ws.cell(row=ind, column=2, value=dict_solicitacao['escola'])
            ws.cell(row=ind, column=3, value=dict_solicitacao['aluno'])
            ws.cell(row=ind, column=4, value=dict_solicitacao['lote'])
            ws.cell(row=ind, column=5, value=dict_solicitacao['edital'])
            ws.cell(row=ind, column=6, value=dict_solicitacao['nome_protocolo'])

        wb.save(f'relacao-planilha-nao-relacionados.xlsx')

    def handle(self, *args, **options):
        dados_planilha = self.extrair_dados_planilha()
        solicitacoes_nao_relacionadas = []
        for dado in dados_planilha:
            solicitacao = SolicitacaoDietaEspecial.objects.filter(uuid=dado['uuid'])
            if solicitacao:
                solicitacao = solicitacao.first()
                editais = Edital.objects.filter(numero=dado['editais'])
                protocolos_uuids = editais.values_list('protocolos_padroes_dieta_especial__uuid')
                nome_protocolo = dado[None]
                protocolos = ProtocoloPadraoDietaEspecial.objects.filter(uuid__in=protocolos_uuids,
                                                                         nome_protocolo=nome_protocolo)
                if not protocolos:
                    objeto = {'uuid': dado['uuid'], 'escola': dado['escola'], 'aluno': dado['aluno'],
                              'lote': dado['lote'], 'edital': dado['editais'], 'nome_protocolo': dado[None]}
                    solicitacoes_nao_relacionadas.append(objeto)
                else:
                    solicitacao.protocolo_padrao = protocolos.first()
                    solicitacao.save()
        self.exportar_planilha(solicitacoes_nao_relacionadas)
