from django.core.management.base import BaseCommand
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from sme_terceirizadas.dieta_especial.models import ProtocoloPadraoDietaEspecial, SolicitacaoDietaEspecial
from sme_terceirizadas.terceirizada.models import Edital


class Command(BaseCommand):
    help = """
    Processar planilhas de devolutiva das POs para Dietas Especiais com protocolos válidos
    """

    def extrair_dados_planilha(self):
        active_sheet = load_workbook('validos.xlsx').active
        linhas = list(active_sheet.rows)
        lista_objetos = []
        for linha in enumerate(linhas[2:]):
            objeto = {}
            for c, coluna in enumerate(linha[1]):
                objeto[linhas[0][c].value] = coluna.value
            lista_objetos.append(objeto)
        return lista_objetos

    def formatar_tamanho_celulas(self, ws):
        for column in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[column].width = 50

    def exportar_planilha(self, lista):
        wb = Workbook()
        ws = wb.active
        self.formatar_tamanho_celulas(ws)
        ws.title = 'Dietas não relacionadas'
        cabecalho = ['uuid', 'nome protocolo na planilha', 'escola', 'aluno', 'lote', 'erro']
        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = str(title)
            celula.font = Font(size='13', bold=True)

        for ind, dict_solicitacao in enumerate(lista, 2):
            ws.cell(row=ind, column=1, value=str(dict_solicitacao['uuid']))
            ws.cell(row=ind, column=2, value=dict_solicitacao['nome protocolo na planilha'])
            ws.cell(row=ind, column=3, value=dict_solicitacao['escola'])
            ws.cell(row=ind, column=4, value=dict_solicitacao['aluno'])
            ws.cell(row=ind, column=5, value=dict_solicitacao['lote'])
            ws.cell(row=ind, column=6, value=dict_solicitacao['erro'])

        wb.save(f'relacao-planilha-validos-protocolos-com-erro.xlsx')

    def handle(self, *args, **options):
        dados_planilha = self.extrair_dados_planilha()
        solicitacoes_nao_relacionadas = []
        for dado in dados_planilha:
            solicitacao = SolicitacaoDietaEspecial.objects.filter(uuid=dado['UUID'])
            if solicitacao:
                solicitacao = solicitacao.first()
                editais_uuids = solicitacao.escola.lote.contratos_do_lote.values_list('edital__uuid', flat=True)
                editais = Edital.objects.filter(uuid__in=editais_uuids)
                protocolos_uuids = editais.values_list('protocolos_padroes_dieta_especial__uuid')
                nome_protocolo = dado['NOME PROTOCOLO NO DB']
                if nome_protocolo == 'ALERGIA - CORANTES':
                    nome_protocolo = 'ALERGIA - CORANTE'

                protocolos = ProtocoloPadraoDietaEspecial.objects.filter(uuid__in=protocolos_uuids,
                                                                         nome_protocolo=nome_protocolo)
                if not protocolos:
                    objeto = {'uuid': dado['UUID'], 'nome protocolo na planilha': nome_protocolo,
                              'escola': solicitacao.escola.nome, 'aluno': solicitacao.aluno.nome,
                              'lote': solicitacao.escola.lote.nome,
                              'erro': 'Protocolo não encontrado para o Edital e Lote da Escola '}
                    solicitacoes_nao_relacionadas.append(objeto)
                else:
                    solicitacao.protocolo_padrao = protocolos.first()
                    solicitacao.save()
        self.exportar_planilha(solicitacoes_nao_relacionadas)
