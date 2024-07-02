from django.http import HttpResponse
from openpyxl import Workbook, styles
from openpyxl.worksheet.datavalidation import DataValidation
from itertools import groupby
from operator import attrgetter
from datetime import datetime
from sme_terceirizadas.dados_comuns.helper_planilha_modelo import (
    cria_validacao_lista_em_sheet_oculto,
)
from sme_terceirizadas.imr.models import (
    CategoriaOcorrencia,
    TipoGravidade,
    TipoPenalidade,
)
from sme_terceirizadas.terceirizada.models import Edital


def exportar_planilha_importacao_tipos_penalidade(request, **kwargs):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=planilha_importacao_tipos_penalidade.xlsx"
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = "Tipos de Penalidade"
    headers = [
        "Edital",
        "Número da Cláusula/Item",
        "Gravidade",
        "Obrigações (separadas por ;)",
        "Descrição da Cláusula/Item",
        "Status",
    ]
    sheet = workbook["Tipos de Penalidade"]
    for column in "ABCDEFGF":
        sheet.column_dimensions[column].width = 30
    _font = styles.Font(name="Calibri", sz=10)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill("solid", fgColor="ffff99")
        cabecalho.font = styles.Font(name="Calibri", size=10, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style="thin", color="000000"),
            right=styles.Side(border_style="thin", color="000000"),
            top=styles.Side(border_style="thin", color="000000"),
            bottom=styles.Side(border_style="thin", color="000000"),
        )

    hidden_sheet = workbook.create_sheet(title="HiddenSheet")
    hidden_sheet.sheet_state = "hidden"

    editais = [edital.numero for edital in Edital.objects.all()]
    ws = cria_validacao_lista_em_sheet_oculto(
        editais, hidden_sheet, ws, workbook, "A", 1, "Edital", "o", "A"
    )

    tipos_gravidade = ", ".join(
        [gravidade.tipo for gravidade in TipoGravidade.objects.all()]
    )
    dv2 = DataValidation(type="list", formula1=f'"{tipos_gravidade}"', allow_blank=True)
    dv2.error = "Tipo de Gravidade Inválido"
    dv2.errorTitle = "Tipo de Gravidade não permitido"
    ws.add_data_validation(dv2)
    dv2.add("C2:C1048576")

    dv3 = DataValidation(type="list", formula1='"Ativo,Inativo"', allow_blank=True)
    dv3.error = "Status Inválido"
    dv3.errorTitle = "Status não permitido"
    ws.add_data_validation(dv3)
    dv3.add("F2:F1048576")

    workbook.save(response)

    return response


def exportar_planilha_importacao_tipos_ocorrencia(request, **kwargs):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename=planilha_importacao_tipos_ocorrencia.xlsx"
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = "Tipos de Ocorrência"
    headers = [
        "Posição",
        "Perfis",
        "Edital",
        "Categoria da Ocorrência",
        "Título",
        "Descrição",
        "Penalidade",
        "É IMR?",
        "Pontuação (IMR)",
        "Tolerância",
        "% de Desconto",
        "Status",
        "Aceita múltiplas respostas?",
    ]
    sheet = workbook["Tipos de Ocorrência"]
    for column in "ABCHIJKL":
        sheet.column_dimensions[column].width = 20
    for column in "EFM":
        sheet.column_dimensions[column].width = 35
    for column in "DG":
        sheet.column_dimensions[column].width = 50
    _font = styles.Font(name="Calibri", sz=10)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill("solid", fgColor="ffff99")
        cabecalho.font = styles.Font(name="Calibri", size=10, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style="thin", color="000000"),
            right=styles.Side(border_style="thin", color="000000"),
            top=styles.Side(border_style="thin", color="000000"),
            bottom=styles.Side(border_style="thin", color="000000"),
        )

    hidden_sheet = workbook.create_sheet(title="HiddenSheet")
    hidden_sheet.sheet_state = "hidden"

    dv = DataValidation(
        type="list",
        formula1='"DIRETOR,SUPERVISAO,DIRETOR E SUPERVISAO"',
        allow_blank=True,
    )
    dv.error = "Perfil Inválido"
    dv.errorTitle = "Perfil não permitido"
    ws.add_data_validation(dv)
    dv.add("B2:B1048576")

    editais = [edital.numero for edital in Edital.objects.all()]
    ws = cria_validacao_lista_em_sheet_oculto(
        editais, hidden_sheet, ws, workbook, "A", 1, "Edital", "o", "C"
    )

    categorias = [categoria.nome for categoria in CategoriaOcorrencia.objects.all()]
    ws = cria_validacao_lista_em_sheet_oculto(
        categorias, hidden_sheet, ws, workbook, "B", 2, "Categoria", "a", "D"
    )

    penalidades = [
        f"{penalidade.edital.numero} - {penalidade.numero_clausula}"
        for penalidade in TipoPenalidade.objects.all()
    ]
    ws = cria_validacao_lista_em_sheet_oculto(
        penalidades, hidden_sheet, ws, workbook, "C", 3, "Penalidade", "a", "G"
    )

    dv4 = DataValidation(type="list", formula1='"SIM,NÃO"', allow_blank=True)
    dv4.error = "Opção Inválida"
    dv4.errorTitle = "Opção não permitida"
    ws.add_data_validation(dv4)
    dv4.add("H2:H1048576")

    dv5 = DataValidation(type="list", formula1='"Ativo,Inativo"', allow_blank=True)
    dv5.error = "Status Inválido"
    dv5.errorTitle = "Status não permitido"
    ws.add_data_validation(dv5)
    dv5.add("L2:L1048576")

    dv6 = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv6.error = "Valor Inválido"
    dv6.errorTitle = "Valor não permitido"
    ws.add_data_validation(dv6)
    dv6.add("M2:M1048576")

    workbook.save(response)

    return response


class RelatorioNotificacaoService:

    def __init__(self, formulario_supervisao):
        self.categoria = "QUANTIDADE/QUALIDADE DE UTENSÍLIOS/MOBILIÁRIOS/EQUIPAMENTOS"
        self.formulario_supervisao = formulario_supervisao

    def get_lista_respostas(self):
        respostas = self.formulario_supervisao.formulario_base.buscar_respostas(categoria=self.categoria)
        respostas.sort(key=attrgetter('parametrizacao.tipo_ocorrencia', 'grupo'))
        grouped = groupby(respostas, key=attrgetter('parametrizacao.tipo_ocorrencia', 'grupo'))
        lista_respostas = []

        for _, group in grouped:
            _respostas = []
            for resposta in group:
                _respostas.append(resposta)

            _respostas_sorted = sorted(_respostas, key=lambda x: int(x.parametrizacao.posicao))
            lista_respostas.append(_respostas_sorted)
        return lista_respostas

    def retornar_dados_formatados(self):
        dados = {
            'diretoria_regional': self.formulario_supervisao.escola.diretoria_regional.nome,
            'unidade': self.formulario_supervisao.escola.nome,
            'maior_frequencia_no_periodo': self.formulario_supervisao.maior_frequencia_no_periodo,
            'total_matriculados_por_data': self.formulario_supervisao.escola.quantidade_alunos_matriculados_por_data(self.formulario_supervisao.formulario_base.data),
            'data_visita': self.formulario_supervisao.formulario_base.data.strftime("%d/%m/%Y"),
            'usuario': self.formulario_supervisao.formulario_base.usuario,
            'lote': self.formulario_supervisao.escola.lote.nome,
            'terceirizada': self.formulario_supervisao.escola.lote.terceirizada.nome_fantasia,
            'edital': self.formulario_supervisao.escola.edital if self.formulario_supervisao.escola.edital else "-",
            'respostas': self.get_lista_respostas(),
            'data_geracao': datetime.now().strftime("%d/%m/%Y"),
            'hora_geracao': datetime.now().strftime("%H:%M:%S"),
        }
        return dados
