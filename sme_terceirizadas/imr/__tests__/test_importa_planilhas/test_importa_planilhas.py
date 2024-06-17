from unittest.mock import Mock

import pytest
from django.contrib.admin import AdminSite
from django.core.files.uploadedfile import SimpleUploadedFile
from django.http import HttpResponse
from openpyxl.reader.excel import load_workbook

from sme_terceirizadas.imr.admin import ImportacaoPlanilhaTipoPenalidadeAdmin
from sme_terceirizadas.imr.models import (
    ImportacaoPlanilhaTipoPenalidade,
    TipoPenalidade,
)

pytestmark = pytest.mark.django_db


def test_importa_planilha_tipo_penalidade(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    edital_factory,
    tipo_gravidade_factory,
    importacao_planilha_tipo_penalidade_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    edital = edital_factory.create(numero="EDITAL MODELO IMR")
    gravidade = tipo_gravidade_factory.create(tipo="Leve")

    workbook = load_workbook(
        "sme_terceirizadas/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_penalidade.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file = SimpleUploadedFile(
        name="teste_carga_tipos_penalidades.xlsx",
        content=response.content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    importacao_planilha_tipo_penalidade_factory.create(conteudo=file)

    queryset = ImportacaoPlanilhaTipoPenalidade.objects.all()

    importacao_planilha_tipo_penalidade_admin = ImportacaoPlanilhaTipoPenalidadeAdmin(
        model=ImportacaoPlanilhaTipoPenalidade, admin_site=AdminSite()
    )
    importacao_planilha_tipo_penalidade_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )
    assert TipoPenalidade.objects.count() == 1
    tipo_penalidade = TipoPenalidade.objects.get()
    assert tipo_penalidade.edital == edital
    assert tipo_penalidade.gravidade == gravidade
    assert tipo_penalidade.obrigacoes.count() == 3
    assert tipo_penalidade.descricao == "teste descrição"
    assert tipo_penalidade.status is True


def test_importa_planilha_tipo_penalidade_erro_seleciona_mais_de_um_arquivo(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    edital_factory,
    tipo_gravidade_factory,
    importacao_planilha_tipo_penalidade_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    edital_factory.create(numero="EDITAL MODELO IMR")
    tipo_gravidade_factory.create(tipo="Leve")

    workbook = load_workbook(
        "sme_terceirizadas/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_penalidade.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file = SimpleUploadedFile(
        name="teste_carga_tipos_penalidades.xlsx",
        content=response.content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    importacao_planilha_tipo_penalidade_factory.create(conteudo=file)
    importacao_planilha_tipo_penalidade_factory.create(conteudo=file)

    queryset = ImportacaoPlanilhaTipoPenalidade.objects.all()

    importacao_planilha_tipo_penalidade_admin = ImportacaoPlanilhaTipoPenalidadeAdmin(
        model=ImportacaoPlanilhaTipoPenalidade, admin_site=AdminSite()
    )
    importacao_planilha_tipo_penalidade_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )

    # não criou porque selecionou mais de um arquivo
    assert TipoPenalidade.objects.count() == 0


def test_importa_planilha_tipo_penalidade_erro_arquivo_formato_incorreto(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    edital_factory,
    tipo_gravidade_factory,
    importacao_planilha_tipo_penalidade_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    edital_factory.create(numero="EDITAL MODELO IMR")
    tipo_gravidade_factory.create(tipo="Leve")

    workbook = load_workbook(
        "sme_terceirizadas/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_penalidade.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file_pdf = SimpleUploadedFile(
        name="teste_carga_tipos_penalidades.pdf",
        content=response.content,
        content_type="application/pdf",
    )

    importacao_planilha_tipo_penalidade_factory.create(conteudo=file_pdf)

    queryset = ImportacaoPlanilhaTipoPenalidade.objects.all()

    importacao_planilha_tipo_penalidade_admin = ImportacaoPlanilhaTipoPenalidadeAdmin(
        model=ImportacaoPlanilhaTipoPenalidade, admin_site=AdminSite()
    )
    importacao_planilha_tipo_penalidade_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )

    # não criou porque o formato estava incorreto (PDF)
    assert TipoPenalidade.objects.count() == 0
