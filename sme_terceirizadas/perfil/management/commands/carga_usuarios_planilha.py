import re
from datetime import date

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from sme_terceirizadas.escola.models import Codae
from sme_terceirizadas.perfil.models import Perfil, Usuario, Vinculo


class Command(BaseCommand):
    help = 'Carga de dados para criação de usuários com perfil nutrisupervisão.'

    def handle(self, *args, **kwrgs):
        self.stdout.write('Lendo dados da planilha.')
        self.preparar_dados('Dados pessoais - Nutricionistas Supervisao.xlsx')
        self.stdout.write('Processo de carga de dados finalizado.')

    def preparar_dados(self, nome_arquivo):
        wb = load_workbook(nome_arquivo)
        sheets = [
            {'nome': 'Planilha1', 'inicio': 3},
            {'nome': 'Supervisão interna', 'inicio': 5}
        ]
        codae = Codae.objects.get(nome='CODAE - SUPERVISÃO DE NUTRIÇÃO')
        perfil = Perfil.objects.get(nome='COORDENADOR_SUPERVISAO_NUTRICAO')
        data_atual = date.today()

        for sheet in sheets:
            sheet_ranges = wb[sheet['nome']]
            linhas = sheet_ranges.max_row

            for i in range(sheet['inicio'], linhas + 1):
                usuario = self.criar_usuario(sheet, sheet_ranges, i)
                self.criar_vinculo(codae, perfil, usuario, data_atual)

    def criar_usuario(self, sheet, sheet_ranges, i):
        fix_coluna = 0 if sheet['nome'] == 'Planilha1' else 1
        try:
            cpf = re.sub('[^0-9]', '', str(sheet_ranges.cell(row=i, column=(4 - fix_coluna)).value))
            email = (sheet_ranges.cell(row=i, column=(5 - fix_coluna)).value).strip()
            rf = re.sub('[^0-9]', '', str(sheet_ranges.cell(row=i, column=(6 - fix_coluna)).value))

            for user in Usuario.objects.filter(email=email):
                Vinculo.objects.filter(usuario=user).delete()
                user.delete()

            self.stdout.write(f'criando usuario com email: {email} cpf: {cpf} index: {i}')
            usuario = Usuario.objects.create_superuser(
                email=email,
                password='SIGPAE123',
                cpf=cpf,
                registro_funcional=rf,
                nome=sheet_ranges.cell(row=i, column=(2 - fix_coluna)).value,
                crn_numero=sheet_ranges.cell(row=i, column=(7 - fix_coluna)).value,
                cargo='ANALISTA DE SAUDE',
            )
            return usuario

        except Exception as e:
            self.stdout.write(f'Houve um erro na criação do usuário: {e}')
            pass

    def criar_vinculo(self, codae, perfil, usuario, data_atual):
        Vinculo.objects.create(
            instituicao=codae,
            perfil=perfil,
            usuario=usuario,
            data_inicial=data_atual,
        )
