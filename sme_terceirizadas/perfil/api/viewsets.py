import datetime
import logging

from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.http import HttpResponse
from django_filters import rest_framework as filters
from openpyxl import Workbook, styles
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST

from sme_terceirizadas.perfil.models.perfil import Vinculo
from sme_terceirizadas.perfil.models.usuario import (
    ImportacaoPlanilhaUsuarioExternoCoreSSO,
    ImportacaoPlanilhaUsuarioServidorCoreSSO
)

from ...dados_comuns.constants import ADMINISTRADOR_EMPRESA, DIRETOR_UE, USUARIO_EMPRESA
from ...dados_comuns.permissions import PermissaoParaCriarUsuarioComCoresso, UsuarioSuperCodae
from ...escola.api.serializers import UsuarioDetalheSerializer
from ...escola.models import Codae
from ...terceirizada.models import Terceirizada
from ..api.helpers import ofuscar_email
from ..models import Perfil, PerfisVinculados, Usuario
from ..tasks import (
    busca_cargo_de_usuario,
    processa_planilha_usuario_externo_coresso_async,
    processa_planilha_usuario_servidor_coresso_async
)
from ..utils import PerfilPagination
from .filters import ImportacaoPlanilhaUsuarioCoreSSOFilter, VinculoFilter
from .serializers import (
    AlteraEmailSerializer,
    ImportacaoPlanilhaUsuarioExternoCoreSSOCreateSerializer,
    ImportacaoPlanilhaUsuarioExternoCoreSSOSerializer,
    ImportacaoPlanilhaUsuarioServidorCoreSSOCreateSerializer,
    ImportacaoPlanilhaUsuarioServidorCoreSSOSerializer,
    PerfilSimplesSerializer,
    PerfisVinculadosSerializer,
    RedefinirSenhaSerializer,
    UsuarioComCoreSSOCreateSerializer,
    UsuarioSerializer,
    UsuarioUpdateSerializer,
    VinculoSerializer,
    VinculoSimplesSerializer
)

logger = logging.getLogger(__name__)


class UsuarioViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = (IsAuthenticated,)
    lookup_field = 'uuid'
    queryset = Usuario.objects.all()
    serializer_class = UsuarioDetalheSerializer

    @action(detail=False, url_path='atualizar-email', methods=['patch'])
    def atualizar_email(self, request):
        usuario = request.user
        tipo_email = request.data.get('tipo_email')
        usuario.tipo_email = tipo_email
        usuario.email = request.data.get('email')
        usuario.save()
        serializer = self.get_serializer(usuario)
        return Response(serializer.data)

    @action(detail=False, url_path='atualizar-cargo', methods=['get'])
    def atualizar_cargo(self, request):
        usuario = request.user
        registro_funcional = usuario.registro_funcional
        busca_cargo_de_usuario.delay(registro_funcional)
        return Response({'detail': 'sucesso'}, status=status.HTTP_200_OK)

    @action(detail=False, url_path='atualizar-senha', methods=['patch'])
    @transaction.atomic
    def atualizar_senha(self, request):
        serializer = RedefinirSenhaSerializer()
        validated_data = serializer.validate(request.data)
        usuario = request.user
        result = serializer.update(usuario, validated_data)
        usuario.last_login = datetime.datetime.now()
        usuario.save()
        if isinstance(result, Response):
            usuario.last_login = None
            usuario.save()
            logger.error('Erro ao alterar a senha:', result)
            return result
        return Response({'detail': 'Senha alterada com sucesso'}, status=status.HTTP_200_OK)

    @action(detail=False, url_path='meus-dados')
    def meus_dados(self, request):
        usuario = request.user
        serializer = self.get_serializer(usuario)
        return Response(serializer.data)


class UsuarioUpdateViewSet(viewsets.GenericViewSet):
    permission_classes = (AllowAny,)
    serializer_class = UsuarioUpdateSerializer

    def get_authenticators(self, *args, **kwargs):
        if 'post' in self.action_map and self.action_map['post'] == 'create':
            return []
        return super().get_authenticators(*args, **kwargs)

    def _get_usuario(self, request):
        if request.data.get('registro_funcional') is not None:
            return Usuario.objects.get(registro_funcional=request.data.get('registro_funcional'))
        else:
            return Usuario.objects.get(email=request.data.get('email'))

    def _get_usuario_por_rf_cpf(self, registro_funcional_ou_cpf):
        return Usuario.objects.get(username=registro_funcional_ou_cpf)

    def create(self, request):  # noqa C901
        try:
            usuario = self._get_usuario(request)
            # TODO: ajeitar isso aqui
            usuario = UsuarioUpdateSerializer(
                usuario).partial_update(usuario, request.data)
            if not isinstance(usuario.vinculo_atual.instituicao, Terceirizada):
                usuario.enviar_email_confirmacao()
            return Response(UsuarioDetalheSerializer(usuario).data)
        except ValidationError as e:
            return Response({'detail': e.detail[0].title()}, status=status.HTTP_400_BAD_REQUEST)
        except ObjectDoesNotExist:
            if request.data.get('registro_funcional'):
                mensagem = 'RF não cadastrado no sistema'
            else:
                mensagem = 'E-mail não cadastrado no sistema'
            return Response({'detail': mensagem}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, url_path='recuperar-senha/(?P<registro_funcional_ou_cpf>.*)')
    def recuperar_senha(self, request, registro_funcional_ou_cpf=None):
        try:
            usuario = self._get_usuario_por_rf_cpf(
                registro_funcional_ou_cpf)
        except ObjectDoesNotExist:
            return Response({'detail': 'Não existe usuário com este CPF ou RF'},
                            status=status.HTTP_400_BAD_REQUEST)
        usuario.enviar_email_recuperacao_senha()
        return Response({'email': f'{ofuscar_email(usuario.email)}'})

    @action(detail=False, methods=['POST'], url_path='atualizar-senha/(?P<usuario_uuid>.*)/(?P<token_reset>.*)')  # noqa
    def atualizar_senha(self, request, usuario_uuid=None, token_reset=None):
        # TODO: melhorar este método
        senha1 = request.data.get('senha1')
        senha2 = request.data.get('senha2')
        if senha1 != senha2:
            return Response({'detail': 'Senhas divergem'}, status.HTTP_400_BAD_REQUEST)
        try:
            usuario = Usuario.objects.get(uuid=usuario_uuid)
        except ObjectDoesNotExist:
            return Response({'detail': 'Não existe usuário com este e-mail ou RF'},
                            status=status.HTTP_400_BAD_REQUEST)
        if usuario.atualiza_senha(senha=senha1, token=token_reset):
            return Response({'sucesso!': 'senha atualizada com sucesso'})
        else:
            return Response({'detail': 'Token inválido'}, status.HTTP_400_BAD_REQUEST)


class PerfilViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = (IsAuthenticated,)
    lookup_field = 'uuid'
    queryset = Perfil.objects.all()
    serializer_class = PerfilSimplesSerializer

    @action(detail=False)
    def visoes(self, request):
        return Response(Perfil.visoes_to_json())


class PerfisVinculadosViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = (IsAuthenticated,)
    queryset = PerfisVinculados.objects.all()
    serializer_class = PerfisVinculadosSerializer

    @action(detail=True, methods=['get'], url_path='perfis-subordinados')
    def retorna_perfis_subordinados(self, request, pk=None):
        try:
            perfil_buscado = Perfil.objects.get(nome=pk)
            perfis_vinculados = PerfisVinculados.objects.get(perfil_master=perfil_buscado)
            retorno = []
            for p in perfis_vinculados.perfis_subordinados.all():
                retorno.append(p.nome)
            return Response(retorno)

        except ObjectDoesNotExist:
            return Response({'detail': 'Perfil ou Perfis Vinculados não encontrado'},
                            status=status.HTTP_400_BAD_REQUEST)


class UsuarioConfirmaEmailViewSet(viewsets.GenericViewSet):
    permission_classes = (AllowAny,)
    serializer_class = UsuarioDetalheSerializer

    # TODO: ajeitar isso
    def list(self, request, uuid, confirmation_key):  # noqa C901
        try:
            usuario = Usuario.objects.get(uuid=uuid)
            usuario.confirm_email(confirmation_key)
            usuario.is_active = usuario.is_confirmed
        except ObjectDoesNotExist:
            return Response({'detail': 'Erro ao confirmar email'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            vinculo_esperando_ativacao = usuario.vinculos.get(
                ativo=False, data_inicial=None, data_final=None)
            vinculo_esperando_ativacao.ativo = True
            vinculo_esperando_ativacao.data_inicial = datetime.date.today()
            vinculo_esperando_ativacao.save()
        except ObjectDoesNotExist:
            return Response({'detail': 'Não possui vínculo'},
                            status=status.HTTP_400_BAD_REQUEST)

        usuario.save()
        return Response(UsuarioDetalheSerializer(usuario).data)


class VinculoViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = (IsAuthenticated,)
    lookup_field = 'uuid'
    queryset = Vinculo.objects.all()
    serializer_class = VinculoSerializer
    pagination_class = PerfilPagination
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = VinculoFilter

    @action(detail=False, methods=['GET'], url_path='vinculos-ativos', permission_classes=(IsAuthenticated,))
    def lista_vinculos_ativos(self, request):
        usuario = request.user
        if usuario.vinculo_atual.perfil.nome in [DIRETOR_UE, ADMINISTRADOR_EMPRESA, USUARIO_EMPRESA]:
            queryset = self.get_queryset().filter(
                content_type=usuario.vinculo_atual.content_type,
                object_id=usuario.vinculo_atual.object_id
            ).order_by('-data_inicial')
        else:
            queryset = self.get_queryset().order_by('-data_inicial')

        queryset = [vinc for vinc in self.filter_queryset(
            queryset) if vinc.status is Vinculo.STATUS_ATIVO]
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = VinculoSimplesSerializer(page, many=True)
            response = self.get_paginated_response(
                serializer.data
            )
            return response
        response = VinculoSimplesSerializer(queryset, many=True).data
        return Response(response)

    @action(detail=False, methods=['GET'], url_path='vinculo-empresa', permission_classes=(IsAuthenticated,))
    def vinculo_empresa(self, request):
        usuario = request.user
        if usuario.vinculo_atual.perfil.nome in [ADMINISTRADOR_EMPRESA, USUARIO_EMPRESA]:
            vinculos = Vinculo.objects.filter(usuario=usuario)
            vinculo = [vinc for vinc in vinculos if vinc.status is Vinculo.STATUS_ATIVO][0]
            response = {'nome': vinculo.instituicao.nome, 'cnpj': vinculo.instituicao.cnpj}
        else:
            response = {'erro': 'Vínculo não encontrato.'}
        return Response(response, status=status.HTTP_200_OK)


def exportar_planilha_importacao_usuarios_perfil_codae(request, **kwargs):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=planilha_importacao_usuarios_perfil_CODAE.xlsx'
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = 'CODAE'
    headers = [
        'Nome do Usuário',
        'Cargo',
        'Email',
        'CPF',
        'Telefone',
        'RF',
        'Perfil',
        'Número CRN'
    ]
    _font = styles.Font(name='Calibri', sz=10)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill('solid', fgColor='ffff99')
        cabecalho.font = styles.Font(name='Calibri', size=10, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style='thin', color='000000'),
            right=styles.Side(border_style='thin', color='000000'),
            top=styles.Side(border_style='thin', color='000000'),
            bottom=styles.Side(border_style='thin', color='000000')
        )
    dv = DataValidation(
        type='list',
        formula1='"COORDENADOR_GESTAO_ALIMENTACAO_TERCEIRIZADA,'
                 'COORDENADOR_DIETA_ESPECIAL,'
                 'COORDENADOR_SUPERVISAO_NUTRICAO,'
                 'COORDENADOR_GESTAO_PRODUTO"',
        allow_blank=True
    )
    dv.error = 'Perfil Inválido'
    dv.errorTitle = 'Perfil não permitido'
    ws.add_data_validation(dv)
    dv.add('G2:G1048576')
    workbook.save(response)

    return response


def exportar_planilha_importacao_usuarios_perfil_escola(request, **kwargs):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=planilha_importacao_usuarios_perfil_ESCOLA.xlsx'
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = 'ESCOLA'
    headers = [
        'Cód. EOL da U.E',
        'Nome do Usuário',
        'Cargo',
        'Email',
        'CPF',
        'Telefone',
        'RF',
        'Perfil',
    ]
    _font = styles.Font(name='Calibri', sz=10)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill('solid', fgColor='ffff99')
        cabecalho.font = styles.Font(name='Calibri', size=10, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style='thin', color='000000'),
            right=styles.Side(border_style='thin', color='000000'),
            top=styles.Side(border_style='thin', color='000000'),
            bottom=styles.Side(border_style='thin', color='000000')
        )
    dv = DataValidation(
        type='list',
        formula1='"DIRETOR_UE"',
        allow_blank=True
    )
    dv.error = 'Perfil Inválido'
    dv.errorTitle = 'Perfil não permitido'
    ws.add_data_validation(dv)
    dv.add('H2:H1048576')
    workbook.save(response)

    return response


def exportar_planilha_importacao_usuarios_perfil_dre(request, **kwargs):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=planilha_importacao_usuarios_perfil_DRE.xlsx'
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = 'DRE'
    headers = [
        'Cód. EOL da DRE',
        'Nome do Usuário',
        'Cargo',
        'Email',
        'CPF',
        'Telefone',
        'RF',
        'Perfil',
    ]
    _font = styles.Font(name='Calibri', sz=10)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill('solid', fgColor='ffff99')
        cabecalho.font = styles.Font(name='Calibri', size=10, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style='thin', color='000000'),
            right=styles.Side(border_style='thin', color='000000'),
            top=styles.Side(border_style='thin', color='000000'),
            bottom=styles.Side(border_style='thin', color='000000')
        )
    dv = DataValidation(
        type='list',
        formula1='"COGESTOR, ADMINISTRADOR_DRE"',
        allow_blank=True
    )
    dv.error = 'Perfil Inválido'
    dv.errorTitle = 'Perfil não permitido'
    ws.add_data_validation(dv)
    dv.add('H2:H1048576')
    workbook.save(response)

    return response


def exportar_planilha_importacao_usuarios_servidor_coresso(request, **kwargs):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=planilha_importacao_usuarios_perfil_servidor_coresso.xlsx'
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = 'Servidores CoreSSO'
    headers = [
        'Cód. EOL da Instituição (Escola, DRE ou CODAE)',
        'Nome do Usuário',
        'Cargo',
        'Email',
        'CPF',
        'RF',
        'Tipo de perfil',
        'Perfil',
        'CODAE'
    ]
    _font = styles.Font(name='Calibri', sz=10)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill('solid', fgColor='ffff99')
        cabecalho.font = styles.Font(name='Calibri', size=10, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style='thin', color='000000'),
            right=styles.Side(border_style='thin', color='000000'),
            top=styles.Side(border_style='thin', color='000000'),
            bottom=styles.Side(border_style='thin', color='000000')
        )
    tipos_de_perfil = 'escola,dre,codae'
    dv = DataValidation(
        type='list',
        formula1=f'{tipos_de_perfil}',
        allow_blank=True
    )
    dv.error = 'Tipo de perfil Inválido'
    dv.errorTitle = 'Tipo de Perfil não permitido'
    ws.add_data_validation(dv)
    dv.add('G2:G1048576')

    perfis = ', '.join([p.nome for p in Perfil.objects.all()])
    dv2 = DataValidation(
        type='list',
        formula1=f'{perfis}',
        allow_blank=True
    )
    dv2.error = 'Perfil Inválido'
    dv2.errorTitle = 'Perfil não permitido'
    ws.add_data_validation(dv2)
    dv2.add('H2:H1048576')

    codaes = ', '.join([c.nome.split(' - ')[1] for c in Codae.objects.all()]).replace(' ', '')
    dv3 = DataValidation(
        type='list',
        formula1=f'{codaes}',
        allow_blank=True
    )
    dv3.error = 'CODAE Inválida'
    dv3.errorTitle = 'CODAE não permitida'
    ws.add_data_validation(dv3)
    dv3.add('I2:I1048576')

    workbook.save(response)

    return response


def exportar_planilha_importacao_usuarios_externos_coresso(request, **kwargs):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=planilha_importacao_usuarios_perfil_servidor_coresso.xlsx'
    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = 'Usuários Externos CoreSSO'
    headers = [
        'Nome do Usuário',
        'Email',
        'CPF',
        'Perfil',
        'CNPJ da Empresa'
    ]
    _font = styles.Font(name='Calibri', sz=10)
    {k: setattr(styles.DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    for i in range(0, len(headers)):
        cabecalho = ws.cell(row=1, column=1 + i, value=headers[i])
        cabecalho.fill = styles.PatternFill('solid', fgColor='ffff99')
        cabecalho.font = styles.Font(name='Calibri', size=10, bold=True)
        cabecalho.border = styles.Border(
            left=styles.Side(border_style='thin', color='000000'),
            right=styles.Side(border_style='thin', color='000000'),
            top=styles.Side(border_style='thin', color='000000'),
            bottom=styles.Side(border_style='thin', color='000000')
        )
    perfis = ', '.join([p.nome for p in Perfil.objects.all()])
    dv = DataValidation(
        type='list',
        formula1=f'{perfis}',
        allow_blank=True
    )
    dv.error = 'Perfil Inválido'
    dv.errorTitle = 'Perfil não permitido'
    ws.add_data_validation(dv)
    dv.add('D2:D1048576')

    workbook.save(response)

    return response


class UsuarioComCoreSSOViewSet(mixins.CreateModelMixin, viewsets.GenericViewSet):
    lookup_field = 'username'
    permission_classes = (IsAuthenticated,)
    serializer_class = UsuarioComCoreSSOCreateSerializer
    permission_action_classes = {
        'create': [PermissaoParaCriarUsuarioComCoresso],
        'delete': [PermissaoParaCriarUsuarioComCoresso]
    }
    queryset = Usuario.objects.all()

    @action(detail=True, permission_classes=(UsuarioSuperCodae,),
            methods=['post'], url_path='finalizar-vinculo')
    def finaliza_vinculo(self, request, username):
        """(post) /cadastro-com-coresso/{usuario.username}/finalizar-vinculo/."""
        try:
            user = Usuario.objects.get(username=username)
            user.vinculo_atual.finalizar_vinculo()
            return Response(dict(detail=f'Acesso removido com sucesso!'), status=HTTP_200_OK)
        except ObjectDoesNotExist as e:
            return Response(dict(detail=f'Usuário não encontrado: {e}'), status=HTTP_400_BAD_REQUEST)

    @action(detail=True, permission_classes=(UsuarioSuperCodae,),
            url_path='alterar-email', methods=['patch'])
    @transaction.atomic
    def altera_email(self, request, username):
        """(patch) /cadastro-com-coresso/{usuario.username}/alterar-email/."""
        data = request.data
        serialize = AlteraEmailSerializer()
        validated_data = serialize.validate(data)
        user = Usuario.objects.get(username=username)
        instance = serialize.update(user, validated_data)
        if isinstance(instance, Response):
            return instance
        return Response(UsuarioSerializer(instance, context={'request': request}).data, status=status.HTTP_200_OK)


class ImportacaoPlanilhaUsuarioServidorCoreSSOViewSet(mixins.RetrieveModelMixin,
                                                      mixins.ListModelMixin,
                                                      mixins.CreateModelMixin,
                                                      viewsets.GenericViewSet):

    permission_classes = (UsuarioSuperCodae,)
    lookup_field = 'uuid'
    queryset = ImportacaoPlanilhaUsuarioServidorCoreSSO.objects.all().order_by('-criado_em')
    serializer_class = ImportacaoPlanilhaUsuarioServidorCoreSSOSerializer
    pagination_class = PerfilPagination
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = ImportacaoPlanilhaUsuarioCoreSSOFilter

    def get_serializer_class(self):
        if self.action in ['retrieve', 'list']:
            return ImportacaoPlanilhaUsuarioServidorCoreSSOSerializer
        else:
            return ImportacaoPlanilhaUsuarioServidorCoreSSOCreateSerializer

    @action(detail=False, methods=['GET'], permission_classes=(UsuarioSuperCodae,),
            url_path='download-planilha-servidor')
    def exportar_planilha_servidor(self, request):
        return exportar_planilha_importacao_usuarios_servidor_coresso(request)

    @action(detail=True, permission_classes=(UsuarioSuperCodae,),
            methods=['post'], url_path='processar-importacao')
    def processar_importacao_usuario_servidor(self, request, uuid):
        """(post) /planilha-coresso-servidor/{ImportacaoPlanilhaUsuarioServidorCoreSSO.uuid}/processar-importacao/."""
        logger.info('Processando arquivo de carga de usuário externo com uuid %s.', uuid)
        username = request.user.get_username()

        arquivo = self.get_object()
        if not arquivo:
            msg = f'Arquivo com uuid {uuid} não encontrado.'
            logger.info(msg)
            return Response(msg, status=status.HTTP_404_NOT_FOUND)
        processa_planilha_usuario_servidor_coresso_async.delay(username=username, arquivo_uuid=uuid)

        return Response(dict(detail='Processamento de importação iniciado com sucesso.'), status=HTTP_200_OK)

    @action(detail=True, permission_classes=(UsuarioSuperCodae,),
            methods=['patch'], url_path='remover')
    def remover_planilha_usuario_servidor(self, request, uuid):
        """(patch) /planilha-coresso-servidor/{ImportacaoPlanilhaUsuarioServidorCoreSSO.uuid}/remover/."""
        arquivo = self.get_object()
        if not arquivo:
            msg = f'Arquivo com uuid {uuid} não encontrado.'
            logger.info(msg)
            return Response(msg, status=status.HTTP_404_NOT_FOUND)

        arquivo.removido()

        return Response(dict(detail='Arquivo removido com sucesso.'), status=HTTP_200_OK)


class ImportacaoPlanilhaUsuarioExternoCoreSSOViewSet(mixins.RetrieveModelMixin,
                                                     mixins.ListModelMixin,
                                                     mixins.CreateModelMixin,
                                                     viewsets.GenericViewSet):

    permission_classes = (UsuarioSuperCodae,)
    lookup_field = 'uuid'
    queryset = ImportacaoPlanilhaUsuarioExternoCoreSSO.objects.all().order_by('-criado_em')
    serializer_class = ImportacaoPlanilhaUsuarioExternoCoreSSOSerializer
    pagination_class = PerfilPagination
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = ImportacaoPlanilhaUsuarioCoreSSOFilter

    def get_serializer_class(self):
        if self.action in ['retrieve', 'list']:
            return ImportacaoPlanilhaUsuarioExternoCoreSSOSerializer
        else:
            return ImportacaoPlanilhaUsuarioExternoCoreSSOCreateSerializer

    @action(detail=False, methods=['GET'], permission_classes=(UsuarioSuperCodae,),
            url_path='download-planilha-nao-servidor')
    def exportar_planilha_externos(self, request):
        return exportar_planilha_importacao_usuarios_externos_coresso(request)

    @action(detail=True, permission_classes=(UsuarioSuperCodae,),
            methods=['post'], url_path='processar-importacao')
    def processar_importacao_usuario_externo(self, request, uuid):
        """(post) /planilha-coresso-externo/{ImportacaoPlanilhaUsuarioExternoCoreSSO.uuid}/processar-importacao/."""
        logger.info('Processando arquivo de carga de usuário externo com uuid %s.', uuid)
        username = request.user.get_username()

        arquivo = self.get_object()
        if not arquivo:
            msg = f'Arquivo com uuid {uuid} não encontrado.'
            logger.info(msg)
            return Response(msg, status=status.HTTP_404_NOT_FOUND)

        processa_planilha_usuario_externo_coresso_async.delay(username=username, arquivo_uuid=uuid)

        return Response(dict(detail='Processamento de importação iniciado com sucesso.'), status=HTTP_200_OK)

    @action(detail=True, permission_classes=(UsuarioSuperCodae,),
            methods=['patch'], url_path='remover')
    def remover_planilha_usuario_externo(self, request, uuid):
        """(patch) /planilha-coresso-externo/{ImportacaoPlanilhaUsuarioExternoCoreSSO.uuid}/remover/."""
        arquivo = self.get_object()
        if not arquivo:
            msg = f'Arquivo com uuid {uuid} não encontrado.'
            logger.info(msg)
            return Response(msg, status=status.HTTP_404_NOT_FOUND)

        arquivo.removido()

        return Response(dict(detail='Arquivo removido com sucesso.'), status=HTTP_200_OK)
