import logging
from tempfile import NamedTemporaryFile
from typing import Type

from django.core.files import File
from django.db import transaction
from openpyxl import Workbook, load_workbook, styles

from sme_terceirizadas.dados_comuns.behaviors import StatusAtivoInativo
from sme_terceirizadas.imr.models import (
    ImportacaoPlanilhaTipoPenalidade,
    ObrigacaoPenalidade,
    TipoGravidade,
    TipoPenalidade,
)
from sme_terceirizadas.perfil.models import Usuario
from sme_terceirizadas.terceirizada.models import Edital
from utility.carga_dados.imr.schemas import ImportacaoPlanilhaTipoPenalidadeSchema

logger = logging.getLogger("sigpae.carga_dados_tipo_penalidade_importa_dados")


def importa_tipos_penalidade(
    usuario: Usuario, arquivo: ImportacaoPlanilhaTipoPenalidade
) -> None:
    logger.debug(f"Iniciando o processamento do arquivo: {arquivo.uuid}")

    try:
        processador = ProcessadorPlanilhaTipoPenalidade(
            usuario, arquivo, ImportacaoPlanilhaTipoPenalidadeSchema
        )
        processador.processamento()
        processador.finaliza_processamento()
    except Exception as exc:
        logger.error(f"Erro genérico: {exc}")


class ProcessadorPlanilhaTipoPenalidade:
    LINHA_2 = 2

    def __init__(
        self,
        usuario: Usuario,
        arquivo: ImportacaoPlanilhaTipoPenalidade,
        class_schema: Type[ImportacaoPlanilhaTipoPenalidadeSchema],
    ) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.class_schema = class_schema
        self.erros = []
        self.worksheet = self.abre_worksheet()

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def processamento(self):
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        logger.info(
            f"Quantidade de linhas: {len(linhas)} -- Quantidade de colunas: {len(linhas[0])}"
        )

        for ind, linha in enumerate(
            linhas[1:], self.LINHA_2
        ):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)
                usuario_schema = self.class_schema(**dicionario_dados)

                logger.info(
                    f"Criando tipo de penalidade: {usuario_schema.edital} -- {usuario_schema.numero_clausula}"
                )
                self.cria_tipo_penalidade(ind, usuario_schema)
            except Exception as exc:
                self.erros.append(f"Linha {ind} - {exc}")

    def cria_tipo_penalidade(
        self, ind: int, usuario_schema: ImportacaoPlanilhaTipoPenalidadeSchema
    ):
        try:
            self.__cria_tipo_penalidade(usuario_schema)
        except Exception as exd:
            self.erros.append(f"Linha {ind} - {exd}")

    def get_edital(self, numero_edital: str):
        return Edital.objects.get(numero=numero_edital)

    def get_tipo_gravidade(self, tipo_gravidade: str):
        return TipoGravidade.objects.get(tipo=tipo_gravidade)

    def __cria_tipo_penalidade_obj(
        self, usuario_schema: ImportacaoPlanilhaTipoPenalidadeSchema
    ):
        edital = self.get_edital(usuario_schema.edital)
        gravidade = self.get_tipo_gravidade(usuario_schema.gravidade)
        status = usuario_schema.status == StatusAtivoInativo.ATIVO
        tipo_penalidade, _ = TipoPenalidade.objects.update_or_create(
            edital=edital,
            numero_clausula=usuario_schema.numero_clausula,
            defaults={
                "criado_por": self.usuario,
                "gravidade": gravidade,
                "descricao": usuario_schema.descricao_clausula,
                "status": status,
            },
        )
        return tipo_penalidade

    def normaliza_obrigacoes(self, obrigacoes: str):
        if obrigacoes[-1] == ";":
            obrigacoes = obrigacoes[: len(obrigacoes) - 1]
        return obrigacoes.split(";")

    def __cria_obrigacao_penalidade(
        self, tipo_penalidade: TipoPenalidade, descricao: str
    ):
        ObrigacaoPenalidade.objects.create(
            tipo_penalidade=tipo_penalidade, descricao=descricao
        )

    @transaction.atomic
    def __cria_tipo_penalidade(
        self, usuario_schema: ImportacaoPlanilhaTipoPenalidadeSchema
    ):
        tipo_penalidade = self.__cria_tipo_penalidade_obj(usuario_schema)
        lista_obrigacoes = self.normaliza_obrigacoes(usuario_schema.obrigacoes)
        for descricao_obrigacao in lista_obrigacoes:
            self.__cria_obrigacao_penalidade(tipo_penalidade, descricao_obrigacao)

    def validacao_inicial(self) -> bool:
        return self.existe_conteudo()

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = "Não foi feito o upload da planilha"
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {
            key: linha[index].value
            for index, key in enumerate(self.class_schema.schema()["properties"].keys())
        }

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = "\n".join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
            logger.error(f'Arquivo "{self.arquivo.uuid}" processado com erro(s).')
        else:
            self.arquivo.log = "Planilha processada com sucesso."
            self.arquivo.processamento_com_sucesso()
            logger.info(f'Arquivo "{self.arquivo.uuid}" processado com sucesso.')

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = "Erros"
        cabecalho = ws.cell(
            row=1, column=1, value="Erros encontrados no processamento da planilha"
        )
        cabecalho.fill = styles.PatternFill("solid", fgColor="808080")
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f"arquivo_resultado_{self.arquivo.pk}.xlsx"
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))
