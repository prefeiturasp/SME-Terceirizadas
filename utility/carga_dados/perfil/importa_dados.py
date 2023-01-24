import logging
from datetime import date
from tempfile import NamedTemporaryFile

from django.core.files import File
from django.db import transaction
from django.db.models import F, Func, Q, Value
from openpyxl import Workbook, load_workbook, styles

from sme_terceirizadas.perfil.models.usuario import ImportacaoPlanilhaUsuarioServidorCoreSSO, \
    ImportacaoPlanilhaUsuarioExternoCoreSSO
from sme_terceirizadas.perfil.services.usuario_coresso_service import EOLUsuarioCoreSSO
from sme_terceirizadas.terceirizada.models import Terceirizada
from utility.carga_dados.escola.helper import bcolors
from utility.carga_dados.helper import ja_existe, progressbar

from sme_terceirizadas.dados_comuns.models import Contato
from sme_terceirizadas.dados_comuns.constants import DJANGO_ADMIN_TREINAMENTO_PASSWORD
from sme_terceirizadas.escola.models import Codae, DiretoriaRegional, Escola
from sme_terceirizadas.perfil.data.perfis import data_perfis
from sme_terceirizadas.perfil.models import Perfil, Usuario, Vinculo
from sme_terceirizadas.perfil.models import (
    ImportacaoPlanilhaUsuarioPerfilCodae,
    ImportacaoPlanilhaUsuarioPerfilDre,
    ImportacaoPlanilhaUsuarioPerfilEscola
)

import magic

from .schemas import (
    ImportacaoPlanilhaUsuarioPerfilCodaeSchema,
    ImportacaoPlanilhaUsuarioPerfilDreSchema,
    ImportacaoPlanilhaUsuarioPerfilEscolaSchema,
    ImportacaoPlanilhaUsuarioServidorCoreSSOSchema, ImportacaoPlanilhaUsuarioExternoCoreSSOSchema
)

logger = logging.getLogger('sigpae.carga_dados_perfil_importa_dados')


def cria_perfis():
    for item in progressbar(data_perfis, 'Perfil'):
        _, created = Perfil.objects.get_or_create(
            nome=item['nome'],
            ativo=item['ativo'],
            super_usuario=item['super_usuario'],
        )
        if not created:
            ja_existe('Perfil', item['nome'])


def cria_vinculos():
    perfil = {
        'perfil_diretor_escola': Perfil.objects.get(nome='DIRETOR_UE'),
        'perfil_administrador_ue': Perfil.objects.get(nome='ADMINISTRADOR_UE'),
        'perfil_cogestor_dre': Perfil.objects.get(nome='COGESTOR'),
        'perfil_usuario_codae': Perfil.objects.get(nome='COORDENADOR_GESTAO_ALIMENTACAO_TERCEIRIZADA'),  # noqa
        'perfil_usuario_dilog': Perfil.objects.get(nome='COORDENADOR_LOGISTICA'),  # noqa
        'perfil_usuario_nutri_codae': Perfil.objects.get(nome='COORDENADOR_DIETA_ESPECIAL'),  # noqa
        'perfil_usuario_nutri_supervisao': Perfil.objects.get(nome='COORDENADOR_SUPERVISAO_NUTRICAO'),  # noqa
        'perfil_usuario_nutri_manifestacao': Perfil.objects.get(nome='COORDENADOR_SUPERVISAO_NUTRICAO_MANIFESTACAO'),
        'perfil_coordenador_gestao_produto': Perfil.objects.get(nome='COORDENADOR_GESTAO_PRODUTO'),  # noqa
        'perfil_usuario_terceirizada': Perfil.objects.get(nome='ADMINISTRADOR_EMPRESA'),  # noqa
        'perfil_usuario_codae_dilog': Perfil.objects.get(nome='COORDENADOR_CODAE_DILOG_LOGISTICA'),
        'perfil_usuario_codae_gabinete': Perfil.objects.get(nome='ADMINISTRADOR_CODAE_GABINETE'),
        'perfil_usuario_dilog_contabil': Perfil.objects.get(nome='ADMINISTRADOR_CODAE_DILOG_CONTABIL'),
        'perfil_usuario_dilog_juridico': Perfil.objects.get(nome='ADMINISTRADOR_CODAE_DILOG_JURIDICO'),
        'perfil_admin_fornecedor': Perfil.objects.get(nome='ADMINSTRADOR_FORNECEDOR'),
        'perfil_usuario_cronograma': Perfil.objects.get(nome='DILOG_CRONOGRAMA'),
        'perfil_usuario_qualidade': Perfil.objects.get(nome='DILOG_QUALIDADE'),
        'perfil_usuario_dilog_diretoria': Perfil.objects.get(nome='DILOG_DIRETORIA'),
        'perfil_usuario_dinutre_diretoria': Perfil.objects.get(nome='DINUTRE_DIRETORIA'),
        'perfil_usuario_representante_codae': Perfil.objects.get(nome='ADMINISTRADOR_REPRESENTANTE_CODAE')
    }

    usuario = {
        'usuario_escola': Usuario.objects.get(email='escola@admin.com'),
        'usuario_escola_cei': Usuario.objects.get(email='escolacei@admin.com'),
        'usuario_escola_cei_ceu': Usuario.objects.get(email='escolaceiceu@admin.com'),
        'usuario_escola_cci': Usuario.objects.get(email='escolacci@admin.com'),
        'usuario_escola_emef': Usuario.objects.get(email='escolaemef@admin.com'),
        'usuario_escola_emef3': Usuario.objects.get(email='escolaemef3@admin.com'),
        'usuario_escola_emebs': Usuario.objects.get(email='escolaemebs@admin.com'),
        'usuario_escola_cieja': Usuario.objects.get(email='escolacieja@admin.com'),
        'usuario_escola_emei': Usuario.objects.get(email='escolaemei@admin.com'),
        'usuario_escola_ceu_emei': Usuario.objects.get(email='escolaceuemei@admin.com'),
        'usuario_escola_ceu_emef': Usuario.objects.get(email='escolaceuemef@admin.com'),
        'usuario_dre': Usuario.objects.get(email='dre@admin.com'),
        'usuario_codae': Usuario.objects.get(email='codae@admin.com'),
        'usuario_dilog': Usuario.objects.get(email='dilog@admin.com'),
        'usuario_nutri_codae': Usuario.objects.get(email='nutricodae@admin.com'),
        'usuario_nutri_supervisao': Usuario.objects.get(email='nutrisupervisao@admin.com'),
        'usuario_nutri_manifestacao': Usuario.objects.get(email='nutricionistamanifestacao@admin.com'),
        'usuario_gestao_produto_codae': Usuario.objects.get(email='gpcodae@admin.com'),
        'usuario_terceirizada': Usuario.objects.get(email='terceirizada@admin.com'),
        'usuario_ue': Usuario.objects.get(email='ue@admin.com'),
        'usuario_codae_gabinete': Usuario.objects.get(email='codaegabinete@admin.com'),
        'usuario_codae_logistica': Usuario.objects.get(email='codaelogistica@admin.com'),
        'usuario_codae_contabio': Usuario.objects.get(email='codaecontabil@admin.com'),
        'usuario_codae_juridico': Usuario.objects.get(email='codaejuridico@admin.com'),
        'usuario_ue_mista': Usuario.objects.get(email='uemista@admin.com'),
        'usuario_ue_direta': Usuario.objects.get(email='uedireta@admin.com'),
        'usuario_ue_parceira': Usuario.objects.get(email='ueparceira@admin.com'),
        'usuario_diretor_ue_abastecimento': Usuario.objects.get(email='diretorabastecimento@admin.com'),
        'usuario_admin_fornecedor': Usuario.objects.get(email='fornecedor@admin.com'),
        'usuario_cronograma': Usuario.objects.get(email='cronograma@admin.com'),
        'usuario_qualidade': Usuario.objects.get(email='qualidade@admin.com'),
        'usuario_dilog_diretoria': Usuario.objects.get(email='dilogdiretoria@admin.com'),
        'usuario_dinutre_diretoria': Usuario.objects.get(email='dinutrediretoria@admin.com'),
        'usuario_representante_codae': Usuario.objects.get(email='representantecodae@admin.com'),
    }

    items = [
        {
            'nome': 'EMEF JOSE ERMIRIO DE MORAIS, SEN.',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola'],
        },
        {
            'nome': 'CEI DIRET ENEDINA DE SOUSA CARVALHO',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_cei'],
        },
        {
            'nome': 'CEU CEI MENINOS',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_cei_ceu'],
        },
        {
            'nome': 'CCI/CIPS CAMARA MUNICIPAL DE SAO PAULO',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_cci'],
        },
        {
            'nome': 'EMEF PERICLES EUGENIO DA SILVA RAMOS',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_emef'],
        },
        {
            'nome': 'EMEF PLINIO SALGADO',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_emef3'],
        },
        {
            'nome': 'EMEBS HELEN KELLER',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_emebs'],
        },
        {
            'nome': 'CIEJA CLOVIS CAITANO MIQUELAZZO - IPIRANGA',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_cieja'],
        },
        {
            'nome': 'EMEI SENA MADUREIRA',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_emei'],
        },
        {
            'nome': 'CEU EMEI BENNO HUBERT STOLLENWERK, PE.',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_ceu_emei'],
        },
        {
            'nome': 'CEU EMEF MARA CRISTINA TARTAGLIA SENA, PROFA.',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_escola_ceu_emef'],
        },
        {
            'nome': 'CEI DIRET ROBERTO ARANTES LANHOSO',
            'perfil': perfil['perfil_administrador_ue'],
            'usuario': usuario['usuario_ue'],
        },
        {
            'nome': 'CEI DIRET PINHEIROS',
            'perfil': perfil['perfil_administrador_ue'],
            'usuario': usuario['usuario_ue_mista'],
        },
        {
            'nome': 'CIEJA ALUNA JESSICA NUNES HERCULANO',
            'perfil': perfil['perfil_administrador_ue'],
            'usuario': usuario['usuario_ue_direta'],
        },
        {
            'nome': 'CR.P.CONV FRATERNIDADE MARIA DE NAZARE',
            'perfil': perfil['perfil_administrador_ue'],
            'usuario': usuario['usuario_ue_parceira'],
        },
        {
            'nome': 'CEI DIRET ROBERTO ARANTES LANHOSO',
            'perfil': perfil['perfil_diretor_escola'],
            'usuario': usuario['usuario_diretor_ue_abastecimento'],
        },
        {
            'nome': 'FORNECEDOR ADMIN',
            'perfil': perfil['perfil_admin_fornecedor'],
            'usuario': usuario['usuario_admin_fornecedor'],
        },
        {
            'nome': 'REPRESENTANTE CODAE ADM',
            'perfil': perfil['perfil_usuario_representante_codae'],
            'usuario': usuario['usuario_representante_codae'],
        },
    ]

    data_atual = date.today()

    for item in progressbar(items, 'Perfil'):
        escola = Escola.objects.get(nome=item['nome'])
        vinculo = Vinculo.objects.filter(
            perfil=item['perfil'],
            usuario=item['usuario'],
        )
        if vinculo.first():
            print(f'{bcolors.FAIL}Vinculo já existe.{bcolors.ENDC}')
        else:
            Vinculo.objects.create(
                instituicao=escola,
                perfil=item['perfil'],
                usuario=item['usuario'],
                data_inicial=data_atual
            )

    diretoria_regional = DiretoriaRegional.objects.get(nome='DIRETORIA REGIONAL DE EDUCACAO IPIRANGA')  # noqa
    codae_alimentacao, created = Codae.objects.get_or_create(nome='CODAE - GESTAO ALIMENTAÇÃO')
    dilog, created = Codae.objects.get_or_create(nome='CODAE - DILOG')
    codae_dieta_especial, created = Codae.objects.get_or_create(nome='CODAE - GESTÃO DIETA ESPECIAL')
    codae_produtos, created = Codae.objects.get_or_create(nome='CODAE - GESTÃO PRODUTOS')
    codae_nutrisupervisao, created = Codae.objects.get_or_create(nome='CODAE - SUPERVISÃO DE NUTRIÇÃO')
    codae_nutrisupervisao_manifestacao, created = Codae.objects.get_or_create(nome='CODAE - NUTRIMANIFESTAÇÃO')
    escola = Escola.objects.get(nome='EMEF JOSE ERMIRIO DE MORAIS, SEN.')  # noqa
    terceirizada = escola.lote.terceirizada

    items_especificos = [
        {
            'instituicao': diretoria_regional,
            'perfil': perfil['perfil_cogestor_dre'],
            'usuario': usuario['usuario_dre'],
        },
        {
            'instituicao': codae_alimentacao,
            'perfil': perfil['perfil_usuario_codae'],
            'usuario': usuario['usuario_codae'],
        },
        {
            'instituicao': dilog,
            'perfil': perfil['perfil_usuario_dilog'],
            'usuario': usuario['usuario_dilog'],
        },
        {
            'instituicao': codae_dieta_especial,
            'perfil': perfil['perfil_usuario_nutri_codae'],
            'usuario': usuario['usuario_nutri_codae'],
        },
        {
            'instituicao': codae_nutrisupervisao,
            'perfil': perfil['perfil_usuario_nutri_supervisao'],
            'usuario': usuario['usuario_nutri_supervisao'],
        },
        {
            'instituicao': codae_nutrisupervisao_manifestacao,
            'perfil': perfil['perfil_usuario_nutri_supervisao'],
            'usuario': usuario['usuario_nutri_manifestacao'],
        },
        {
            'instituicao': codae_produtos,
            'perfil': perfil['perfil_coordenador_gestao_produto'],
            'usuario': usuario['usuario_gestao_produto_codae'],
        },
        {
            'instituicao': terceirizada,
            'perfil': perfil['perfil_usuario_terceirizada'],
            'usuario': usuario['usuario_terceirizada'],
        },
        {
            'instituicao': dilog,
            'perfil': perfil['perfil_usuario_cronograma'],
            'usuario': usuario['usuario_cronograma'],
        },
        {
            'instituicao': dilog,
            'perfil': perfil['perfil_usuario_qualidade'],
            'usuario': usuario['usuario_qualidade'],
        },
        {
            'instituicao': dilog,
            'perfil': perfil['perfil_usuario_dilog_diretoria'],
            'usuario': usuario['usuario_qualidade'],
        },
        {
            'instituicao': dilog,
            'perfil': perfil['perfil_usuario_dinutre_diretoria'],
            'usuario': usuario['usuario_dinutre_diretoria'],
        },
    ]

    for item in progressbar(items_especificos, 'Perfil especifico'):
        vinculo = Vinculo.objects.filter(
            perfil=item['perfil'],
            usuario=item['usuario'],
        )
        if vinculo.first():
            print(f'{bcolors.FAIL}Vinculo já existe.{bcolors.ENDC}')
        else:
            Vinculo.objects.create(
                instituicao=item['instituicao'],
                perfil=item['perfil'],
                usuario=item['usuario'],
                data_inicial=data_atual,
            )


def valida_arquivo_importacao_usuarios(arquivo) -> bool:
    logger.debug(f'Iniciando validação do arquivo {arquivo.conteudo}: {arquivo.uuid}')

    mime_types_validos = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
    ]
    extensoes_validas = ['.xlsx', '.xls']
    arquivo_mime_type = magic.from_buffer(arquivo.conteudo.read(2048), mime=True)
    if arquivo_mime_type not in mime_types_validos:
        arquivo.log = 'Formato de arquivo não suportado.'
        arquivo.erro_no_processamento()
        logger.error(f'Arquivo inválido: {arquivo.uuid}')
        return False
    if not arquivo.conteudo.path.endswith(tuple(extensoes_validas)):
        arquivo.log = 'Extensão de arquivo não suportada'
        arquivo.erro_no_processamento()
        logger.error(f'Arquivo inválido: {arquivo.uuid}')
        return False

    logger.info('Arquivo válido.')
    return True


class ProcessadorPlanilhaUsuarioPerfilEscola:
    def __init__(self, usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioPerfilEscola) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.erros = []
        self.worksheet = self.abre_worksheet()

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def processamento(self):  # noqa C901
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        logger.info(f'Quantidade de linhas: {len(linhas)} -- Quantidade de colunas: {len(linhas[0])}')

        for ind, linha in enumerate(linhas[1:], 2):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)
                usuario_schema = ImportacaoPlanilhaUsuarioPerfilEscolaSchema(**dicionario_dados)

                logger.info(f'Criando usuário: {usuario_schema.nome} -- {usuario_schema.email}')
                self.cria_usuario_perfil_escola(ind, usuario_schema)
            except Exception as exc:
                self.erros.append(f'Linha {ind} - {exc}')

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def validacao_inicial(self) -> bool:
        return self.existe_conteudo()

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = 'Não foi feito o upload da planilha'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {key: linha[index].value
                for index, key in enumerate(ImportacaoPlanilhaUsuarioPerfilEscolaSchema.schema()['properties'].keys())}

    def cria_usuario_perfil_escola(self, ind, usuario_schema: ImportacaoPlanilhaUsuarioPerfilEscolaSchema):  # noqa C901
        try:
            self.__criar_usuario_perfil_escola(usuario_schema)
        except Exception as exd:
            self.erros.append(f'Linha {ind} - {exd}')

    @transaction.atomic
    def __criar_usuario_perfil_escola(self, usuario_schema: ImportacaoPlanilhaUsuarioPerfilEscolaSchema):
        escola = self.consulta_escola(usuario_schema.codigo_eol_escola)
        self.checa_usuario(usuario_schema.rf, usuario_schema.email, usuario_schema.nome)
        self.__criar_usuario(escola, usuario_schema)

    def consulta_escola(self, codigo_eol_escola):
        escola = Escola.objects.filter(codigo_eol=codigo_eol_escola).first()
        if not escola:
            raise Exception(f'Escola não encontrada para o código: {codigo_eol_escola}.')
        return escola

    def checa_usuario(self, registro_funcional, email, nome_planilha):
        usuario = Usuario.objects.filter(Q(registro_funcional=registro_funcional) | Q(email=email)).first()
        if usuario:
            logger.error(f'Usuário com email "{email}" ou com rf "{registro_funcional}" já existe.')
            raise Exception(f'Usuário com email "{email}" ou com rf '
                            + f'"{registro_funcional}" já existe. Nome na planilha: {nome_planilha}.')

    def __criar_usuario(self, escola, usuario_schema):
        try:
            perfil = Perfil.objects.get(nome=usuario_schema.perfil)
        except Perfil.DoesNotExist:
            logger.error(f'Este perfil não existe: {usuario_schema.perfil}')
            raise Exception(f'Este perfil não existe: {usuario_schema.perfil}')
        data_atual = date.today()
        usuario = Usuario.objects.create_user(
            email=usuario_schema.email,
            password=DJANGO_ADMIN_TREINAMENTO_PASSWORD,
            nome=usuario_schema.nome,
            cargo=usuario_schema.cargo,
            registro_funcional=usuario_schema.rf,
            cpf=usuario_schema.cpf,
        )
        Vinculo.objects.create(
            instituicao=escola,
            perfil=perfil,
            usuario=usuario,
            data_inicial=data_atual,
            ativo=True,
        )
        contato = Contato.objects.create(
            nome=usuario_schema.nome,
            email=usuario_schema.email,
            telefone=usuario_schema.telefone
        )
        usuario.contatos.add(contato)

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = '\n'.join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
            logger.error(f'Arquivo "{self.arquivo.uuid}" processado com erro(s).')
        else:
            self.arquivo.log = 'Planilha processada com sucesso.'
            self.arquivo.processamento_com_sucesso()
            logger.info(f'Arquivo "{self.arquivo.uuid}" processado com sucesso.')

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = 'Erros'
        cabecalho = ws.cell(row=1, column=1, value='Erros encontrados no processamento da planilha')
        cabecalho.fill = styles.PatternFill('solid', fgColor='808080')
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f'arquivo_resultado_{self.arquivo.pk}.xlsx'
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))


def importa_usuarios_perfil_escola(usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioPerfilEscola) -> None:
    logger.debug(f'Iniciando o processamento do arquivo: {arquivo.uuid}')

    try:
        processador = ProcessadorPlanilhaUsuarioPerfilEscola(usuario, arquivo)
        processador.processamento()
        processador.finaliza_processamento()
    except Exception as exc:
        logger.error(f'Erro genérico: {exc}')


class ProcessadorPlanilhaUsuarioPerfilCodae:
    def __init__(self, usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioPerfilCodae) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.erros = []
        self.worksheet = self.abre_worksheet()

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def processamento(self):  # noqa C901
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        logger.info(f'Quantidade de linhas: {len(linhas)} -- Quantidade de colunas: {len(linhas[0])}')

        for ind, linha in enumerate(linhas[1:], 2):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)
                usuario_schema = ImportacaoPlanilhaUsuarioPerfilCodaeSchema(**dicionario_dados)

                logger.info(f'Criando usuário: {usuario_schema.nome} -- {usuario_schema.email}')
                self.cria_usuario_perfil_codae(ind, usuario_schema)
            except Exception as exc:
                self.erros.append(f'Linha {ind} - {exc}')

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def validacao_inicial(self) -> bool:
        return self.existe_conteudo()

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = 'Não foi feito o upload da planilha'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {key: linha[index].value
                for index, key in enumerate(ImportacaoPlanilhaUsuarioPerfilCodaeSchema.schema()['properties'].keys())}

    def cria_usuario_perfil_codae(self, ind, usuario_schema: ImportacaoPlanilhaUsuarioPerfilCodaeSchema):  # noqa C901
        try:
            self.__criar_usuario_perfil_codae(usuario_schema)
        except Exception as exd:
            self.erros.append(f'Linha {ind} - {exd}')

    @transaction.atomic
    def __criar_usuario_perfil_codae(self, usuario_schema: ImportacaoPlanilhaUsuarioPerfilCodaeSchema):
        codae = self.consulta_codae_instituicao(usuario_schema)
        self.checa_usuario(usuario_schema.rf, usuario_schema.email, usuario_schema.nome)
        self.__criar_usuario(codae, usuario_schema)

    def consulta_codae_instituicao(self, usuario_schema):
        perfis = {
            'perfil_usuario_codae': Perfil.objects.get(nome__unaccent__icontains='COORDENADOR_GESTAO_ALIMENTACAO_TERCEIRIZADA').nome,
            'perfil_usuario_nutri_codae': Perfil.objects.get(nome__unaccent__icontains='COORDENADOR_DIETA_ESPECIAL').nome,
            'perfil_usuario_nutri_supervisao': Perfil.objects.get(nome__unaccent__icontains='COORDENADOR_SUPERVISAO_NUTRICAO').nome,
            'perfil_usuario_nutri_manifestacao': Perfil.objects.get(nome__unaccent__icontains='COORDENADOR_SUPERVISAO_NUTRICAO_MANIFESTACAO').nome,
            'perfil_coordenador_gestao_produto': Perfil.objects.get(nome__unaccent__icontains='COORDENADOR_GESTAO_PRODUTO').nome
        }
        instituicoes_codae = {
            'codae_alimentacao': Codae.objects.get(nome__unaccent__icontains='ALIMENTAÇÃO'),
            'codae_dieta_especial': Codae.objects.get(nome__unaccent__icontains='ESPECIAL'),
            'codae_nutrisupervisao': Codae.objects.get(nome__unaccent__icontains='NUTRIÇÃO'),
            'codae_nutrimanifestacao': Codae.objects.get(nome__unaccent__icontains='NUTRIMANIFESTAÇÃO'),
            'codae_produtos': Codae.objects.get(nome__unaccent__icontains='PRODUTOS')
        }

        try:
            perfil = Perfil.objects.get(nome=usuario_schema.perfil)
        except Perfil.DoesNotExist:
            logger.error(f'Este perfil não existe: {usuario_schema.perfil}')
            raise Exception(f'Este perfil não existe: {usuario_schema.perfil}')
        if usuario_schema.perfil == perfis['perfil_usuario_codae']:
            codae = instituicoes_codae['codae_alimentacao']
        elif usuario_schema.perfil == perfis['perfil_usuario_nutri_codae']:
            codae = instituicoes_codae['codae_dieta_especial']
        elif usuario_schema.perfil == perfis['perfil_usuario_nutri_supervisao']:
            codae = instituicoes_codae['codae_nutrisupervisao']
        elif usuario_schema.perfil == perfis['perfil_usuario_nutri_manifestacao']:
            codae = instituicoes_codae['codae_nutrimanifestacao']
        elif usuario_schema.perfil == perfis['perfil_coordenador_gestao_produto']:
            codae = instituicoes_codae['codae_produtos']
        else:
            codae = None

        if not codae:
            raise Exception(f'Instituição CODAE não encontrada para perfil: {usuario_schema.perfil}.')
        return codae

    def checa_usuario(self, registro_funcional, email, nome_planilha):
        usuario = Usuario.objects.filter(Q(registro_funcional=registro_funcional) | Q(email=email)).first()
        if usuario:
            logger.error(f'Usuário com email "{email}" ou com rf "{registro_funcional}" já existe.')
            raise Exception(f'Usuário com email "{email}" ou com rf '
                            + f'"{registro_funcional}" já existe. Nome na planilha: {nome_planilha}.')

    def __criar_usuario(self, codae, usuario_schema):
        try:
            perfil = Perfil.objects.get(nome=usuario_schema.perfil)
        except Perfil.DoesNotExist:
            logger.error(f'Este perfil não existe: {usuario_schema.perfil}')
            raise Exception(f'Este perfil não existe: {usuario_schema.perfil}')
        data_atual = date.today()
        usuario = Usuario.objects.create_user(
            email=usuario_schema.email,
            password=DJANGO_ADMIN_TREINAMENTO_PASSWORD,
            nome=usuario_schema.nome,
            cargo=usuario_schema.cargo,
            registro_funcional=usuario_schema.rf,
            cpf=usuario_schema.cpf,
            crn_numero=usuario_schema.crn_numero
        )
        Vinculo.objects.create(
            instituicao=codae,
            perfil=perfil,
            usuario=usuario,
            data_inicial=data_atual,
            ativo=True,
        )
        contato = Contato.objects.create(
            nome=usuario_schema.nome,
            email=usuario_schema.email,
            telefone=usuario_schema.telefone,
        )
        usuario.contatos.add(contato)

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = '\n'.join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
            logger.error(f'Arquivo "{self.arquivo.uuid}" processado com erro(s).')
        else:
            self.arquivo.log = 'Planilha processada com sucesso.'
            self.arquivo.processamento_com_sucesso()
            logger.info(f'Arquivo "{self.arquivo.uuid}" processado com sucesso.')

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = 'Erros'
        cabecalho = ws.cell(row=1, column=1, value='Erros encontrados no processamento da planilha')
        cabecalho.fill = styles.PatternFill('solid', fgColor='808080')
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f'arquivo_resultado_{self.arquivo.pk}.xlsx'
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))


def importa_usuarios_perfil_codae(usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioPerfilCodae) -> None:
    logger.debug(f'Iniciando o processamento do arquivo: {arquivo.uuid}')

    try:
        processador = ProcessadorPlanilhaUsuarioPerfilCodae(usuario, arquivo)
        processador.processamento()
        processador.finaliza_processamento()
    except Exception as exc:
        logger.error(f'Erro genérico: {exc}')


class ProcessadorPlanilhaUsuarioPerfilDre:
    def __init__(self, usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioPerfilDre) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.erros = []
        self.worksheet = self.abre_worksheet()

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def processamento(self):  # noqa C901
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        logger.info(f'Quantidade de linhas: {len(linhas)} -- Quantidade de colunas: {len(linhas[0])}')

        for ind, linha in enumerate(linhas[1:], 2):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)
                usuario_schema = ImportacaoPlanilhaUsuarioPerfilDreSchema(**dicionario_dados)

                logger.info(f'Criando usuário: {usuario_schema.nome} -- {usuario_schema.email}')
                self.cria_usuario_perfil_codae(ind, usuario_schema)
            except Exception as exc:
                self.erros.append(f'Linha {ind} - {exc}')

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def validacao_inicial(self) -> bool:
        return self.existe_conteudo()

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = 'Não foi feito o upload da planilha'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {key: linha[index].value
                for index, key in enumerate(ImportacaoPlanilhaUsuarioPerfilDreSchema.schema()['properties'].keys())}

    def cria_usuario_perfil_codae(self, ind, usuario_schema: ImportacaoPlanilhaUsuarioPerfilDreSchema):  # noqa C901
        try:
            self.__criar_usuario_perfil_codae(usuario_schema)
        except Exception as exd:
            self.erros.append(f'Linha {ind} - {exd}')

    @transaction.atomic
    def __criar_usuario_perfil_codae(self, usuario_schema: ImportacaoPlanilhaUsuarioPerfilDreSchema):
        dre = self.consulta_dre(usuario_schema.codigo_eol_dre)
        self.checa_usuario(usuario_schema.rf, usuario_schema.email, usuario_schema.nome)
        self.__criar_usuario(dre, usuario_schema)

    def consulta_dre(self, codigo_eol_dre):
        dre = DiretoriaRegional.objects.filter(codigo_eol=codigo_eol_dre).first()
        if not dre:
            raise Exception(f'Dre não encontrada para o código: {codigo_eol_dre}.')
        return dre

    def checa_usuario(self, registro_funcional, email, nome_planilha):
        usuario = Usuario.objects.filter(Q(registro_funcional=registro_funcional) | Q(email=email)).first()
        if usuario:
            logger.error(f'Usuário com email "{email}" ou com rf "{registro_funcional}" já existe.')
            raise Exception(f'Usuário com email "{email}" ou com rf '
                            + f'"{registro_funcional}" já existe. Nome na planilha: {nome_planilha}.')

    def __criar_usuario(self, dre, usuario_schema):
        try:
            perfil = Perfil.objects.get(nome=usuario_schema.perfil)
        except Perfil.DoesNotExist:
            logger.error(f'Este perfil não existe: {usuario_schema.perfil}')
            raise Exception(f'Este perfil não existe: {usuario_schema.perfil}')
        data_atual = date.today()
        usuario = Usuario.objects.create_user(
            email=usuario_schema.email,
            password=DJANGO_ADMIN_TREINAMENTO_PASSWORD,
            nome=usuario_schema.nome,
            cargo=usuario_schema.cargo,
            registro_funcional=usuario_schema.rf,
            cpf=usuario_schema.cpf,
        )
        Vinculo.objects.create(
            instituicao=dre,
            perfil=perfil,
            usuario=usuario,
            data_inicial=data_atual,
            ativo=True,
        )
        contato = Contato.objects.create(
            nome=usuario_schema.nome,
            email=usuario_schema.email,
            telefone=usuario_schema.telefone,
        )
        usuario.contatos.add(contato)

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = '\n'.join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
            logger.error(f'Arquivo "{self.arquivo.uuid}" processado com erro(s).')
        else:
            self.arquivo.log = 'Planilha processada com sucesso.'
            self.arquivo.processamento_com_sucesso()
            logger.info(f'Arquivo "{self.arquivo.uuid}" processado com sucesso.')

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = 'Erros'
        cabecalho = ws.cell(row=1, column=1, value='Erros encontrados no processamento da planilha')
        cabecalho.fill = styles.PatternFill('solid', fgColor='808080')
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f'arquivo_resultado_{self.arquivo.pk}.xlsx'
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))


def importa_usuarios_perfil_dre(usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioPerfilDre) -> None:
    logger.debug(f'Iniciando o processamento do arquivo: {arquivo.uuid}')

    try:
        processador = ProcessadorPlanilhaUsuarioPerfilDre(usuario, arquivo)
        processador.processamento()
        processador.finaliza_processamento()
    except Exception as exc:
        logger.error(f'Erro genérico: {exc}')


class ProcessaPlanilhaUsuarioServidorCoreSSO:
    def __init__(self, usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioServidorCoreSSO) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.erros = []
        self.worksheet = self.abre_worksheet()

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def processamento(self):  # noqa C901
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        logger.info(f'Quantidade de linhas: {len(linhas)} -- Quantidade de colunas: {len(linhas[0])}')

        for ind, linha in enumerate(linhas[1:], 2):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)
                usuario_schema = ImportacaoPlanilhaUsuarioServidorCoreSSOSchema(**dicionario_dados)
                logger.info(f'Criando usuário: {usuario_schema.nome} -- {usuario_schema.email}')
                self.cria_usuario_servidor(ind, usuario_schema)
                self.loga_sucesso_carga_usuario(usuario_schema)

            except Exception as exc:
                self.erros.append(f'Linha {ind} - {exc}')

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def get_instituicao(self, dados_usuario):
        if dados_usuario.tipo_perfil == 'ESCOLA':
            return Escola.objects.get(codigo_eol=format(int(dados_usuario.codigo_eol), '06d'))
        elif dados_usuario.tipo_perfil == 'DRE':
            return DiretoriaRegional.objects.get(codigo_eol=format(int(dados_usuario.codigo_eol), '06d'))
        else:
            return Codae.objects.annotate(nome_sem_espacos=Func(F('nome'), Value(' '), Value(''), function='replace')
                                          ).get(nome_sem_espacos__icontains=f'codae-{dados_usuario.codae}')

    def get_perfil(self, dados_usuario):
        return Perfil.objects.get(nome__iexact=dados_usuario.perfil)

    def loga_sucesso_carga_usuario(self, dados_usuario):
        mensagem = f'Usuário {dados_usuario.rf} criado/atualizado com sucesso.'
        logger.info(mensagem)

    def cria_ou_atualiza_usuario_admin(self, dados_usuario):
        usuario, criado = Usuario.objects.update_or_create(
            username=dados_usuario.rf,
            registro_funcional=dados_usuario.rf,
            defaults={
                'email': dados_usuario.email if dados_usuario.email else "",
                'nome': dados_usuario.nome,
                'cargo': dados_usuario.cargo or "",
                'cpf': dados_usuario.cpf or "",
            }
        )
        return usuario

    def cria_vinculo(self, usuario, dados_usuario):
        if usuario.existe_vinculo_ativo:
            vinculo = usuario.vinculo_atual
            vinculo.ativo = False
            vinculo.data_final = date.today()
            vinculo.save()
        Vinculo.objects.create(
            instituicao=self.get_instituicao(dados_usuario),
            perfil=self.get_perfil(dados_usuario),
            usuario=usuario,
            data_inicial=date.today(),
            ativo=True,
        )

    def validacao_inicial(self) -> bool:
        return self.existe_conteudo()

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = 'Não foi feito o upload da planilha'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {key: linha[index].value
                for index, key in enumerate(ImportacaoPlanilhaUsuarioServidorCoreSSOSchema.schema()['properties'].keys())}

    def cria_usuario_servidor(self, ind, usuario_schema: ImportacaoPlanilhaUsuarioServidorCoreSSOSchema):  # noqa C901
        try:
            self.__criar_usuario_servidor(usuario_schema)
        except Exception as exd:
            self.erros.append(f'Linha {ind} - {exd}')

    @transaction.atomic
    def __criar_usuario_servidor(self, usuario_schema: ImportacaoPlanilhaUsuarioServidorCoreSSOSchema):
        usuario = self.cria_ou_atualiza_usuario_admin(usuario_schema)
        self.cria_vinculo(usuario, usuario_schema)
        eolusuariocoresso = EOLUsuarioCoreSSO()
        eolusuariocoresso.cria_ou_atualiza_usuario_core_sso(usuario_schema, login=usuario_schema.rf, eh_servidor='S')

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = '\n'.join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
            logger.error(f'Arquivo "{self.arquivo.uuid}" processado com erro(s).')
        else:
            self.arquivo.log = 'Planilha processada com sucesso.'
            self.arquivo.processamento_com_sucesso()
            logger.info(f'Arquivo "{self.arquivo.uuid}" processado com sucesso.')

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = 'Erros'
        cabecalho = ws.cell(row=1, column=1, value='Erros encontrados no processamento da planilha')
        cabecalho.fill = styles.PatternFill('solid', fgColor='808080')
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f'arquivo_resultado_{self.arquivo.pk}.xlsx'
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))


class ProcessaPlanilhaUsuarioExternoCoreSSO:
    def __init__(self, usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioServidorCoreSSO) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.erros = []
        self.worksheet = self.abre_worksheet()

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def processamento(self):  # noqa C901
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        logger.info(f'Quantidade de linhas: {len(linhas)} -- Quantidade de colunas: {len(linhas[0])}')

        for ind, linha in enumerate(linhas[1:], 2):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)
                usuario_schema = ImportacaoPlanilhaUsuarioExternoCoreSSOSchema(**dicionario_dados)
                logger.info(f'Criando usuário: {usuario_schema.nome} -- {usuario_schema.email}')
                self.cria_usuario_externo(ind, usuario_schema)
                self.loga_sucesso_carga_usuario(usuario_schema)

            except Exception as exc:
                self.erros.append(f'Linha {ind} - {exc}')

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def get_instituicao(self, dados_usuario):
        return Terceirizada.objects.get(cnpj=dados_usuario.cnpj_terceirizada)

    def get_perfil(self, dados_usuario):
        return Perfil.objects.get(nome__iexact=dados_usuario.perfil)

    def loga_sucesso_carga_usuario(self, dados_usuario):
        mensagem = f'Usuário {dados_usuario.cpf} criado/atualizado com sucesso.'
        logger.info(mensagem)

    def cria_ou_atualiza_usuario_admin(self, dados_usuario):
        usuario, criado = Usuario.objects.update_or_create(
            username=dados_usuario.cpf,
            defaults={
                'email': dados_usuario.email if dados_usuario.email else "",
                'nome': dados_usuario.nome,
                'cpf': dados_usuario.cpf
            }
        )
        return usuario

    def cria_vinculo(self, usuario, dados_usuario):
        if usuario.existe_vinculo_ativo:
            vinculo = usuario.vinculo_atual
            vinculo.ativo = False
            vinculo.data_final = date.today()
            vinculo.save()
        Vinculo.objects.create(
            instituicao=self.get_instituicao(dados_usuario),
            perfil=self.get_perfil(dados_usuario),
            usuario=usuario,
            data_inicial=date.today(),
            ativo=True,
        )

    def validacao_inicial(self) -> bool:
        return self.existe_conteudo()

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = 'Não foi feito o upload da planilha'
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {key: linha[index].value
                for index, key in enumerate(ImportacaoPlanilhaUsuarioExternoCoreSSOSchema.schema()['properties'].keys())}

    def cria_usuario_externo(self, ind, usuario_schema: ImportacaoPlanilhaUsuarioExternoCoreSSOSchema):  # noqa C901
        try:
            self.__criar_usuario_externo(usuario_schema)
        except Exception as exd:
            self.erros.append(f'Linha {ind} - {exd}')

    @transaction.atomic
    def __criar_usuario_externo(self, usuario_schema: ImportacaoPlanilhaUsuarioExternoCoreSSOSchema):
        usuario = self.cria_ou_atualiza_usuario_admin(usuario_schema)
        self.cria_vinculo(usuario, usuario_schema)
        eolusuariocoresso = EOLUsuarioCoreSSO()
        eolusuariocoresso.cria_ou_atualiza_usuario_core_sso(usuario_schema, login=usuario_schema.cpf, eh_servidor='N')

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = '\n'.join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
            logger.error(f'Arquivo "{self.arquivo.uuid}" processado com erro(s).')
        else:
            self.arquivo.log = 'Planilha processada com sucesso.'
            self.arquivo.processamento_com_sucesso()
            logger.info(f'Arquivo "{self.arquivo.uuid}" processado com sucesso.')

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = 'Erros'
        cabecalho = ws.cell(row=1, column=1, value='Erros encontrados no processamento da planilha')
        cabecalho.fill = styles.PatternFill('solid', fgColor='808080')
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f'arquivo_resultado_{self.arquivo.pk}.xlsx'
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))


def importa_usuarios_servidores_coresso(usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioServidorCoreSSO) -> None:
    logger.debug(f'Iniciando o processamento do arquivo: {arquivo.uuid}')

    try:
        processador = ProcessaPlanilhaUsuarioServidorCoreSSO(usuario, arquivo)
        processador.processamento()
        processador.finaliza_processamento()
    except Exception as exc:
        logger.error(f'Erro genérico: {exc}')


class ProcessaPlanilhaUsuarioServidorCoreSSOException(Exception):
    pass


def importa_usuarios_externos_coresso(usuario: Usuario, arquivo: ImportacaoPlanilhaUsuarioExternoCoreSSO) -> None:
    logger.debug(f'Iniciando o processamento do arquivo: {arquivo.uuid}')

    try:
        processador = ProcessaPlanilhaUsuarioExternoCoreSSO(usuario, arquivo)
        processador.processamento()
        processador.finaliza_processamento()
    except Exception as exc:
        logger.error(f'Erro genérico: {exc}')


class ProcessaPlanilhaUsuarioExternoCoreSSOException(Exception):
    pass
