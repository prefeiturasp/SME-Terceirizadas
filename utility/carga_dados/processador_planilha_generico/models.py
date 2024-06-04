import logging
from tempfile import NamedTemporaryFile
from typing import Type, TypeVar

from django.core.files import File
from openpyxl import Workbook, load_workbook, styles

from sme_terceirizadas.perfil.models import Usuario

T = TypeVar("T")


class ProcessadorDePlanilha:
    LINHA_2 = 2

    def __init__(
        self,
        usuario: Usuario,
        arquivo: T,
        class_schema: Type[T],
        atributo_1: str,
        atributo_2: str,
        logger: logging.getLogger,
    ) -> None:
        """Prepara atributos importantes para o processamento da planilha."""
        self.usuario = usuario
        self.arquivo = arquivo
        self.class_schema = class_schema
        self.erros = []
        self.worksheet = self.abre_worksheet()
        self.atributo_1 = atributo_1
        self.atributo_2 = atributo_2
        self.logger = logger

    @property
    def path(self):
        return self.arquivo.conteudo.path

    def cria_objeto(self, idx: int, schema: Type[T]):
        raise NotImplementedError("Deve criar um método cria_objeto")

    def abre_worksheet(self):
        return load_workbook(self.path).active

    def processamento(self):
        self.arquivo.inicia_processamento()
        if not self.validacao_inicial():
            return

        linhas = list(self.worksheet.rows)
        self.logger.info(
            f"Quantidade de linhas: {len(linhas)} -- Quantidade de colunas: {len(linhas[0])}"
        )

        for ind, linha in enumerate(
            linhas[1:], self.LINHA_2
        ):  # Começando em 2 pois a primeira linha é o cabeçalho da planilha
            try:
                dicionario_dados = self.monta_dicionario_de_dados(linha)
                schema = self.class_schema(**dicionario_dados)

                self.logger.info(
                    f"Criando objeto: {getattr(schema, self.atributo_1)} -- "
                    f"{getattr(schema, self.atributo_2)}"
                )
                self.cria_objeto(ind, schema)
            except Exception as exc:
                self.erros.append(f"Linha {ind} - {exc}")

    def validacao_inicial(self) -> bool:
        return self.existe_conteudo()

    def existe_conteudo(self) -> bool:
        if not self.arquivo.conteudo:
            self.arquivo.log = "Não foi feito o upload da planilha"
            self.arquivo.erro_no_processamento()
            return False
        return True

    def monta_dicionario_de_dados(self, linha: tuple) -> dict:
        return {
            key: linha[index].value
            for index, key in enumerate(self.class_schema.schema()["properties"].keys())
        }

    def finaliza_processamento(self) -> None:
        if self.erros:
            self.arquivo.log = "\n".join(self.erros)
            self.arquivo.processamento_com_erro()
            self.cria_planilha_de_erros()
            self.logger.error(f'Arquivo "{self.arquivo.uuid}" processado com erro(s).')
        else:
            self.arquivo.log = "Planilha processada com sucesso."
            self.arquivo.processamento_com_sucesso()
            self.logger.info(f'Arquivo "{self.arquivo.uuid}" processado com sucesso.')

    def cria_planilha_de_erros(self) -> None:
        workbook: Workbook = Workbook()
        ws = workbook.active
        ws.title = "Erros"
        cabecalho = ws.cell(
            row=1, column=1, value="Erros encontrados no processamento da planilha"
        )
        cabecalho.fill = styles.PatternFill("solid", fgColor="808080")
        for index, erro in enumerate(self.erros, 2):
            ws.cell(row=index, column=1, value=erro)

        filename = f"arquivo_resultado_{self.arquivo.pk}.xlsx"
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            self.arquivo.resultado.save(name=filename, content=File(tmp))
