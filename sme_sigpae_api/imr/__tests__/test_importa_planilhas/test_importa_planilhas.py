from unittest.mock import Mock

import pytest
from django.contrib.admin import AdminSite
from django.core.files.uploadedfile import SimpleUploadedFile
from django.http import HttpResponse
from openpyxl.reader.excel import load_workbook

from sme_sigpae_api.imr.admin import (
    ImportacaoPlanilhaTipoOcorrenciaAdmin,
    ImportacaoPlanilhaTipoPenalidadeAdmin,
)
from sme_sigpae_api.imr.models import (
    ImportacaoPlanilhaTipoOcorrencia,
    ImportacaoPlanilhaTipoPenalidade,
    TipoOcorrencia,
    TipoPenalidade,
)

pytestmark = pytest.mark.django_db


def test_importa_planilha_tipo_penalidade(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    edital_factory,
    tipo_gravidade_factory,
    importacao_planilha_tipo_penalidade_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    edital = edital_factory.create(numero="EDITAL MODELO IMR")
    gravidade = tipo_gravidade_factory.create(tipo="Leve")

    workbook = load_workbook(
        "sme_terceirizadas/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_penalidade.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file = SimpleUploadedFile(
        name="teste_carga_tipos_penalidades.xlsx",
        content=response.content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    importacao_planilha_tipo_penalidade_factory.create(conteudo=file)

    queryset = ImportacaoPlanilhaTipoPenalidade.objects.all()

    importacao_planilha_tipo_penalidade_admin = ImportacaoPlanilhaTipoPenalidadeAdmin(
        model=ImportacaoPlanilhaTipoPenalidade, admin_site=AdminSite()
    )
    importacao_planilha_tipo_penalidade_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )
    assert TipoPenalidade.objects.count() == 1
    tipo_penalidade = TipoPenalidade.objects.get()
    assert tipo_penalidade.edital == edital
    assert tipo_penalidade.gravidade == gravidade
    assert tipo_penalidade.obrigacoes.count() == 3
    assert tipo_penalidade.descricao == "teste descrição"
    assert tipo_penalidade.status is True


def test_importa_planilha_tipo_penalidade_erro_seleciona_mais_de_um_arquivo(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    edital_factory,
    tipo_gravidade_factory,
    importacao_planilha_tipo_penalidade_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    edital_factory.create(numero="EDITAL MODELO IMR")
    tipo_gravidade_factory.create(tipo="Leve")

    workbook = load_workbook(
        "sme_sigpae_api/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_penalidade.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file = SimpleUploadedFile(
        name="teste_carga_tipos_penalidades.xlsx",
        content=response.content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    importacao_planilha_tipo_penalidade_factory.create(conteudo=file)
    importacao_planilha_tipo_penalidade_factory.create(conteudo=file)

    queryset = ImportacaoPlanilhaTipoPenalidade.objects.all()

    importacao_planilha_tipo_penalidade_admin = ImportacaoPlanilhaTipoPenalidadeAdmin(
        model=ImportacaoPlanilhaTipoPenalidade, admin_site=AdminSite()
    )
    importacao_planilha_tipo_penalidade_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )

    # não criou porque selecionou mais de um arquivo
    assert TipoPenalidade.objects.count() == 0


def test_importa_planilha_tipo_penalidade_erro_arquivo_formato_incorreto(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    edital_factory,
    tipo_gravidade_factory,
    importacao_planilha_tipo_penalidade_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    edital_factory.create(numero="EDITAL MODELO IMR")
    tipo_gravidade_factory.create(tipo="Leve")

    workbook = load_workbook(
        "sme_sigpae_api/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_penalidade.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file_pdf = SimpleUploadedFile(
        name="teste_carga_tipos_penalidades.pdf",
        content=response.content,
        content_type="application/pdf",
    )

    importacao_planilha_tipo_penalidade_factory.create(conteudo=file_pdf)

    queryset = ImportacaoPlanilhaTipoPenalidade.objects.all()

    importacao_planilha_tipo_penalidade_admin = ImportacaoPlanilhaTipoPenalidadeAdmin(
        model=ImportacaoPlanilhaTipoPenalidade, admin_site=AdminSite()
    )
    importacao_planilha_tipo_penalidade_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )

    # não criou porque o formato estava incorreto (PDF)
    assert TipoPenalidade.objects.count() == 0


def test_importa_planilha_tipo_ocorrencia(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    edital_factory,
    categoria_ocorrencia_factory,
    importacao_planilha_tipo_ocorrencia_factory,
    tipo_penalidade_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    edital = edital_factory.create(numero="EDITAL MODELO IMR")
    categoria = categoria_ocorrencia_factory.create(
        perfis=["DIRETOR"], nome="RECEBIMENTO DE ALIMENTOS"
    )
    tipo_penalidade = tipo_penalidade_factory.create(
        numero_clausula="1.1.", edital=edital
    )

    workbook = load_workbook(
        "sme_sigpae_api/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_ocorrencia.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file = SimpleUploadedFile(
        name="teste_carga_tipos_ocorrencia.xlsx",
        content=response.content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    importacao_planilha_tipo_ocorrencia_factory.create(conteudo=file)

    queryset = ImportacaoPlanilhaTipoOcorrencia.objects.all()

    importacao_planilha_tipo_ocorrencia_admin = ImportacaoPlanilhaTipoOcorrenciaAdmin(
        model=ImportacaoPlanilhaTipoOcorrencia, admin_site=AdminSite()
    )
    importacao_planilha_tipo_ocorrencia_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )
    assert TipoOcorrencia.objects.count() == 1
    tipo_ocorrencia = TipoOcorrencia.objects.get()
    assert tipo_ocorrencia.posicao == 1
    assert tipo_ocorrencia.perfis == [TipoOcorrencia.DIRETOR]
    assert tipo_ocorrencia.edital == edital
    assert tipo_ocorrencia.categoria == categoria
    assert tipo_ocorrencia.titulo == "titulo teste"
    assert tipo_ocorrencia.descricao == "descrição teste"
    assert tipo_ocorrencia.penalidade == tipo_penalidade
    assert tipo_ocorrencia.eh_imr is True
    assert tipo_ocorrencia.pontuacao == 1
    assert tipo_ocorrencia.tolerancia == 1
    assert tipo_ocorrencia.status is True
    assert tipo_ocorrencia.aceita_multiplas_respostas is True


def test_importa_planilha_tipo_ocorrencia_erro_seleciona_mais_de_um_arquivo(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    importacao_planilha_tipo_ocorrencia_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    workbook = load_workbook(
        "sme_sigpae_api/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_ocorrencia.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file = SimpleUploadedFile(
        name="teste_carga_tipos_ocorrencia.xlsx",
        content=response.content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    importacao_planilha_tipo_ocorrencia_factory.create(conteudo=file)
    importacao_planilha_tipo_ocorrencia_factory.create(conteudo=file)

    queryset = ImportacaoPlanilhaTipoOcorrencia.objects.all()

    importacao_planilha_tipo_ocorrencia_admin = ImportacaoPlanilhaTipoOcorrenciaAdmin(
        model=ImportacaoPlanilhaTipoOcorrencia, admin_site=AdminSite()
    )
    importacao_planilha_tipo_ocorrencia_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )

    # não criou porque selecionou mais de um arquivo
    assert TipoOcorrencia.objects.count() == 0


def test_importa_planilha_tipo_ocorrencia_erro_arquivo_formato_incorreto(
    client_autenticado_vinculo_coordenador_supervisao_nutricao,
    importacao_planilha_tipo_ocorrencia_factory,
):
    client, usuario = client_autenticado_vinculo_coordenador_supervisao_nutricao

    workbook = load_workbook(
        "sme_sigpae_api/imr/__tests__/test_importa_planilhas/arquivos_teste/"
        "planilha_importacao_tipos_ocorrencia.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="my_data.xlsx"'
    workbook.save(response)

    file_pdf = SimpleUploadedFile(
        name="teste_carga_tipos_ocorrencia.pdf",
        content=response.content,
        content_type="application/pdf",
    )

    importacao_planilha_tipo_ocorrencia_factory.create(conteudo=file_pdf)

    queryset = ImportacaoPlanilhaTipoOcorrencia.objects.all()

    importacao_planilha_tipo_ocorrencia_admin = ImportacaoPlanilhaTipoOcorrenciaAdmin(
        model=ImportacaoPlanilhaTipoOcorrencia, admin_site=AdminSite()
    )
    importacao_planilha_tipo_ocorrencia_admin.processa_planilha(
        request=Mock(user=usuario), queryset=queryset
    )

    # não criou porque o formato estava incorreto (PDF)
    assert TipoOcorrencia.objects.count() == 0
