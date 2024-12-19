import os
import tempfile

import openpyxl
import pytest
from rest_framework import status

pytestmark = pytest.mark.django_db


def test_modelo_excel_tipos_penalidade(client_admin_django):
    response = client_admin_django.get(
        "/admin/imr/importacaoplanilhatipopenalidade/exportar_planilha_importacao_tipos_penalidade",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        follow=True,
    )
    assert response.status_code == status.HTTP_200_OK

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_file.write(response.content)
        temp_file_path = temp_file.name

    assert os.path.exists(temp_file_path), "Falha ao criar o arquivo"

    workbook = openpyxl.load_workbook(temp_file_path)
    sheet = workbook.active

    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert first_row == (
        "Edital",
        "Número da Cláusula/Item",
        "Gravidade",
        "Obrigações (separadas por ;)",
        "Descrição da Cláusula/Item",
        "Status",
    )

    os.remove(temp_file_path)


def test_modelo_excel_tipos_ocorrencia(client_admin_django):
    response = client_admin_django.get(
        "/admin/imr/importacaoplanilhatipoocorrencia/exportar_planilha_importacao_tipos_ocorrencia",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        follow=True,
    )
    assert response.status_code == status.HTTP_200_OK

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_file.write(response.content)
        temp_file_path = temp_file.name

    assert os.path.exists(temp_file_path), "Falha ao criar o arquivo"

    workbook = openpyxl.load_workbook(temp_file_path)
    sheet = workbook.active

    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert first_row == (
        "Posição",
        "Perfis",
        "Edital",
        "Categoria da Ocorrência",
        "Título",
        "Descrição",
        "Penalidade",
        "É IMR?",
        "Pontuação (IMR)",
        "Tolerância",
        "% de Desconto",
        "Status",
        "Aceita múltiplas respostas?",
    )

    os.remove(temp_file_path)
